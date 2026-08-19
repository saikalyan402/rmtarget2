from __future__ import annotations

# PATCH VERSION: v7 dynamic-scenario-copy + larger-amounts + simplified-revenue
 
import inspect
import io
from html import escape
from typing import Any, Dict, List, Optional, Sequence, Tuple
 
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook as openpyxl_load_workbook
 
 
 
# =============================================================================
# 1. APPLICATION CONFIGURATION  (management assumptions only - no business data)
# =============================================================================
 
APP_TITLE = "Sales Performance"
APP_SUBTITLE = "FY27 · Current reality → scenario planning → target achievement"
 
# --- Financial-year timeline -------------------------------------------------
MONTHS_COMPLETED = 3        # April, May, June are complete
MONTHS_REMAINING = 9        # July .. March
MONTHS_JUL_JAN = 7          # July .. January
MONTHS_FEB_MAR = 2          # February, March
 
FUTURE_MONTHS: List[str] = [
    "July", "August", "September", "October",
    "November", "December", "January", "February", "March",
]
MOMENTUM_MONTHS = FUTURE_MONTHS[:MONTHS_JUL_JAN]      # July .. January
LEAKAGE_MONTHS = FUTURE_MONTHS[MONTHS_JUL_JAN:]       # February, March
MONTH_DATES = pd.date_range("2026-07-01", periods=len(FUTURE_MONTHS), freq="MS")
 
# --- Revenue assumptions (basis points) --------------------------------------
REVENUE_BPS: Dict[str, float] = {"Equity": 60.0, "Debt": 20.0, "Liquid": 10.0}
REVENUE_RATE: Dict[str, float] = {k: v / 10000.0 for k, v in REVENUE_BPS.items()}
 
ASSETS: List[str] = ["Equity", "Debt", "Liquid"]
SALES_TYPES: List[str] = ["GS", "NS"]
SALES_LABEL: Dict[str, str] = {"GS": "Gross Sales", "NS": "Net Sales"}
VERTICALS: List[str] = ["Retail", "DHNI", "VRM"]
 
# Revenue is always measured on one basis so Gross and Net are never added.
REVENUE_BASIS = "NS"
 
# --- Scenario assumptions -----------------------------------------------------
S1_RUNRATE_UPLIFT = 0.20            # +20% on current run rate
S2_EQUITY_TARGET = 1.00             # 100% of Equity FY target by January
S2_OVERALL_TARGET = 0.75            # 75% of overall FY target by January
S3_TARGET = 1.00                    # 100% of FY target by January
S3_DEFAULT_DIP = 0.20               # default Feb-Mar run-rate dip
S4_TARGET = 1.20                    # 120% of FY target by March
S5_EQUITY_TARGET = 1.20             # 120% Equity by March
S5_OVERALL_TARGET = 1.00            # 100% overall by March
S6_SEGMENT_TARGETS: Dict[str, float] = {
    "Digital": 1.40,
    "Retail B30": 1.25,
    "Others": 1.15,
}
S7_DEFAULT_JAN_TARGET = 1.00        # January milestone = 100% of FY target
S7_DEFAULT_MAR_TARGET = 1.00        # March outcome = 100% of FY target
S7_DEFAULT_LEAKAGE = 0.20           # Feb-Mar AUM leakage / run-rate pressure
 
SEGMENT_ORDER: List[str] = ["Digital", "Retail B30", "Others"]
 
# --- Scenario navigator definitions -------------------------------------------
# "short" and "thesis" are presentation copy only; every calculation key
# ("kind", milestones, multipliers) is unchanged.
SCENARIOS: Dict[int, Dict[str, str]] = {
    1: {
        "label": "Scenario 1 · +20% Run-Rate Push",
        "name": "+20% Run-Rate Push",
        "short": "Run Rate",
        "kind": "runrate",
        "thesis": "Lift the current pace by a fifth and let the year compound.",
        "explanation": (
            "Increase the current Apr-Jun monthly run rate by 20% from July onward "
            "and measure the resulting March achievement."
        ),
        "milestone": "March 2027 · run rate lifted 20% for the remaining 9 months",
    },
    2: {
        "label": "Scenario 2 · 75% Overall by Jan + 100% Equity",
        "name": "75% Overall by January + 100% Equity",
        "short": "Jan Milestone",
        "kind": "jan_target",
        "thesis": "Finish Equity early, carry three quarters of the book by January.",
        "explanation": (
            "Reach 100% of the Equity FY target and 75% of the overall FY target by January. "
            "The residual requirement is allocated to Debt and Liquid in FY-target proportion."
        ),
        "milestone": "January 2027 · Equity 100% of FY target, portfolio 75% of FY target",
    },
    3: {
        "label": "Scenario 3 · 100% by Jan, Then Feb-Mar Dip",
        "name": "100% by January, then Feb-Mar dip",
        "short": "Jan + Leakage",
        "kind": "jan_target",
        "thesis": "Land the full year by January, then absorb the closing dip.",
        "explanation": (
            "Reach 100% of the FY target by January, followed by a configurable "
            "February-March run-rate decline."
        ),
        "milestone": "January 2027 · 100% of FY target, then a reduced Feb-Mar run rate",
    },
    4: {
        "label": "Scenario 4 · 120% by March",
        "name": "120% by March",
        "short": "March Target",
        "kind": "march_target",
        "thesis": "Hold one pace for nine months and close the year at 120%.",
        "explanation": (
            "Determine the monthly run rate required to finish March at 120% of the FY target."
        ),
        "milestone": "March 2027 · 120% of FY target",
    },
    5: {
        "label": "Scenario 5 · 120% Equity + 100% Overall",
        "name": "120% Equity + 100% Overall by March",
        "short": "Equity + Overall",
        "kind": "march_target",
        "thesis": "Push Equity beyond target while the portfolio still lands at 100%.",
        "explanation": (
            "Reach 120% of the Equity FY target and 100% of the overall FY target by March, "
            "with Debt and Liquid balancing the remaining requirement."
        ),
        "milestone": "March 2027 · Equity 120% of FY target, portfolio 100% of FY target",
    },
    6: {
        "label": "Scenario 6 · Digital 140% + B30 125% + Others 115%",
        "name": "Digital 140% + Retail B30 125% + Others 115%",
        "short": "Segment",
        "kind": "march_target",
        "thesis": "Let the fastest segments carry a heavier share of the year.",
        "explanation": (
            "Model differentiated performance where Digital achieves 140%, Retail B30 achieves "
            "125% and Others achieve 115% of their respective FY targets."
        ),
        "milestone": "March 2027 · differentiated achievement by business segment",
    },
    7: {
        "label": "Scenario 7 · Momentum Build-Up to March 2027",
        "name": "Momentum Build-Up to March 2027",
        "short": "Momentum",
        "kind": "momentum",
        "thesis": "Build momentum now. Create a January buffer. Protect March.",
        "explanation": (
            "Build progressive month-on-month momentum from July 2026 to reach the January 2027 "
            "milestone, create sufficient buffer to absorb Feb-Mar run-rate leakage, and protect the "
            "March 2027 target."
        ),
        "milestone": "January 2027 milestone → Feb-Mar leakage absorbed → March 2027 target held",
    },
    8: {
        "label": "Scenario 8 · Channel Growth & Target Simulator",
        "name": "Channel Growth & Target Simulator",
        "short": "Channel Simulator",
        "kind": "channel_simulator",
        "thesis": "Nine channels, nine dials. Set the pace channel by channel.",
        "explanation": (
            "Independently adjust monthly growth, January 2027 target achievement and March 2027 "
            "target achievement for Digital, VRM, EM, B30, T30, T8, DHNI, Retail and Institutional."
        ),
        "milestone": "January 2027 target → February/March leakage → March 2027 target",
    },
    9: {
        "label": "Scenario 9 · Channel Mix Optimiser",
        "name": "Channel Mix Optimiser",
        "short": "Optimiser",
        "kind": "channel_optimizer",
        "thesis": "Find the minimum growth each channel must deliver. Nothing more.",
        "explanation": (
            "Find the minimum channel growth trajectory required to achieve a selected portfolio March "
            "ambition, while preserving the January milestone and leakage assumption."
        ),
        "milestone": "Portfolio March ambition optimised across nine channels",
    },
    10: {
        "label": "Scenario 10 · Asset × Channel Target Simulator",
        "name": "Asset × Channel Target Simulator",
        "short": "Asset × Channel",
        "kind": "asset_channel_target",
        "thesis": "Edit Retail, DHNI, VRM and Insti inside Equity, Debt and Liquid; Digital stays out of the calculation.",
        "explanation": (
            "Read Current and FY27 Budget directly from FINAL, then set simulation achievement percentages "
            "for Retail, DHNI, VRM and Insti inside each asset class. Projected Number = Simulation % × FY27 Budget. "
            "Digital is excluded, and T2/T6/T30/B30/EM are read from the FINAL market matrix when present."
        ),
        "milestone": "March 2027 · editable Asset × Channel achievement targets",
    },
}
SCENARIO_ORDER: List[int] = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
 
# Scenario 8/9 channel planning universe.  Each channel has independent
# momentum, January 2027 target and March 2027 target controls.
CHANNELS: List[str] = [
    "Digital", "VRM", "EM", "B30", "T30", "T8", "DHNI", "Retail", "Institutional"
]
S8_DEFAULT_GROWTH: Dict[str, float] = {c: 0.05 for c in CHANNELS}
S8_DEFAULT_JAN_TARGET: Dict[str, float] = {c: 1.00 for c in CHANNELS}
S8_DEFAULT_MAR_TARGET: Dict[str, float] = {c: 1.00 for c in CHANNELS}
S8_DEFAULT_LEAKAGE = 0.20
 
# --- Workbook contract --------------------------------------------------------
SHEET_ALIASES: Dict[str, List[str]] = {
    "Retail": ["RM Retail Sales", "RM Retail", "Retail Sales"],
    "DHNI": ["RM DHNI", "DHNI", "RM D-HNI"],
    "VRM": ["VRM", "RM VRM", "VRM Sales"],
}
 
# Presentation/dashboard sheet. It is deliberately NOT included in
# SHEET_ALIASES because it is not an employee-level calculation sheet.
FINAL_SHEET_ALIASES: List[str] = ["FINAL", "Final", "Final Dashboard"]
 
COLUMN_SPEC: Dict[Tuple[str, str], Dict[str, List[str]]] = {
    ("GS", "Equity"): {
        "fy": ["FY 26 TGT EQ", "Equity GS Targets"],
        "ytd_tgt": ["YTD June EQ TGT", "Q1 Equity GS Targets"],
        "ach": ["Equity GS Ach YTD June", "Equity GS Actuals"],
    },
    ("GS", "Debt"): {
        "fy": ["FY 26 TGT DT", "Debt GS Targets"],
        "ytd_tgt": ["YTD June DT TGT", "Q1 Debt GS Targets"],
        "ach": ["Debt GS Ach", "Debt GS Actuals"],
    },
    ("GS", "Liquid"): {
        "fy": ["FY 26 TGT LIQ", "Liquid GS Targets"],
        "ytd_tgt": ["YTD June LIQ TGT", "Q1 Liquid GS Targets"],
        "ach": ["Liquid GS Ach", "Liquid GS Actuals"],
    },
    ("NS", "Equity"): {
        "fy": ["FY 26 TGT EQ NS", "Equity Net Targets"],
        "ytd_tgt": ["YTD June EQ NS TGT", "Q1 Equity Net Targets"],
        "ach": ["Equity NS Ach YTD June", "Equity Net Actuals"],
    },
    ("NS", "Debt"): {
        "fy": ["FY 26 TGT DT NS", "Debt Net Targets"],
        "ytd_tgt": ["YTD June DT NS TGT", "Q1 Debt Net Targets"],
        "ach": ["Debt NS Ach", "Debt Net Actuals"],
    },
    ("NS", "Liquid"): {
        "fy": ["FY 26 TGT LIQ NS", "Liquid Net Targets"],
        "ytd_tgt": ["YTD June LIQ NS TGT", "Q1 Liquid Net Targets"],
        "ach": ["Liquid NS Ach", "Liquid Net Actuals"],
    },
}
 
META_ALIASES: Dict[str, List[str]] = {
    "Employee Name": ["Employee Name", "Emp Name", "Name"],
    "Emp Code": ["Emp Code", "Employee Code"],
    "ADID": ["ADID", "AD ID"],
    "Status": ["Status", "Employee Status"],
    "Type": ["Type", "Employment Type", "Functional Designation"],
    "ZONE": ["ZONE", "Zone"],
    "REGION": ["REGION", "Region"],
    "EM City": ["EM City", "City", "Location"],
    "MKT TYPE": ["MKT TYPE", "Market Type", "Mkt Type"],
}
META_FIELDS: List[str] = list(META_ALIASES.keys())

# Optional employee/location-level AUM target retained for workbook compatibility.
# Scenario 10 no longer uses AUM weighting; its percentages are based on
# Projected Number / FY27 Target.
AUM_TARGET_ALIASES: List[str] = [
    "Target AUM", "AUM Target", "FY27 AUM Target", "FY 27 AUM Target",
    "FY27 Target AUM", "FY 27 TGT AUM", "AUM TGT", "Target_AUM",
]
 
# -----------------------------------------------------------------------------
# SEGMENT CLASSIFICATION CONFIGURATION (Scenario 6)
# Edit this block to change how business segments are identified. The scenario
# calculation engine reads the resulting mapping and never needs to change.
# -----------------------------------------------------------------------------
SEGMENT_RULES: Dict[str, Dict[str, Any]] = {
    "Digital": {
        "search_columns": ["MKT TYPE", "Type", "REGION", "ZONE", "EM City", "Status"],
        "keywords": ["digital", "online", "d2c", "e-com", "ecom", "virtual", "vrm",
                     "inside sales", "web"],
    },
    "Retail B30": {
        "search_columns": ["MKT TYPE", "REGION", "Type"],
        "keywords": ["b30"],
    },
}
FALLBACK_SEGMENT = "Others"
 
 
# =============================================================================
# 2. GENERIC HELPERS
# =============================================================================
 
class WorkbookError(Exception):
    """Raised when the uploaded workbook does not satisfy the data contract."""
 
 
def normalize_column_name(column: Any) -> str:
    """Collapse non-breaking spaces, repeated spaces and stray padding."""
    return " ".join(str(column).replace("\u00a0", " ").strip().split())
 
 
def _norm_key(column: Any) -> str:
    return normalize_column_name(column).casefold()
 
 
def _squash_key(column: Any) -> str:
    return "".join(ch for ch in _norm_key(column) if ch.isalnum())
 
 
def normalize_frame(frame: pd.DataFrame) -> pd.DataFrame:
    """Return a copy of the frame with normalised column labels."""
    out = frame.copy()
    out.columns = [normalize_column_name(c) for c in out.columns]
    return out
 
 
def clean_numeric(series: Optional[pd.Series]) -> pd.Series:
    """Coerce a column to numeric, tolerating text-formatted numbers."""
    if series is None:
        return pd.Series(dtype="float64")
    if series.dtype == object:
        series = (
            series.astype(str)
            .str.replace("\u00a0", " ", regex=False)
            .str.replace(",", "", regex=False)
            .str.replace("\u20b9", "", regex=False)
            .str.replace("%", "", regex=False)
            .str.strip()
            .replace({"": np.nan, "-": np.nan, "nan": np.nan, "NA": np.nan, "N/A": np.nan})
        )
    return pd.to_numeric(series, errors="coerce")
 
 
def as_text(series: pd.Series) -> pd.Series:
    """Coerce any column to clean text, mapping every missing marker to ''."""
    filled = series.where(series.notna(), "")
    return (
        filled.astype(str)
        .str.replace("\u00a0", " ", regex=False)
        .str.strip()
        .replace({"nan": "", "NaN": "", "NaT": "", "None": "", "<NA>": ""})
    )
 
 
def text_column(frame: pd.DataFrame, column: str) -> pd.Series:
    """Safe accessor for a metadata column that may be absent."""
    if column not in frame.columns:
        return pd.Series([""] * len(frame), index=frame.index, dtype=object)
    return as_text(frame[column])
 
 
def safe_div(numerator: Optional[float], denominator: Optional[float]) -> Optional[float]:
    """Division that never raises and never returns inf/nan."""
    try:
        if numerator is None or denominator is None:
            return None
        n = float(numerator)
        d = float(denominator)
        if not np.isfinite(n) or not np.isfinite(d) or abs(d) < 1e-12:
            return None
        result = n / d
        return result if np.isfinite(result) else None
    except (TypeError, ValueError):
        return None
 
 
def _num(value: Any) -> Optional[float]:
    """Normalise a possibly-missing numeric to float or None."""
    if value is None:
        return None
    try:
        f = float(value)
    except (TypeError, ValueError):
        return None
    return f if np.isfinite(f) else None
 
 
def _z(value: Any) -> float:
    """Numeric value with missing treated as zero (for summation)."""
    v = _num(value)
    return 0.0 if v is None else v
 
 
def _ssum(series: pd.Series) -> Optional[float]:
    """Sum that returns None when every entry is missing."""
    total = series.sum(min_count=1)
    return _num(total)
 
 
# --- Display formatting -------------------------------------------------------
 
NA_TEXT = "0"
 
 
def fmt_cr(value: Any, decimals: int = 0) -> str:
    v = _num(value)
    if v is None:
        return NA_TEXT
    return f"\u20b9 {v:,.{decimals}f} Cr"
 
 
def fmt_cr_signed(value: Any, decimals: int = 0) -> str:
    v = _num(value)
    if v is None:
        return NA_TEXT
    return f"\u20b9 {v:+,.{decimals}f} Cr"
 
 
def fmt_pct(value: Any, decimals: int = 1) -> str:
    v = _num(value)
    if v is None:
        return NA_TEXT
    return f"{v * 100:,.{decimals}f}%"
 
 
def fmt_pct_signed(value: Any, decimals: int = 1) -> str:
    v = _num(value)
    if v is None:
        return NA_TEXT
    return f"{v * 100:+,.{decimals}f}%"
 
 
def fmt_pts(value: Any, decimals: int = 1) -> str:
    """Percentage-point delta, e.g. +18.4 pts."""
    v = _num(value)
    if v is None:
        return NA_TEXT
    return f"{v * 100:+,.{decimals}f} pts"
 
 
def fmt_num(value: Any, decimals: int = 0) -> str:
    v = _num(value)
    if v is None:
        return NA_TEXT
    return f"{v:,.{decimals}f}"
 
 
FORMATTERS = {
    "cr": fmt_cr,
    "cr1": lambda v: fmt_cr(v, 1),
    "cr_signed": fmt_cr_signed,
    "cr1_signed": lambda v: fmt_cr_signed(v, 1),
    "pct": fmt_pct,
    "pct_signed": fmt_pct_signed,
    "pts": fmt_pts,
    "num": fmt_num,
    "txt": lambda v: NA_TEXT if v is None or (isinstance(v, float) and not np.isfinite(v)) else str(v),
}
 
# Formats whose values are read as numbers (right aligned in glass tables).
NUMERIC_FORMATS = {"cr", "cr1", "cr_signed", "cr1_signed", "pct", "pct_signed", "pts", "num"}
AMOUNT_FORMATS = {"cr", "cr1", "cr_signed", "cr1_signed"}
SIGNED_FORMATS = {"cr_signed", "cr1_signed", "pct_signed", "pts"}
 
 
def format_table(frame: pd.DataFrame, formats: Dict[str, str]) -> pd.DataFrame:
    """Return a display-ready copy of a numeric frame using the given formats."""
    out = pd.DataFrame(index=frame.index)
    for column in frame.columns:
        kind = formats.get(column, "txt")
        formatter = FORMATTERS.get(kind, FORMATTERS["txt"])
        out[column] = [formatter(v) for v in frame[column]]
    return out
 
 
# =============================================================================
# 3. WORKBOOK LOADING, VALIDATION & CLEANING
# =============================================================================
 
def _build_column_index(frame: pd.DataFrame) -> Dict[str, int]:
    """Map normalised column keys to positional index (first occurrence wins)."""
    index: Dict[str, int] = {}
    for position, column in enumerate(frame.columns):
        for key in (_norm_key(column), _squash_key(column)):
            if key and key not in index:
                index[key] = position
    return index
 
 
def _resolve_column(index: Dict[str, int], aliases: Sequence[str]) -> Optional[int]:
    for alias in aliases:
        for key in (_norm_key(alias), _squash_key(alias)):
            if key in index:
                return index[key]
    return None
 
 
def _expected_header_keys() -> set:
    keys = set()
    for spec in COLUMN_SPEC.values():
        for aliases in spec.values():
            for alias in aliases:
                keys.add(_norm_key(alias))
    for aliases in META_ALIASES.values():
        for alias in aliases:
            keys.add(_norm_key(alias))
    for alias in AUM_TARGET_ALIASES:
        keys.add(_norm_key(alias))
    return keys
 
 
def _detect_header_row(raw: pd.DataFrame, max_scan: int = 15) -> int:
    """Find the row that actually holds the column headers."""
    expected = _expected_header_keys()
    best_row, best_score = 0, -1
    for row in range(min(max_scan, len(raw))):
        values = {_norm_key(v) for v in raw.iloc[row].tolist() if str(v).strip().lower() != "nan"}
        score = len(values & expected)
        if score > best_score:
            best_row, best_score = row, score
    return best_row if best_score >= 4 else 0
 
 
def _match_sheet(available: Sequence[str], aliases: Sequence[str]) -> Optional[str]:
    normalised = {_norm_key(s): s for s in available}
    for alias in aliases:
        key = _norm_key(alias)
        if key in normalised:
            return normalised[key]
    for alias in aliases:
        key = _norm_key(alias)
        for sheet_key, sheet in normalised.items():
            if key in sheet_key:
                return sheet
    return None
 
 
def validate_frame(frame: pd.DataFrame, index: Dict[str, int], sheet_label: str) -> List[str]:
    """Return a list of human-readable descriptions of missing required columns."""
    missing: List[str] = []
    for (sales, asset), spec in COLUMN_SPEC.items():
        for role, aliases in spec.items():
            if _resolve_column(index, aliases) is None:
                role_label = {
                    "fy": "FY target",
                    "ytd_tgt": "YTD June target",
                    "ach": "YTD June achievement",
                }[role]
                missing.append(
                    f"{sheet_label} · {SALES_LABEL[sales]} · {asset} {role_label} "
                    f"(expected column '{aliases[0]}')"
                )
    return missing
 
 
def _extract_records(frame: pd.DataFrame, vertical: str) -> pd.DataFrame:
    """Turn one workbook sheet into a tidy per-employee record frame."""
    index = _build_column_index(frame)
    records = pd.DataFrame(index=frame.index)
    records["Vertical"] = vertical
 
    for field, aliases in META_ALIASES.items():
        position = _resolve_column(index, aliases)
        if position is None:
            records[field] = ""
        else:
            records[field] = as_text(frame.iloc[:, position]).to_numpy()

    aum_position = _resolve_column(index, AUM_TARGET_ALIASES)
    if aum_position is None:
        records["aum_target"] = 0.0
    else:
        records["aum_target"] = clean_numeric(frame.iloc[:, aum_position]).fillna(0.0).to_numpy()
 
    for (sales, asset), spec in COLUMN_SPEC.items():
        for role, aliases in spec.items():
            position = _resolve_column(index, aliases)
            series = frame.iloc[:, position] if position is not None else None
            records[f"{sales}_{asset}_{role}"] = clean_numeric(series).to_numpy()
 
    return records
 
 
def _clean_records(records: pd.DataFrame) -> pd.DataFrame:
    """Drop non-employee rows while preserving legitimate negative values."""
    names = text_column(records, "Employee Name").str.casefold()
    invalid = names.isin({"", "nan", "none", "total", "grand total", "sum", "subtotal"})
    numeric_columns = [c for c in records.columns if c.split("_")[0] in SALES_TYPES]
    empty_rows = records[numeric_columns].isna().all(axis=1)
    cleaned = records.loc[~(invalid | empty_rows)].copy()
    cleaned[numeric_columns] = cleaned[numeric_columns].fillna(0.0)
    if "aum_target" in cleaned.columns:
        cleaned["aum_target"] = pd.to_numeric(cleaned["aum_target"], errors="coerce").fillna(0.0)
    return cleaned.reset_index(drop=True)
 
 
@st.cache_data(show_spinner=False)
def load_workbook(payload: bytes) -> pd.DataFrame:
    """Read, validate and clean the workbook into a single tidy record frame."""
    try:
        excel = pd.ExcelFile(io.BytesIO(payload), engine="openpyxl")
    except Exception as exc:  # pragma: no cover - defensive
        raise WorkbookError(
            "The file could not be opened as an Excel workbook. "
            "Please upload a valid .xlsx file."
        ) from exc
 
    available = list(excel.sheet_names)
    resolved: Dict[str, str] = {}
    for vertical, aliases in SHEET_ALIASES.items():
        sheet = _match_sheet(available, aliases)
        if sheet is not None:
            resolved[vertical] = sheet
 
    missing_sheets = [v for v in SHEET_ALIASES if v not in resolved]
    if missing_sheets:
        wanted = ", ".join(f"'{SHEET_ALIASES[v][0]}'" for v in missing_sheets)
        raise WorkbookError(
            f"The workbook is missing the required calculation sheet(s): {wanted}. "
            "Please upload the standard RM scorecard workbook."
        )
 
    frames: List[pd.DataFrame] = []
    problems: List[str] = []
    for vertical, sheet in resolved.items():
        raw = pd.read_excel(excel, sheet_name=sheet, header=None, nrows=20)
        header_row = _detect_header_row(raw)
        frame = normalize_frame(pd.read_excel(excel, sheet_name=sheet, header=header_row))
        index = _build_column_index(frame)
        problems.extend(validate_frame(frame, index, vertical))
        if not problems:
            frames.append(_extract_records(frame, vertical))
 
    if problems:
        raise WorkbookError(
            "The workbook is missing required columns:\n\n- " + "\n- ".join(problems[:12])
            + ("\n- \u2026" if len(problems) > 12 else "")
        )
 
    records = _clean_records(pd.concat(frames, ignore_index=True))
    if records.empty:
        raise WorkbookError("No employee records were found in the calculation sheets.")
    return records
 
 
# =============================================================================
# 3A. FINAL SHEET - MANAGEMENT SOURCE LAYER
# =============================================================================
 
def _find_final_sheet_name(sheet_names: Sequence[str]) -> Optional[str]:
    """Resolve the sixth workbook sheet named FINAL."""
    return _match_sheet(sheet_names, FINAL_SHEET_ALIASES)
 
 
def _excel_rgb(color: Any) -> Optional[str]:
    """Convert an openpyxl RGB colour to a CSS hex colour when possible."""
    try:
        if color is None or color.type != "rgb" or not color.rgb:
            return None
        raw = str(color.rgb)
        rgb = raw[-6:]
        if len(rgb) == 6 and all(ch in "0123456789abcdefABCDEF" for ch in rgb):
            return f"#{rgb}"
    except Exception:
        return None
    return None
 
 
def _display_excel_value(value: Any, number_format: str = "") -> str:
    """Format Excel values for the FINAL dashboard without exposing formulas."""
    if value is None:
        return ""
 
    if isinstance(value, (pd.Timestamp,)):
        return value.strftime("%d-%b-%Y")
 
    # Date/datetime objects from openpyxl.
    if hasattr(value, "strftime") and not isinstance(value, str):
        try:
            return value.strftime("%d-%b-%Y")
        except Exception:
            pass
 
    if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
        numeric = float(value)
        if not np.isfinite(numeric):
            return ""
 
        fmt = str(number_format or "")
        if "%" in fmt:
            # Excel stores percentages as fractions.
            decimals = 0
            if "." in fmt.split("%")[0]:
                decimals = len(fmt.split("%")[0].split(".")[-1].replace("0", "0"))
            decimals = min(max(decimals, 0), 2)
            return f"{numeric * 100:,.{decimals}f}%"
 
        # Respect accounting-style parentheses when possible.
        negative_parentheses = numeric < 0 and "(" in fmt and ")" in fmt
        abs_value = abs(numeric)
 
        if abs(abs_value - round(abs_value)) < 1e-9:
            rendered = f"{abs_value:,.0f}"
        else:
            rendered = f"{abs_value:,.2f}".rstrip("0").rstrip(".")
 
        if numeric < 0:
            return f"({rendered})" if negative_parentheses else f"-{rendered}"
        return rendered
 
    return str(value)
 
 
@st.cache_data(show_spinner=False)
def load_final_sheet_frame(payload: bytes) -> pd.DataFrame:
    """
    Read the FINAL sheet as raw cells for a fallback / inspection table.
 
    The sheet intentionally has no single header row, so header=None is used.
    """
    excel = pd.ExcelFile(io.BytesIO(payload), engine="openpyxl")
    sheet = _find_final_sheet_name(excel.sheet_names)
    if sheet is None:
        raise WorkbookError(
            "The workbook does not contain the sixth sheet 'FINAL'. "
            "Please upload the workbook that contains Summary, Summary-Achievement, "
            "RM Retail Sales, RM DHNI, VRM and FINAL."
        )
 
    frame = pd.read_excel(excel, sheet_name=sheet, header=None)
    frame = frame.dropna(axis=0, how="all").dropna(axis=1, how="all")
    return frame.reset_index(drop=True)
 
 
@st.cache_data(show_spinner=False)
def build_final_sheet_html(payload: bytes) -> str:
    """
    Render the Excel FINAL sheet into a scrollable dark-glass HTML table.
 
    Merged headings, alignment and the workbook's own fills are preserved,
    using cached formula results (data_only=True).
    """
    workbook = openpyxl_load_workbook(io.BytesIO(payload), data_only=True)
    sheet_name = _find_final_sheet_name(workbook.sheetnames)
    if sheet_name is None:
        raise WorkbookError(
            "The workbook does not contain the sixth sheet 'FINAL'. "
            "Please upload the workbook that contains the FINAL dashboard sheet."
        )
 
    ws = workbook[sheet_name]
 
    # Find the real content boundary from non-empty values rather than Excel's
    # formatted used-range, which can extend thousands of blank rows/columns.
    non_empty = [
        (cell.row, cell.column)
        for row in ws.iter_rows()
        for cell in row
        if cell.value not in (None, "")
    ]
    if not non_empty:
        return "<div class='glass-note'>The FINAL sheet is empty.</div>"
 
    min_row = min(r for r, _ in non_empty)
    max_row = max(r for r, _ in non_empty)
    min_col = min(c for _, c in non_empty)
    max_col = max(c for _, c in non_empty)
 
    # Merged-cell lookup.
    merge_anchor: Dict[Tuple[int, int], Tuple[int, int]] = {}
    merge_covered: set = set()
    for merged in ws.merged_cells.ranges:
        # Only consider merged ranges intersecting the content boundary.
        if (
            merged.max_row < min_row or merged.min_row > max_row
            or merged.max_col < min_col or merged.min_col > max_col
        ):
            continue
        anchor = (merged.min_row, merged.min_col)
        merge_anchor[anchor] = (
            merged.max_row - merged.min_row + 1,
            merged.max_col - merged.min_col + 1,
        )
        for rr in range(merged.min_row, merged.max_row + 1):
            for cc in range(merged.min_col, merged.max_col + 1):
                if (rr, cc) != anchor:
                    merge_covered.add((rr, cc))
 
    html_parts = [
        "<div class='final-sheet-scroll'><table class='final-sheet-table'>"
    ]
 
    for row_idx in range(min_row, max_row + 1):
        row_values = [
            ws.cell(row=row_idx, column=col_idx).value
            for col_idx in range(min_col, max_col + 1)
        ]
 
        # Keep dashboard spacing, but collapse very large blank areas to a
        # single slim spacer row.
        if all(v in (None, "") for v in row_values):
            html_parts.append(
                "<tr><td colspan='{}' class='final-spacer'></td></tr>".format(
                    max_col - min_col + 1
                )
            )
            continue
 
        html_parts.append("<tr>")
        for col_idx in range(min_col, max_col + 1):
            if (row_idx, col_idx) in merge_covered:
                continue
 
            cell = ws.cell(row=row_idx, column=col_idx)
            rowspan, colspan = merge_anchor.get((row_idx, col_idx), (1, 1))
 
            value = _display_excel_value(cell.value, cell.number_format)
            fill_color = _excel_rgb(cell.fill.fgColor)
            font_color = _excel_rgb(cell.font.color)
 
            styles: List[str] = []
            if fill_color and fill_color.lower() not in {"#000000", "#ffffff"}:
                # Workbook fills are light; keep them, but darken the ink so the
                # cell stays readable inside the dark command centre.
                styles.append(f"background:{fill_color}")
                styles.append("color:#0B0D12")
            elif font_color and font_color.lower() in {"#ff0000", "#c00000", "#e60000"}:
                styles.append("color:#FF8A8A")
 
            if cell.font.bold:
                styles.append("font-weight:650")
            if cell.font.italic:
                styles.append("font-style:italic")
 
            horizontal = getattr(cell.alignment, "horizontal", None)
            if horizontal in {"center", "centerContinuous"}:
                styles.append("text-align:center")
            elif horizontal == "right":
                styles.append("text-align:right")
 
            attrs = []
            if rowspan > 1:
                attrs.append(f"rowspan='{rowspan}'")
            if colspan > 1:
                attrs.append(f"colspan='{colspan}'")
 
            html_parts.append(
                f"<td {' '.join(attrs)} style=\"{';'.join(styles)}\">{escape(value)}</td>"
            )
        html_parts.append("</tr>")
 
    html_parts.append("</table></div>")
    return "".join(html_parts)
 
 
FINAL_METRIC_ROWS: List[str] = [
    "Overall", "Equity", "Debt", "Liquid",
    "Retail", "DHNI", "VRM", "Insti", "Digital",
    "Alternatives", "Passives",
]
 
FINAL_ASSET_ROWS: List[str] = ["Equity", "Debt", "Liquid"]
FINAL_CHANNEL_ROWS: List[str] = [
    "Retail", "DHNI", "VRM", "Insti", "Digital", "Alternatives", "Passives",
]
FINAL_DETAIL_CHANNEL_ROWS: List[str] = ["Overall", "Retail", "DHNI", "VRM", "Insti", "Digital"]
FINAL_MARKET_ROWS: List[str] = ["T2", "T6", "T30", "B30", "EM"]
 
 
def _final_key(value: Any) -> str:
    return " ".join(str(value or "").replace("\u00a0", " ").strip().split()).lower()
 
 
def _final_number(value: Any) -> Optional[float]:
    """Convert FINAL-sheet numeric / accounting text into float."""
    if value is None:
        return None
    if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
        numeric = float(value)
        return numeric if np.isfinite(numeric) else None
 
    raw = str(value).strip()
    if not raw or raw.lower() in {"-", "\u2014", "na", "n/a", "none", "nan", "#div/0!"}:
        return None
 
    negative = raw.startswith("(") and raw.endswith(")")
    cleaned = (
        raw.replace(",", "")
        .replace("\u20b9", "")
        .replace("%", "")
        .replace("(", "")
        .replace(")", "")
        .strip()
    )
    try:
        numeric = float(cleaned)
    except ValueError:
        return None
    if negative:
        numeric = -numeric
    return numeric
 
 
def _final_known_label(value: Any) -> Optional[str]:
    key = _final_key(value)
    aliases = {
        "overall": "Overall",
        "equity": "Equity",
        "debt": "Debt",
        "liquid": "Liquid",
        "retail": "Retail",
        "dhni": "DHNI",
        "vrm": "VRM",
        "insti": "Insti",
        "institutional": "Insti",
        "digital": "Digital",
        "alternatives": "Alternatives",
        "alternate": "Alternatives",
        "passives": "Passives",
        "passive": "Passives",
    }
    return aliases.get(key)
 
 
def _scan_final_sheet(ws: Any) -> Tuple[int, int]:
    """Cap the scan to the management-dashboard area, not formatted blank Excel space."""
    return min(max(ws.max_row, 1), 320), min(max(ws.max_column, 1), 180)
 
 
def _find_final_cells(ws: Any, wanted: str) -> List[Tuple[int, int]]:
    wanted_key = _final_key(wanted)
    max_row, max_col = _scan_final_sheet(ws)
    found: List[Tuple[int, int]] = []
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            if _final_key(ws.cell(row=row, column=col).value) == wanted_key:
                found.append((row, col))
    return found
 
 
def _find_header_near(
    ws: Any,
    title_position: Tuple[int, int],
    header_name: str,
    row_window: int = 8,
    col_before: int = 4,
    col_after: int = 24,
) -> Optional[Tuple[int, int]]:
    title_row, title_col = title_position
    max_row, max_col = _scan_final_sheet(ws)
    wanted = _final_key(header_name)
    for row in range(title_row, min(title_row + row_window, max_row) + 1):
        start_col = max(1, title_col - col_before)
        end_col = min(max_col, title_col + col_after)
        for col in range(start_col, end_col + 1):
            if _final_key(ws.cell(row=row, column=col).value) == wanted:
                return row, col
    return None
 
 
def _augment_final_runrate(frame: pd.DataFrame, months_done: int) -> pd.DataFrame:
    """Recreate the run-rate formulas shown in FINAL from Target and YTD."""
    if frame.empty:
        return frame
 
    months = max(int(months_done), 1)
    out = frame.copy()
    out["FY27 Target"] = pd.to_numeric(out["FY27 Target"], errors="coerce")
    out["YTD"] = pd.to_numeric(out["YTD"], errors="coerce")
 
    out["Achievement %"] = np.where(
        (out["FY27 Target"] > 0) & (out["YTD"] >= 0),
        out["YTD"] / out["FY27 Target"],
        np.nan,
    )
    out["Current RR"] = out["YTD"] / months
 
    # Required run rate means the monthly pace needed FROM NOW to close the
    # remaining FY target. With Apr-Jun complete, 9 months remain.
    # IMPORTANT: this is NOT FY27 Target / 12.
    remaining_months = max(12 - months, 1)
    out["Required RR to Target"] = (out["FY27 Target"] - out["YTD"]) / remaining_months
 
    # Annualise the current Apr-Jun run rate to a full 12-month FY.
    out["Estimated FY @ Current RR"] = out["Current RR"] * 12.0
    out["Projected FY %"] = np.where(
        (out["FY27 Target"] > 0) & (out["Estimated FY @ Current RR"] >= 0),
        out["Estimated FY @ Current RR"] / out["FY27 Target"],
        np.nan,
    )
    return out
 
 
def _parse_final_sales_block(ws: Any, title: str, months_done: int) -> pd.DataFrame:
    """Parse the NET SALES / GROSS SALES run-rate block on the FINAL sheet."""
    positions = _find_final_cells(ws, title)
    for position in positions:
        header = _find_header_near(ws, position, "FY27 Target")
        if header is None:
            continue
 
        header_row, target_col = header
 
        # In the FINAL block the row label sits immediately to the left of FY27 Target.
        label_col = max(1, target_col - 1)
        ytd_col = target_col + 1
 
        rows: List[Dict[str, Any]] = []
        seen: set = set()
        max_row, _ = _scan_final_sheet(ws)
 
        for row in range(header_row + 1, min(header_row + 28, max_row) + 1):
            label = _final_known_label(ws.cell(row=row, column=label_col).value)
            if label is None or label in seen:
                continue
 
            target = _final_number(ws.cell(row=row, column=target_col).value)
            ytd = _final_number(ws.cell(row=row, column=ytd_col).value)
 
            # Ignore title/spacer rows accidentally matching a label.
            if target is None and ytd is None:
                continue
 
            rows.append({"Metric": label, "FY27 Target": target, "YTD": ytd})
            seen.add(label)
 
        if rows:
            frame = pd.DataFrame(rows).set_index("Metric")
            order = [label for label in FINAL_METRIC_ROWS if label in frame.index]
            return _augment_final_runrate(
                frame.loc[order].reset_index(), months_done
            ).set_index("Metric")
 
    return pd.DataFrame()
 
 

def _final_sales_matrix_candidates(ws: Any, sales_code: str) -> List[Tuple[int, int, List[int]]]:
    """Find repeated Target/YTD matrix headers such as the detailed GS/NS tables in FINAL."""
    max_row, max_col = _scan_final_sheet(ws)
    wanted = _final_key(sales_code)
    candidates: List[Tuple[int, int, List[int]]] = []
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            if _final_key(ws.cell(row=row, column=col).value) != wanted:
                continue
            target_cols = [
                cc
                for cc in range(col + 1, min(max_col, col + 24) + 1)
                if _final_key(ws.cell(row=row, column=cc).value) == "target"
            ]
            if len(target_cols) >= 4:
                candidates.append((row, col, target_cols[:4]))
    return candidates


def _parse_final_sales_detail_matrix(ws: Any, sales_code: str) -> pd.DataFrame:
    """
    Parse the detailed FINAL matrix:
      Overall / Retail / DHNI / VRM / Insti / Digital
    crossed with:
      Overall / Equity / Debt / Liquid -> Target, YTD, % Achievement.

    Scenario 10 uses this block as its source of truth instead of rebuilding
    current values from the RM calculation sheets.
    """
    max_row, _ = _scan_final_sheet(ws)
    groups = ["Overall", "Equity", "Debt", "Liquid"]

    for header_row, label_col, target_cols in _final_sales_matrix_candidates(ws, sales_code):
        rows: List[Dict[str, Any]] = []
        labels_found: set = set()

        for row in range(header_row + 1, min(header_row + 14, max_row) + 1):
            label = _final_known_label(ws.cell(row=row, column=label_col).value)
            if label not in FINAL_DETAIL_CHANNEL_ROWS:
                continue
            labels_found.add(label)
            for group, target_col in zip(groups, target_cols):
                target = _final_number(ws.cell(row=row, column=target_col).value)
                current = _final_number(ws.cell(row=row, column=target_col + 1).value)
                if target is None and current is None:
                    continue
                rows.append({
                    "Channel": label,
                    "Asset": group,
                    "FY27 Target": 0.0 if target is None else target,
                    "Current": 0.0 if current is None else current,
                })

        # The management matrix must contain at least the main planning channels.
        required = {"Overall", "Retail", "DHNI", "VRM", "Insti"}
        if rows and len(required.intersection(labels_found)) >= 4:
            frame = pd.DataFrame(rows)
            frame["Current Achievement %"] = np.where(
                frame["FY27 Target"] != 0,
                frame["Current"] / frame["FY27 Target"],
                0.0,
            )
            return frame

    return pd.DataFrame()


def _parse_final_market_detail_matrix(ws: Any, sales_code: str) -> pd.DataFrame:
    """
    Parse a FINAL market-type matrix when the workbook contains T2/T6/T30/B30/EM
    under repeated Overall/Equity/Debt/Liquid Target/YTD headers.

    No RM-sheet fallback is used for Scenario 10. If this matrix is not present
    in FINAL, the market-type simulator reports that the FINAL source is absent.
    """
    max_row, _ = _scan_final_sheet(ws)
    groups = ["Overall", "Equity", "Debt", "Liquid"]
    wanted_markets = {_final_key(value): value for value in FINAL_MARKET_ROWS}

    for header_row, label_col, target_cols in _final_sales_matrix_candidates(ws, sales_code):
        rows: List[Dict[str, Any]] = []
        found: set = set()

        # Market rows are expected immediately below their own repeated header.
        # Keep the window tight so a later market table is never accidentally
        # attached to the earlier channel matrix.
        for row in range(header_row + 1, min(header_row + 14, max_row) + 1):
            raw = ws.cell(row=row, column=label_col).value
            market = wanted_markets.get(_final_key(raw))
            if market is None:
                continue
            found.add(market)
            for group, target_col in zip(groups, target_cols):
                target = _final_number(ws.cell(row=row, column=target_col).value)
                current = _final_number(ws.cell(row=row, column=target_col + 1).value)
                if target is None and current is None:
                    continue
                rows.append({
                    "Market Type": market,
                    "Asset": group,
                    "FY27 Target": 0.0 if target is None else target,
                    "Current": 0.0 if current is None else current,
                })

        if rows and len(found) >= 2:
            frame = pd.DataFrame(rows)
            frame["Current Achievement %"] = np.where(
                frame["FY27 Target"] != 0,
                frame["Current"] / frame["FY27 Target"],
                0.0,
            )
            return frame

    return pd.DataFrame()


def _parse_final_aum_block(ws: Any) -> pd.DataFrame:
    """Parse Target / Current AUM from the FINAL management sheet."""
    positions = _find_final_cells(ws, "AUM")
    for position in positions:
        title_row, title_col = position
        max_row, max_col = _scan_final_sheet(ws)
 
        header_row = None
        target_col = None
        current_col = None
 
        for row in range(title_row, min(title_row + 6, max_row) + 1):
            for col in range(max(1, title_col - 4), min(max_col, title_col + 8) + 1):
                if _final_key(ws.cell(row=row, column=col).value) == "target":
                    # Find Current on the same header row.
                    for cc in range(col + 1, min(max_col, col + 5) + 1):
                        if _final_key(ws.cell(row=row, column=cc).value) == "current":
                            header_row = row
                            target_col = col
                            current_col = cc
                            break
                if header_row is not None:
                    break
            if header_row is not None:
                break
 
        if header_row is None or target_col is None or current_col is None:
            continue
 
        label_col = max(1, target_col - 1)
        rows: List[Dict[str, Any]] = []
        seen: set = set()
 
        for row in range(header_row + 1, min(header_row + 28, max_row) + 1):
            label = _final_known_label(ws.cell(row=row, column=label_col).value)
            if label is None or label in seen:
                continue
            target = _final_number(ws.cell(row=row, column=target_col).value)
            current = _final_number(ws.cell(row=row, column=current_col).value)
            if target is None and current is None:
                continue
            rows.append({"Metric": label, "Target": target, "Current": current})
            seen.add(label)
 
        if rows:
            frame = pd.DataFrame(rows).set_index("Metric")
            order = [label for label in FINAL_METRIC_ROWS if label in frame.index]
            frame = frame.loc[order].copy()
            frame["Achievement %"] = np.where(
                (frame["Target"] > 0) & (frame["Current"] >= 0),
                frame["Current"] / frame["Target"],
                np.nan,
            )
            frame["Gap to Target"] = frame["Target"] - frame["Current"]
            return frame
 
    return pd.DataFrame()
 
 
def _parse_months_done(ws: Any) -> int:
    positions = (
        _find_final_cells(ws, "#months done")
        + _find_final_cells(ws, "months done")
        + _find_final_cells(ws, "# months done")
    )
    max_row, max_col = _scan_final_sheet(ws)
 
    for row, col in positions:
        # Search immediately around / below the label for the highlighted count.
        for rr in range(row, min(row + 4, max_row) + 1):
            for cc in range(max(1, col - 2), min(max_col, col + 4) + 1):
                value = _final_number(ws.cell(row=rr, column=cc).value)
                if value is not None and 1 <= value <= 12:
                    return int(round(value))
    return MONTHS_COMPLETED
 
 
@st.cache_data(show_spinner=False)
def parse_final_dashboard_metrics(payload: bytes) -> Dict[str, Any]:
    """Return structured management metrics from the workbook's FINAL sheet."""
    workbook = openpyxl_load_workbook(io.BytesIO(payload), data_only=True)
    sheet_name = _find_final_sheet_name(workbook.sheetnames)
    if sheet_name is None:
        raise WorkbookError(
            "The workbook is missing the required sixth sheet 'FINAL'. "
            "Please upload the workbook containing FINAL."
        )
 
    ws = workbook[sheet_name]
    months_done = _parse_months_done(ws)
 
    gs = _parse_final_sales_block(ws, "GROSS SALES", months_done)
    ns = _parse_final_sales_block(ws, "NET SALES", months_done)
    aum = _parse_final_aum_block(ws)

    # Scenario 10 reads its Current and FY27 Budget values directly from FINAL.
    # These detailed matrices are intentionally kept separate from the regular
    # run-rate blocks used by Scenarios 1-9.
    gs_detail = _parse_final_sales_detail_matrix(ws, "GS")
    ns_detail = _parse_final_sales_detail_matrix(ws, "NS")
    gs_market = _parse_final_market_detail_matrix(ws, "GS")
    ns_market = _parse_final_market_detail_matrix(ws, "NS")
 
    return {
        "sheet_name": sheet_name,
        "months_done": months_done,
        "GS": gs,
        "NS": ns,
        "AUM": aum,
        "GS_DETAIL": gs_detail,
        "NS_DETAIL": ns_detail,
        "GS_MARKET": gs_market,
        "NS_MARKET": ns_market,
    }
 
 
# =============================================================================
# 4. SEGMENT & CHANNEL IDENTIFICATION (Scenarios 6, 8, 9)
# =============================================================================
 
def identify_segments(records: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
    """Suggest a column + values that identify each configured business segment."""
    suggestions: Dict[str, Dict[str, Any]] = {}
    for segment, rule in SEGMENT_RULES.items():
        for column in rule["search_columns"]:
            if column not in records.columns:
                continue
            values = sorted({v for v in text_column(records, column) if v.strip()})
            matches = [
                v for v in values
                if any(keyword in v.casefold() for keyword in rule["keywords"])
            ]
            if matches:
                suggestions[segment] = {"column": column, "values": matches}
                break
    return suggestions
 
 
def map_business_segments(
    records: pd.DataFrame,
    mapping: Dict[str, Dict[str, Any]],
) -> pd.DataFrame:
    """Assign every record to Digital / Retail B30 / Others using the mapping."""
    out = records.copy()
    out["Segment"] = FALLBACK_SEGMENT
    # Reverse priority so that the first segment in SEGMENT_ORDER wins.
    for segment in reversed([s for s in SEGMENT_ORDER if s != FALLBACK_SEGMENT]):
        rule = mapping.get(segment)
        if not rule:
            continue
        column, values = rule.get("column"), set(rule.get("values") or [])
        if not column or column not in out.columns or not values:
            continue
        mask = text_column(out, column).isin(values)
        out.loc[mask, "Segment"] = segment
    return out
 
 
def segment_diagnostics(records: pd.DataFrame) -> Dict[str, int]:
    counts = records["Segment"].value_counts().to_dict()
    return {segment: int(counts.get(segment, 0)) for segment in SEGMENT_ORDER}
 
 
CHANNEL_KEYWORDS: Dict[str, List[str]] = {
    "Digital": ["digital", "online", "d2c", "e-com", "ecom", "virtual", "web"],
    "VRM": ["vrm", "virtual relationship", "virtual rm"],
    "EM": ["em", "emerging market", "em city"],
    "B30": ["b30", "b-30", "b 30"],
    "T30": ["t30", "t-30", "t 30"],
    "T8": ["t8", "t-8", "t 8"],
    "DHNI": ["dhni", "d-hni", "hni", "wealth"],
    "Retail": ["retail", "rm retail"],
    "Institutional": ["insti", "institutional", "institution", "institutional sales"],
}
 
 
def _channel_text_score(row: pd.Series, channel: str) -> int:
    values = " ".join(str(row.get(c, "")) for c in META_FIELDS).casefold()
    return sum(1 for keyword in CHANNEL_KEYWORDS[channel] if keyword in values)
 
 
def identify_channels(records: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
    """Suggest channel mappings from workbook metadata without hard-coding one schema."""
    suggestions: Dict[str, Dict[str, Any]] = {}
    usable = [
        f for f in META_FIELDS
        if f in records.columns and text_column(records, f).ne("").any()
    ]
    for channel in CHANNELS:
        best = (0, None, [])
        for column in usable:
            values = sorted({v for v in text_column(records, column) if v.strip()})
            matches = [v for v in values if any(k in v.casefold() for k in CHANNEL_KEYWORDS[channel])]
            score = len(matches)
            if score > best[0]:
                best = (score, column, matches)
        if best[1] and best[2]:
            suggestions[channel] = {"column": best[1], "values": best[2]}
        elif channel == "VRM" and "Vertical" in records.columns:
            suggestions[channel] = {"column": "Vertical", "values": ["VRM"]}
        elif channel == "DHNI" and "Vertical" in records.columns:
            suggestions[channel] = {"column": "Vertical", "values": ["DHNI"]}
        elif channel == "Retail" and "Vertical" in records.columns:
            suggestions[channel] = {"column": "Vertical", "values": ["Retail"]}
    return suggestions
 
 
def map_business_channels(
    records: pd.DataFrame,
    mapping: Dict[str, Dict[str, Any]],
) -> pd.DataFrame:
    """Assign every record to one of the nine Scenario 8 planning channels."""
    out = records.copy()
    out["Channel"] = "Unclassified"
    # More specific channels are applied first; explicit mapping wins.
    for channel in CHANNELS:
        rule = mapping.get(channel)
        if not rule:
            continue
        column, values = rule.get("column"), set(rule.get("values") or [])
        if not column or column not in out.columns or not values:
            continue
        mask = text_column(out, column).isin(values)
        out.loc[mask, "Channel"] = channel
    # Use the existing vertical as a safe fallback for the three explicit RM populations.
    if "Vertical" in out.columns:
        for channel, vertical in (("VRM", "VRM"), ("DHNI", "DHNI"), ("Retail", "Retail")):
            mask = (out["Channel"] == "Unclassified") & (out["Vertical"] == vertical)
            out.loc[mask, "Channel"] = channel
    return out
 
 
# =============================================================================
# 5. BASE GRID & CURRENT-STATE STATISTICS
# =============================================================================
 
def build_base_grid(records: pd.DataFrame) -> pd.DataFrame:
    """Aggregate records to the finest analytical grain used by the engine."""
    rows: List[Dict[str, Any]] = []
    work = records.copy()
    if "MKT TYPE" in work.columns:
        work["MarketType"] = (
            work["MKT TYPE"].astype(str).str.replace("\u00a0", " ", regex=False).str.strip()
        )
        work.loc[work["MarketType"].isin(["", "nan", "None"]), "MarketType"] = "Unspecified"
    else:
        work["MarketType"] = "Unspecified"

    group_cols = ["Vertical", "Segment", "Channel", "MarketType"]
    grouped = work.groupby(group_cols, dropna=False)
    for (vertical, segment, channel, market_type), block in grouped:
        aum_target = float(pd.to_numeric(block.get("aum_target", 0.0), errors="coerce").fillna(0.0).sum()) \
            if "aum_target" in block.columns else 0.0
        for sales in SALES_TYPES:
            for asset in ASSETS:
                rows.append({
                    "Vertical": vertical,
                    "Segment": segment,
                    "Channel": channel,
                    "MarketType": market_type,
                    "Sales": sales,
                    "Asset": asset,
                    "fy_target": float(block[f"{sales}_{asset}_fy"].sum()),
                    "ytd_target": float(block[f"{sales}_{asset}_ytd_tgt"].sum()),
                    "ytd_ach": float(block[f"{sales}_{asset}_ach"].sum()),
                    "aum_target": aum_target,
                })
    return pd.DataFrame(rows)
 
 
def filter_grid(
    grid: pd.DataFrame,
    sales: Optional[str] = None,
    asset: Optional[str] = None,
    vertical: Optional[str] = None,
    segment: Optional[str] = None,
    channel: Optional[str] = None,
    market_type: Optional[str] = None,
) -> pd.DataFrame:
    mask = pd.Series(True, index=grid.index)
    if sales is not None:
        mask &= grid["Sales"] == sales
    if asset is not None:
        mask &= grid["Asset"] == asset
    if vertical is not None:
        mask &= grid["Vertical"] == vertical
    if segment is not None:
        mask &= grid["Segment"] == segment
    if channel is not None and "Channel" in grid.columns:
        mask &= grid["Channel"] == channel
    if market_type is not None and "MarketType" in grid.columns:
        mask &= grid["MarketType"].astype(str) == str(market_type)
    return grid.loc[mask]
 
 
def current_asset_stats(fy_target: float, ytd_target: float, ytd_ach: float) -> Dict[str, Any]:
    """Baseline statistics for one asset / group. Never scenario dependent."""
    fy_target = _z(fy_target)
    ytd_target = _z(ytd_target)
    ytd_ach = _z(ytd_ach)
    current_rr = ytd_ach / MONTHS_COMPLETED if MONTHS_COMPLETED else None
    current_march = ytd_ach + (current_rr or 0.0) * MONTHS_REMAINING
    return {
        "fy_target": fy_target,
        "ytd_target": ytd_target,
        "ytd_ach": ytd_ach,
        "current_rr": current_rr,
        "ytd_ach_pct": safe_div(ytd_ach, ytd_target),
        "fy_completed_pct": safe_div(ytd_ach, fy_target),
        "current_march": current_march,
        "current_march_pct": safe_div(current_march, fy_target),
    }
 
 
def summarize_current(grid: pd.DataFrame, **filters: Any) -> Dict[str, Any]:
    """Baseline statistics for an arbitrary slice of the base grid."""
    subset = filter_grid(grid, **filters)
    return current_asset_stats(
        subset["fy_target"].sum(),
        subset["ytd_target"].sum(),
        subset["ytd_ach"].sum(),
    )
 
 
# =============================================================================
# 6. SCENARIO ENGINE - SCENARIOS 1 TO 6
# =============================================================================
 
def _blank_cell(stats: Dict[str, Any]) -> Dict[str, Any]:
    cell = dict(stats)
    cell.update({
        "scen_rr": None, "rr_change_pct": None, "feb_mar_rr": None,
        "jan_required": None, "jan_amount": None, "jan_pct": None,
        "jan_buffer": None, "jan_buffer_pct": None,
        "march_required": None, "march_amount": None, "march_pct": None,
        "milestone_pct": None, "incremental_sales": None,
        "headroom_amt": None, "headroom_pct": None,
        "momentum_g": None, "feasible": None, "binding": None,
        "trajectory": None, "note": "",
    })
    return cell
 
 
def compute_cell(
    fy_target: float,
    ytd_target: float,
    ytd_ach: float,
    kind: str,
    multiplier: Optional[float] = None,
    uplift: Optional[float] = None,
    dip: float = 0.0,
) -> Dict[str, Any]:
    """Scenario mathematics for one asset / group (scenarios 1-6)."""
    stats = current_asset_stats(fy_target, ytd_target, ytd_ach)
    cell = _blank_cell(stats)
    ach = stats["ytd_ach"]
    current_rr = stats["current_rr"] or 0.0
 
    if kind == "runrate":
        scen_rr = current_rr * (1.0 + (uplift or 0.0))
        jan_amount = ach + scen_rr * MONTHS_JUL_JAN
        cell.update({
            "scen_rr": scen_rr,
            "feb_mar_rr": scen_rr,
            "jan_amount": jan_amount,
            "march_amount": ach + scen_rr * MONTHS_REMAINING,
            "milestone_pct": None,
        })
 
    elif kind == "jan_target":
        required = max(_z(multiplier) * stats["fy_target"], ach)
        scen_rr = max(required - ach, 0.0) / MONTHS_JUL_JAN
        jan_amount = ach + scen_rr * MONTHS_JUL_JAN
        feb_mar_rr = scen_rr * (1.0 - dip)
        cell.update({
            "scen_rr": scen_rr,
            "feb_mar_rr": feb_mar_rr,
            "jan_required": required,
            "jan_amount": jan_amount,
            "march_amount": jan_amount + feb_mar_rr * MONTHS_FEB_MAR,
            "milestone_pct": multiplier,
        })
 
    elif kind == "march_target":
        required = max(_z(multiplier) * stats["fy_target"], ach)
        scen_rr = max(required - ach, 0.0) / MONTHS_REMAINING
        cell.update({
            "scen_rr": scen_rr,
            "feb_mar_rr": scen_rr,
            "march_required": required,
            "jan_amount": ach + scen_rr * MONTHS_JUL_JAN,
            "march_amount": ach + scen_rr * MONTHS_REMAINING,
            "milestone_pct": multiplier,
        })
 
    else:  # pragma: no cover - guarded by the scenario registry
        raise ValueError(f"Unknown scenario kind: {kind}")
 
    return _finalise_cell(cell)
 
 
def _finalise_cell(cell: Dict[str, Any]) -> Dict[str, Any]:
    """Recompute all derived ratios from the absolute amounts in the cell."""
    fy_target = cell.get("fy_target")
    current_rr = _num(cell.get("current_rr"))
    scen_rr = _num(cell.get("scen_rr"))
 
    cell["ytd_ach_pct"] = safe_div(cell.get("ytd_ach"), cell.get("ytd_target"))
    cell["fy_completed_pct"] = safe_div(cell.get("ytd_ach"), fy_target)
    cell["current_march_pct"] = safe_div(cell.get("current_march"), fy_target)
    cell["jan_pct"] = safe_div(cell.get("jan_amount"), fy_target)
    cell["march_pct"] = safe_div(cell.get("march_amount"), fy_target)
 
    if current_rr is not None and current_rr > 0 and scen_rr is not None:
        cell["rr_change_pct"] = (scen_rr / current_rr) - 1.0
    else:
        cell["rr_change_pct"] = None
 
    march_amount = _num(cell.get("march_amount"))
    current_march = _num(cell.get("current_march"))
    if march_amount is not None and current_march is not None:
        cell["incremental_sales"] = march_amount - current_march
 
    jan_required = _num(cell.get("jan_required"))
    jan_amount = _num(cell.get("jan_amount"))
    if jan_required is not None and jan_amount is not None:
        cell["jan_buffer"] = jan_amount - jan_required
        cell["jan_buffer_pct"] = safe_div(cell["jan_buffer"], jan_required)
 
    march_required = _num(cell.get("march_required"))
    if march_required is not None and march_amount is not None:
        cell["headroom_amt"] = march_amount - march_required
        march_pct = cell.get("march_pct")
        required_pct = safe_div(march_required, fy_target)
        if march_pct is not None and required_pct is not None:
            cell["headroom_pct"] = march_pct - required_pct
        cell["feasible"] = cell["headroom_amt"] >= -1e-6
 
    if cell.get("milestone_pct") is None:
        cell["milestone_pct"] = cell.get("march_pct")
    return cell
 
 
def scenario_multipliers(
    grid: pd.DataFrame,
    scenario_id: int,
    params: Optional[Dict[str, Any]] = None,
) -> Dict[Tuple[str, str, str], float]:
    """Derive per-asset target multipliers from the editable scenario assumptions."""
    params = params or {}
    multipliers: Dict[Tuple[str, str, str], float] = {}
 
    if scenario_id in (1, 7, 8, 9, 10):
        return multipliers
 
    if scenario_id == 3:
        target_pct = float(params.get("target_pct", S3_TARGET))
        for sales in SALES_TYPES:
            for asset in ASSETS:
                multipliers[(sales, asset, "*")] = target_pct
        return multipliers
 
    if scenario_id == 4:
        target_pct = float(params.get("target_pct", S4_TARGET))
        for sales in SALES_TYPES:
            for asset in ASSETS:
                multipliers[(sales, asset, "*")] = target_pct
        return multipliers
 
    if scenario_id == 6:
        segment_targets = dict(params.get("segment_targets", S6_SEGMENT_TARGETS))
        for sales in SALES_TYPES:
            for asset in ASSETS:
                for segment in SEGMENT_ORDER:
                    multipliers[(sales, asset, segment)] = float(
                        segment_targets.get(segment, S6_SEGMENT_TARGETS.get(segment, 1.0))
                    )
        return multipliers
 
    # Scenarios 2 and 5 balance Debt and Liquid around editable Equity / Overall ambitions.
    default_equity = S2_EQUITY_TARGET if scenario_id == 2 else S5_EQUITY_TARGET
    default_overall = S2_OVERALL_TARGET if scenario_id == 2 else S5_OVERALL_TARGET
    equity_mult = float(params.get("equity_target", default_equity))
    overall_mult = float(params.get("overall_target", default_overall))
 
    for sales in SALES_TYPES:
        targets = {
            asset: float(filter_grid(grid, sales=sales, asset=asset)["fy_target"].sum())
            for asset in ASSETS
        }
        total_target = sum(targets.values())
        required_overall = overall_mult * total_target
        required_equity = equity_mult * targets["Equity"]
        remaining = max(required_overall - required_equity, 0.0)
        denominator = targets["Debt"] + targets["Liquid"]
        share = safe_div(remaining, denominator)
        balance_mult = 0.0 if share is None else share
        multipliers[(sales, "Equity", "*")] = equity_mult
        multipliers[(sales, "Debt", "*")] = balance_mult
        multipliers[(sales, "Liquid", "*")] = balance_mult
 
    return multipliers
 
 
def _multiplier_for(
    multipliers: Dict[Tuple[str, str, str], float],
    sales: str,
    asset: str,
    segment: str,
) -> Optional[float]:
    if (sales, asset, segment) in multipliers:
        return multipliers[(sales, asset, segment)]
    return multipliers.get((sales, asset, "*"))
 
 
def apply_scenario_grid(
    grid: pd.DataFrame,
    scenario_id: int,
    params: Dict[str, Any],
    multipliers: Dict[Tuple[str, str, str], float],
) -> pd.DataFrame:
    """Evaluate scenarios 1-6 over every cell of the base grid."""
    kind = SCENARIOS[scenario_id]["kind"]
    dip = float(params.get("dip", 0.0)) if scenario_id == 3 else 0.0
    uplift = float(params.get("runrate_uplift", S1_RUNRATE_UPLIFT)) if scenario_id == 1 else None
 
    results: List[Dict[str, Any]] = []
    for row in grid.to_dict("records"):
        multiplier = _multiplier_for(multipliers, row["Sales"], row["Asset"], row["Segment"])
        cell = compute_cell(
            row["fy_target"], row["ytd_target"], row["ytd_ach"],
            kind=kind, multiplier=multiplier, uplift=uplift, dip=dip,
        )
        cell.update({
            "Vertical": row["Vertical"], "Segment": row["Segment"],
            "Channel": row.get("Channel", "Unclassified"),
            "Sales": row["Sales"], "Asset": row["Asset"],
        })
        results.append(cell)
    return pd.DataFrame(results)
 
 
SUMMABLE_FIELDS = [
    "fy_target", "ytd_target", "ytd_ach", "current_rr", "current_march",
    "scen_rr", "feb_mar_rr", "jan_required", "jan_amount",
    "march_required", "march_amount",
]
 
 
def summarize_cells(subset: pd.DataFrame) -> Dict[str, Any]:
    """Aggregate scenario cells additively and rebuild every derived ratio."""
    cell: Dict[str, Any] = {}
    for field in SUMMABLE_FIELDS:
        cell[field] = _ssum(subset[field]) if field in subset.columns else None
    milestones = subset["milestone_pct"].dropna().unique() if "milestone_pct" in subset else []
    cell["milestone_pct"] = float(milestones[0]) if len(milestones) == 1 else None
    cell["note"] = ""
    cell["momentum_g"] = None
    cell["trajectory"] = None
    cell["binding"] = None
    cell["feasible"] = None
    return _finalise_cell(cell)
 
 
# --- Named scenario entry points (thin wrappers over the shared engine) -------
 
def _scenario_frame(grid: pd.DataFrame, scenario_id: int, params: Dict[str, Any]) -> pd.DataFrame:
    return apply_scenario_grid(grid, scenario_id, params, scenario_multipliers(grid, scenario_id, params))
 
 
def calculate_scenario_1(grid: pd.DataFrame, params: Dict[str, Any]) -> pd.DataFrame:
    """+20% run-rate push from July onward."""
    return _scenario_frame(grid, 1, params)
 
 
def calculate_scenario_2(grid: pd.DataFrame, params: Dict[str, Any]) -> pd.DataFrame:
    """100% Equity and 75% overall FY target by January."""
    return _scenario_frame(grid, 2, params)
 
 
def calculate_scenario_3(grid: pd.DataFrame, params: Dict[str, Any]) -> pd.DataFrame:
    """100% of FY target by January, then a configurable Feb-Mar dip."""
    return _scenario_frame(grid, 3, params)
 
 
def calculate_scenario_4(grid: pd.DataFrame, params: Dict[str, Any]) -> pd.DataFrame:
    """120% of FY target by March."""
    return _scenario_frame(grid, 4, params)
 
 
def calculate_scenario_5(grid: pd.DataFrame, params: Dict[str, Any]) -> pd.DataFrame:
    """120% Equity and 100% overall FY target by March."""
    return _scenario_frame(grid, 5, params)
 
 
def calculate_scenario_6(grid: pd.DataFrame, params: Dict[str, Any]) -> pd.DataFrame:
    """Digital 140%, Retail B30 125%, Others 115% of their FY targets."""
    return _scenario_frame(grid, 6, params)
 
 
SCENARIO_FUNCTIONS = {
    1: calculate_scenario_1,
    2: calculate_scenario_2,
    3: calculate_scenario_3,
    4: calculate_scenario_4,
    5: calculate_scenario_5,
    6: calculate_scenario_6,
}
 
 
# =============================================================================
# 7. SCENARIO 7 - MOMENTUM ENGINE
# =============================================================================
 
def _momentum_sum(growth: float, months: int, tail_factors: Sequence[float]) -> float:
    """Sum of compounding monthly run-rate multiples, including leakage tail."""
    factor = 1.0 + growth
    build = sum(factor ** k for k in range(1, months + 1))
    tail = sum(f * factor ** months for f in tail_factors)
    return build + tail
 
 
def solve_momentum_rate(
    current_rr: float,
    required_amount: float,
    months: int = MONTHS_JUL_JAN,
    tail_factors: Sequence[float] = (),
    upper: float = 3.0,
) -> Optional[float]:
    """
    Back-solve the minimum month-on-month growth rate g such that the
    compounding trajectory delivers the required incremental amount.
 
    Returns 0.0 when no additional momentum is needed and None when the
    requirement cannot be met within the search bounds.
    """
    rr = _num(current_rr)
    need = _num(required_amount)
    if rr is None or need is None or rr <= 0:
        return None
    if need <= 0:
        return 0.0
 
    def shortfall(growth: float) -> float:
        return rr * _momentum_sum(growth, months, tail_factors) - need
 
    if shortfall(0.0) >= 0:
        return 0.0
    if shortfall(upper) < 0:
        return None
 
    low, high = 0.0, upper
    for _ in range(240):
        mid = (low + high) / 2.0
        if shortfall(mid) >= 0:
            high = mid
        else:
            low = mid
    return high
 
 
def calculate_momentum_trajectory(
    current_rr: float,
    growth: Optional[float],
    leakage: float,
    flat_rate: Optional[float] = None,
) -> List[float]:
    """Monthly run rates for July -> March (momentum build, then leakage)."""
    if growth is None:
        base = _z(flat_rate)
        build = [base] * MONTHS_JUL_JAN
    else:
        rr = _z(current_rr)
        build = [rr * (1.0 + growth) ** k for k in range(1, MONTHS_JUL_JAN + 1)]
    january_rr = build[-1] if build else 0.0
    february_rr = january_rr * (1.0 - leakage)
    march_rr = february_rr * (1.0 - leakage)
    return build + [february_rr, march_rr]
 
 
def calculate_leakage_impact(january_rr: float, leakage: float) -> Dict[str, float]:
    """February and March run rates after AUM leakage / run-rate pressure."""
    february_rr = _z(january_rr) * (1.0 - leakage)
    march_rr = february_rr * (1.0 - leakage)
    return {
        "february_rr": february_rr,
        "march_rr": march_rr,
        "feb_mar_sales": february_rr + march_rr,
    }
 
 
def calculate_momentum_headroom(
    march_amount: float,
    march_required: float,
    fy_target: float,
    march_target_pct: float,
) -> Dict[str, Optional[float]]:
    """Scenario achievement versus the March ambition, in Cr and in points."""
    headroom_amt = _z(march_amount) - _z(march_required)
    achieved_pct = safe_div(march_amount, fy_target)
    headroom_pct = None if achieved_pct is None else achieved_pct - march_target_pct
    return {"headroom_amt": headroom_amt, "headroom_pct": headroom_pct}
 
 
def calculate_scenario_7(
    fy_target: float,
    ytd_target: float,
    ytd_ach: float,
    params: Dict[str, Any],
) -> Dict[str, Any]:
    """
    Momentum build-up model.
 
    Solves for the month-on-month growth rate that simultaneously satisfies the
    January milestone and, after Feb-Mar leakage, the March ambition.
    """
    jan_target_pct = float(params.get("jan_target", S7_DEFAULT_JAN_TARGET))
    mar_target_pct = float(params.get("mar_target", S7_DEFAULT_MAR_TARGET))
    leakage = float(params.get("leakage", S7_DEFAULT_LEAKAGE))
 
    stats = current_asset_stats(fy_target, ytd_target, ytd_ach)
    cell = _blank_cell(stats)
    ach = stats["ytd_ach"]
    current_rr = stats["current_rr"] or 0.0
 
    jan_required = max(jan_target_pct * stats["fy_target"], ach)
    mar_required = max(mar_target_pct * stats["fy_target"], ach)
    tail = ((1.0 - leakage), (1.0 - leakage) ** 2)
 
    growth_jan = solve_momentum_rate(current_rr, jan_required - ach, MONTHS_JUL_JAN)
    growth_mar = solve_momentum_rate(current_rr, mar_required - ach, MONTHS_JUL_JAN, tail)
 
    note = ""
    binding = None
    flat_rate = None
 
    if current_rr <= 0:
        # Momentum compounding is undefined on a non-positive run rate:
        # fall back to the flat run rate required to hold both milestones.
        flat_jan = max(jan_required - ach, 0.0) / MONTHS_JUL_JAN
        denominator = MONTHS_JUL_JAN + tail[0] + tail[1]
        flat_mar = max(mar_required - ach, 0.0) / denominator
        flat_rate = max(flat_jan, flat_mar)
        growth = None
        binding = "January" if flat_jan >= flat_mar else "March"
        note = (
            "Current run rate is not positive, so compounding momentum cannot be applied. "
            "The flat monthly run rate required to hold the milestones is shown instead."
        )
    elif growth_jan is None and growth_mar is None:
        growth = None
        flat_rate = max(mar_required - ach, 0.0) / (MONTHS_JUL_JAN + tail[0] + tail[1])
        binding = "March"
        note = "The requirement exceeds the momentum search range; a flat required run rate is shown."
    else:
        candidates = [g for g in (growth_jan, growth_mar) if g is not None]
        growth = max(candidates)
        binding = "March" if (growth_mar is not None and growth == growth_mar
                              and (growth_jan is None or growth_mar >= growth_jan)) else "January"
 
    trajectory = calculate_momentum_trajectory(current_rr, growth, leakage, flat_rate)
    build_phase = trajectory[:MONTHS_JUL_JAN]
    jan_amount = ach + sum(build_phase)
    january_rr = build_phase[-1] if build_phase else 0.0
    leak = calculate_leakage_impact(january_rr, leakage)
    march_amount = jan_amount + leak["feb_mar_sales"]
 
    headroom = calculate_momentum_headroom(
        march_amount, mar_required, stats["fy_target"], mar_target_pct
    )
    shortfall = max(mar_required - march_amount, 0.0)
    denominator = 1.0 + tail[0] + tail[1]
    additional_jan_rr = shortfall / denominator if denominator else None
 
    cell.update({
        "scen_rr": january_rr,
        "feb_mar_rr": leak["february_rr"],
        "march_rr": leak["march_rr"],
        "jan_required": jan_required,
        "jan_amount": jan_amount,
        "march_required": mar_required,
        "march_amount": march_amount,
        "milestone_pct": jan_target_pct,
        "march_target_pct": mar_target_pct,
        "momentum_g": growth,
        "flat_rate": flat_rate,
        "leakage": leakage,
        "trajectory": trajectory,
        "binding": binding,
        "note": note,
        "additional_march_sales": shortfall,
        "additional_jan_rr": additional_jan_rr,
        "avg_scen_rr": (sum(trajectory) / len(trajectory)) if trajectory else None,
    })
    cell = _finalise_cell(cell)
    cell["headroom_amt"] = headroom["headroom_amt"]
    cell["headroom_pct"] = headroom["headroom_pct"]
    cell["feasible"] = shortfall <= 1e-6
    # Momentum run rate versus the current flat run rate, measured on the
    # January exit rate (the pace the business must be running at by then).
    cell["rr_change_pct"] = (
        (january_rr / current_rr) - 1.0 if current_rr and current_rr > 0 else None
    )
    return cell
 
 
# =============================================================================
# 7A. SCENARIO 8/9 - CHANNEL SIMULATOR & MIX OPTIMISER
# =============================================================================
 
def _s8_channel_params(params: Dict[str, Any], channel: str) -> Tuple[float, float, float, float]:
    growth = float(params.get("channel_growth", {}).get(channel, S8_DEFAULT_GROWTH.get(channel, 0.05)))
    jan_target = float(
        params.get("channel_jan_target", {}).get(channel, S8_DEFAULT_JAN_TARGET.get(channel, 1.0))
    )
    mar_target = float(
        params.get("channel_mar_target", {}).get(channel, S8_DEFAULT_MAR_TARGET.get(channel, 1.0))
    )
    leakage = float(params.get("leakage", S8_DEFAULT_LEAKAGE))
    return growth, jan_target, mar_target, leakage
 
 
def calculate_scenario_8_cell(
    fy_target: float,
    ytd_target: float,
    ytd_ach: float,
    channel: str,
    params: Dict[str, Any],
) -> Dict[str, Any]:
    """Fixed-growth channel trajectory with independent Jan-2027 and Mar-2027 targets."""
    stats = current_asset_stats(fy_target, ytd_target, ytd_ach)
    cell = _blank_cell(stats)
    growth, jan_target_pct, mar_target_pct, leakage = _s8_channel_params(params, channel)
    current_rr = _z(stats["current_rr"])
    trajectory = calculate_momentum_trajectory(current_rr, growth, leakage)
    jan_amount = _z(stats["ytd_ach"]) + sum(trajectory[:MONTHS_JUL_JAN])
    march_amount = jan_amount + sum(trajectory[MONTHS_JUL_JAN:])
    jan_required = max(jan_target_pct * stats["fy_target"], _z(stats["ytd_ach"]))
    march_required = max(mar_target_pct * stats["fy_target"], _z(stats["ytd_ach"]))
    jan_gap = jan_amount - jan_required
    march_gap = march_amount - march_required
    cell.update({
        "channel": channel,
        "scen_rr": trajectory[MONTHS_JUL_JAN - 1] if trajectory else current_rr,
        "feb_mar_rr": trajectory[MONTHS_JUL_JAN] if len(trajectory) > MONTHS_JUL_JAN else None,
        "march_rr": trajectory[-1] if trajectory else None,
        "jan_required": jan_required, "jan_amount": jan_amount,
        "march_required": march_required, "march_amount": march_amount,
        "milestone_pct": jan_target_pct, "march_target_pct": mar_target_pct,
        "momentum_g": growth, "leakage": leakage, "trajectory": trajectory,
        "jan_buffer": jan_gap, "jan_buffer_pct": safe_div(jan_gap, jan_required),
        "headroom_amt": march_gap, "headroom_pct": safe_div(march_gap, march_required),
        "feasible": jan_gap >= -1e-6 and march_gap >= -1e-6,
        "binding": "January" if jan_gap < 0 else ("March" if march_gap < 0 else "None"),
        "additional_march_sales": max(-march_gap, 0.0),
        "additional_jan_rr": max(-jan_gap, 0.0) / MONTHS_JUL_JAN,
    })
    return _finalise_cell(cell)
 
 
def calculate_scenario_8_grid(grid: pd.DataFrame, params: Dict[str, Any]) -> pd.DataFrame:
    rows = []
    for row in grid.to_dict("records"):
        channel = row.get("Channel", "Unclassified")
        cell = calculate_scenario_8_cell(
            row["fy_target"], row["ytd_target"], row["ytd_ach"], channel, params
        )
        cell.update({k: row[k] for k in ("Vertical", "Segment", "Channel", "Sales", "Asset")})
        rows.append(cell)
    return pd.DataFrame(rows)
 
 
def calculate_scenario_9_grid(grid: pd.DataFrame, params: Dict[str, Any]) -> pd.DataFrame:
    """Optimiser: solve minimum MoM growth per channel to meet its Jan/Mar targets."""
    rows = []
    for row in grid.to_dict("records"):
        channel = row.get("Channel", "Unclassified")
        _, jan_target, mar_target, leakage = _s8_channel_params(params, channel)
        stats = current_asset_stats(row["fy_target"], row["ytd_target"], row["ytd_ach"])
        ach = _z(stats["ytd_ach"])
        rr = _z(stats["current_rr"])
        jan_req = max(jan_target * stats["fy_target"], ach)
        mar_req = max(mar_target * stats["fy_target"], ach)
        gj = solve_momentum_rate(rr, jan_req - ach, MONTHS_JUL_JAN)
        gm = solve_momentum_rate(rr, mar_req - ach, MONTHS_JUL_JAN, ((1 - leakage), (1 - leakage) ** 2))
        growth = max([g for g in (gj, gm) if g is not None], default=0.0)
        cell = calculate_scenario_8_cell(
            row["fy_target"], row["ytd_target"], row["ytd_ach"], channel,
            {**params, "channel_growth": {**params.get("channel_growth", {}), channel: growth}},
        )
        cell.update({k: row[k] for k in ("Vertical", "Segment", "Channel", "Sales", "Asset")})
        cell["optimized_growth"] = growth
        rows.append(cell)
    return pd.DataFrame(rows)
 
 

# =============================================================================
# 7B. SCENARIO 10 - FINAL-SHEET SIMULATION MODEL
# =============================================================================

# Scenario 10 is intentionally different from Scenarios 1-9:
#   * CURRENT = the YTD value printed on FINAL (not YTD annualised to March)
#   * FY27 BUDGET = the Target value printed on FINAL
#   * PROJECTED NUMBER = editable Achievement % × FY27 BUDGET
#   * Digital is fully excluded from the simulation roll-up
#   * T2/T6/T30/B30/EM use the market matrix from FINAL only
S10_PLANNING_CHANNELS: List[str] = ["Retail", "DHNI", "VRM", "Institutional"]
S10_MARKET_CHANNELS: List[str] = ["Retail", "VRM", "DHNI"]
S10_CHANNEL_LABELS: Dict[str, str] = {
    "Retail": "Retail",
    "DHNI": "DHNI",
    "VRM": "VRM",
    "Institutional": "Insti",
}


def _scenario10_market_bucket(value: Any) -> str:
    """Canonical market bucket: B30 Select -> B30, T30 Ext -> T30."""
    raw = str(value).replace("\u00a0", " ").strip()
    compact = " ".join(raw.replace("-", " ").replace("_", " ").split()).casefold()
    if compact in {"b30", "b30 select", "b30select"}:
        return "B30"
    if compact in {"t30", "t30 ext", "t30ext", "t30 extended"}:
        return "T30"
    return raw or "Unspecified"


def _scenario10_management_channel_from_values(vertical: Any, channel: Any) -> str:
    vertical_text = str(vertical).strip()
    channel_text = str(channel).strip().casefold()
    if channel_text == "digital" or vertical_text.casefold() == "digital":
        return "Digital"
    if channel_text in {"institutional", "insti"} or vertical_text.casefold() in {"institutional", "insti"}:
        return "Institutional"
    if vertical_text in {"Retail", "DHNI", "VRM"}:
        return vertical_text
    return vertical_text or "Unclassified"


def _scenario10_management_channel(row: Dict[str, Any]) -> str:
    if row.get("Scenario10Channel"):
        return str(row.get("Scenario10Channel"))
    return _scenario10_management_channel_from_values(
        row.get("Vertical"), row.get("Channel")
    )


def _scenario10_direct_stats(
    fy_target: float,
    current_value: float,
    months_done: int = MONTHS_COMPLETED,
) -> Dict[str, Any]:
    """
    FINAL-sheet baseline for Scenario 10.

    IMPORTANT: `current_march` is deliberately the FINAL YTD/current value.
    It is NOT multiplied by 4 and is NOT annualised.
    """
    fy_target = _z(fy_target)
    current_value = _z(current_value)
    months = max(int(months_done), 1)
    current_pct = safe_div(current_value, fy_target)
    return {
        "fy_target": fy_target,
        "ytd_target": 0.0,
        "ytd_ach": current_value,
        "current_rr": current_value / months,
        "ytd_ach_pct": current_pct,
        "fy_completed_pct": current_pct,
        "current_march": current_value,
        "current_march_pct": current_pct,
    }


def build_scenario10_final_grid(final_metrics: Dict[str, Any]) -> pd.DataFrame:
    """
    Create Scenario-10 input rows exclusively from FINAL.

    Channel rows come from the detailed FINAL GS/NS matrix. Market rows come
    from a T2/T6/T30/B30/EM matrix in FINAL when available. Digital is never
    inserted into the Scenario-10 calculation grid.
    """
    rows: List[Dict[str, Any]] = []

    for sales in SALES_TYPES:
        detail = final_metrics.get(f"{sales}_DETAIL")
        if not isinstance(detail, pd.DataFrame) or detail.empty:
            continue

        # Main channel × asset matrix. Do not use the Overall group here because
        # Equity + Debt + Liquid are the additive building blocks.
        for channel_label, planning_channel in [
            ("Retail", "Retail"),
            ("DHNI", "DHNI"),
            ("VRM", "VRM"),
            ("Insti", "Institutional"),
        ]:
            for asset in ASSETS:
                hit = detail.loc[
                    (detail["Channel"] == channel_label)
                    & (detail["Asset"] == asset)
                ]
                if hit.empty:
                    target = 0.0
                    current = 0.0
                else:
                    target = _z(hit.iloc[0].get("FY27 Target"))
                    current = _z(hit.iloc[0].get("Current"))

                rows.append({
                    "GridRole": "Channel",
                    "Vertical": channel_label,
                    "Segment": "FINAL",
                    "Channel": "Institutional" if planning_channel == "Institutional" else channel_label,
                    "MarketType": "Unspecified",
                    "Scenario10Channel": planning_channel,
                    "Sales": sales,
                    "Asset": asset,
                    "fy_target": target,
                    "ytd_target": 0.0,
                    "ytd_ach": current,
                    "final_current": current,
                    "source_sheet": "FINAL",
                })

        # Preserve the exact published FINAL Asset and Overall numbers. The
        # detailed channel matrix can differ by 1 Cr because displayed values
        # are rounded. Adjustment rows fix only those rounding differences and
        # never add a simulated projected number.
        published = final_metrics.get(sales)
        if isinstance(published, pd.DataFrame) and not published.empty:
            for published_asset in ASSETS:
                if published_asset not in published.index:
                    continue
                channel_rows = [
                    row for row in rows
                    if row.get("Sales") == sales
                    and row.get("GridRole") == "Channel"
                    and row.get("Asset") == published_asset
                ]
                detail_target = sum(_z(row.get("fy_target")) for row in channel_rows)
                detail_current = sum(_z(row.get("ytd_ach")) for row in channel_rows)
                target_adjustment = _z(published.loc[published_asset].get("FY27 Target")) - detail_target
                current_adjustment = _z(published.loc[published_asset].get("YTD")) - detail_current
                if abs(target_adjustment) > 1e-9 or abs(current_adjustment) > 1e-9:
                    rows.append({
                        "GridRole": "Adjustment",
                        "Vertical": f"{published_asset} Adjustment",
                        "Segment": "FINAL",
                        "Channel": f"{published_asset} Adjustment",
                        "MarketType": "Unspecified",
                        "Scenario10Channel": "Adjustment",
                        "Sales": sales,
                        "Asset": published_asset,
                        "fy_target": target_adjustment,
                        "ytd_target": 0.0,
                        "ytd_ach": current_adjustment,
                        "final_current": current_adjustment,
                        "source_sheet": "FINAL",
                    })

            if "Overall" in published.index:
                # At this point Channel + asset-adjustment rows equal the exact
                # published Equity/Debt/Liquid numbers. Reconcile their sum to
                # the published Overall line as a final rounding adjustment.
                published_target = _z(published.loc["Overall"].get("FY27 Target"))
                published_current = _z(published.loc["Overall"].get("YTD"))
                current_rows = [
                    row for row in rows
                    if row.get("Sales") == sales
                    and row.get("GridRole") in {"Channel", "Adjustment"}
                    and row.get("Asset") != "Overall Adjustment"
                ]
                detail_target = sum(_z(row.get("fy_target")) for row in current_rows)
                detail_current = sum(_z(row.get("ytd_ach")) for row in current_rows)
                target_adjustment = published_target - detail_target
                current_adjustment = published_current - detail_current
                if abs(target_adjustment) > 1e-9 or abs(current_adjustment) > 1e-9:
                    rows.append({
                        "GridRole": "Adjustment",
                        "Vertical": "Overall Adjustment",
                        "Segment": "FINAL",
                        "Channel": "Overall Adjustment",
                        "MarketType": "Unspecified",
                        "Scenario10Channel": "Adjustment",
                        "Sales": sales,
                        "Asset": "Overall Adjustment",
                        "fy_target": target_adjustment,
                        "ytd_target": 0.0,
                        "ytd_ach": current_adjustment,
                        "final_current": current_adjustment,
                        "source_sheet": "FINAL",
                    })

        # Optional FINAL market-type matrix. This is a separate alternate cut;
        # it is excluded from the main channel/asset aggregation to avoid double counting.
        market = final_metrics.get(f"{sales}_MARKET")
        if isinstance(market, pd.DataFrame) and not market.empty:
            for market_type in FINAL_MARKET_ROWS:
                for asset in ASSETS:
                    hit = market.loc[
                        (market["Market Type"] == market_type)
                        & (market["Asset"] == asset)
                    ]
                    if hit.empty:
                        continue
                    rows.append({
                        "GridRole": "Market",
                        "Vertical": "Market",
                        "Segment": "FINAL",
                        "Channel": "Market",
                        "MarketType": _scenario10_market_bucket(market_type),
                        "Scenario10Channel": "Market",
                        "Sales": sales,
                        "Asset": asset,
                        "fy_target": _z(hit.iloc[0].get("FY27 Target")),
                        "ytd_target": 0.0,
                        "ytd_ach": _z(hit.iloc[0].get("Current")),
                        "final_current": _z(hit.iloc[0].get("Current")),
                        "source_sheet": "FINAL",
                    })

    return pd.DataFrame(rows)


def _scenario10_subset(
    grid: pd.DataFrame,
    sales: Optional[str] = None,
    asset: Optional[str] = None,
    planning_channel: Optional[str] = None,
    market_type: Optional[str] = None,
    include_adjustment: bool = False,
) -> pd.DataFrame:
    """Main Scenario-10 channel slice; Digital and market rows are excluded."""
    if grid is None or grid.empty:
        return pd.DataFrame(columns=getattr(grid, "columns", []))

    role = grid.get("GridRole", pd.Series("Channel", index=grid.index)).astype(str)
    allowed_roles = ["Channel", "Adjustment"] if include_adjustment else ["Channel"]
    mask = role.isin(allowed_roles)
    if sales is not None:
        mask &= grid["Sales"] == sales
    if asset is not None:
        mask &= grid["Asset"] == asset
    if planning_channel is not None:
        scenario_channel = grid.get(
            "Scenario10Channel", pd.Series("", index=grid.index)
        ).astype(str)
        mask &= scenario_channel == planning_channel
    if market_type is not None:
        # FINAL channel matrix does not contain a channel × market cross-tab.
        # Keep this unsupported slice empty rather than inventing a split.
        mask &= False
    return grid.loc[mask].copy()


def _scenario10_market_subset(
    grid: pd.DataFrame,
    sales: Optional[str] = None,
    asset: Optional[str] = None,
    market_type: Optional[str] = None,
) -> pd.DataFrame:
    """FINAL-only T2/T6/T30/B30/EM slice."""
    if grid is None or grid.empty:
        return pd.DataFrame(columns=getattr(grid, "columns", []))
    role = grid.get("GridRole", pd.Series("", index=grid.index)).astype(str)
    mask = role == "Market"
    if sales is not None:
        mask &= grid["Sales"] == sales
    if asset is not None:
        mask &= grid["Asset"] == asset
    if market_type is not None:
        requested = _scenario10_market_bucket(market_type)
        canonical = grid.get(
            "MarketType", pd.Series("Unspecified", index=grid.index)
        ).map(_scenario10_market_bucket)
        mask &= canonical == requested
    return grid.loc[mask].copy()


def _scenario10_current_stats(
    grid: pd.DataFrame,
    sales: str,
    asset: str,
    planning_channel: Optional[str] = None,
    market_type: Optional[str] = None,
) -> Dict[str, Any]:
    subset = _scenario10_subset(
        grid,
        sales=sales,
        asset=asset,
        planning_channel=planning_channel,
        market_type=market_type,
    )
    if subset.empty:
        return _scenario10_direct_stats(0.0, 0.0)
    return _scenario10_direct_stats(
        subset["fy_target"].sum(),
        subset["ytd_ach"].sum(),
    )


def _scenario10_safe_current_pct(stats: Dict[str, Any]) -> float:
    value = _num(stats.get("current_march_pct"))
    return 0.0 if value is None else float(value)


def _scenario10_market_stats(
    grid: pd.DataFrame,
    sales: str,
    asset: str,
    market_type: Optional[str] = None,
) -> Dict[str, Any]:
    subset = _scenario10_market_subset(
        grid, sales=sales, asset=asset, market_type=market_type
    )
    if subset.empty:
        return _scenario10_direct_stats(0.0, 0.0)
    return _scenario10_direct_stats(
        subset["fy_target"].sum(),
        subset["ytd_ach"].sum(),
    )


def _scenario10_market_current_pct(
    grid: pd.DataFrame,
    sales: str,
    asset: str,
    market_type: str,
) -> float:
    return _scenario10_safe_current_pct(
        _scenario10_market_stats(grid, sales, asset, market_type)
    )


def _scenario10_market_fy_target(
    grid: pd.DataFrame,
    sales: str,
    asset: str,
    market_type: str,
) -> float:
    subset = _scenario10_market_subset(
        grid, sales=sales, asset=asset, market_type=market_type
    )
    return float(subset["fy_target"].sum()) if not subset.empty else 0.0


def _scenario10_retail_locations(grid: pd.DataFrame) -> List[str]:
    subset = _scenario10_market_subset(grid)
    if subset.empty or "MarketType" not in subset.columns:
        return []
    present = {
        _scenario10_market_bucket(value)
        for value in subset["MarketType"].astype(str)
    }
    return [value for value in FINAL_MARKET_ROWS if value in present]


def _scenario10_current_target_map(
    grid: pd.DataFrame,
) -> Dict[str, Dict[str, Dict[str, float]]]:
    targets: Dict[str, Dict[str, Dict[str, float]]] = {
        sales: {asset: {} for asset in ASSETS} for sales in SALES_TYPES
    }
    for sales in SALES_TYPES:
        for asset in ASSETS:
            for channel in S10_PLANNING_CHANNELS:
                targets[sales][asset][channel] = _scenario10_safe_current_pct(
                    _scenario10_current_stats(
                        grid, sales, asset, planning_channel=channel
                    )
                )
    return targets


def _scenario10_location_current_map(
    grid: pd.DataFrame,
) -> Dict[str, Dict[str, Dict[str, float]]]:
    locations = _scenario10_retail_locations(grid)
    result: Dict[str, Dict[str, Dict[str, float]]] = {
        sales: {asset: {} for asset in ASSETS} for sales in SALES_TYPES
    }
    for sales in SALES_TYPES:
        for asset in ASSETS:
            for location in locations:
                result[sales][asset][location] = _scenario10_market_current_pct(
                    grid, sales, asset, location
                )
    return result


def _scenario10_target_for(
    params: Dict[str, Any],
    sales: str,
    asset: str,
    planning_channel: str,
    market_type: str,
    current_floor: float,
) -> float:
    """Scenario 10 allows any non-negative simulation percentage; no current floor."""
    configured = (
        params.get("asset_vertical_targets", {})
        .get(sales, {})
        .get(asset, {})
        .get(planning_channel, current_floor)
    )
    value = _num(configured)
    return max(0.0, current_floor if value is None else float(value))


def _scenario10_market_target_for(
    params: Dict[str, Any],
    sales: str,
    asset: str,
    market_type: str,
    current_pct: float,
) -> float:
    configured = (
        params.get("retail_location_targets", {})
        .get(sales, {})
        .get(asset, {})
        .get(_scenario10_market_bucket(market_type), current_pct)
    )
    value = _num(configured)
    return max(0.0, current_pct if value is None else float(value))


def _scenario10_compute_cell(
    fy_target: float,
    current_value: float,
    target_pct: float,
    role: str = "Channel",
) -> Dict[str, Any]:
    """FINAL-based Scenario-10 cell: Projected Number = % × FY27 Budget."""
    stats = _scenario10_direct_stats(fy_target, current_value)
    cell = _blank_cell(stats)

    if role == "Adjustment":
        # Preserve FINAL's published overall rounding/current without adding a
        # projected amount to the channel simulation.
        cell.update({
            "scen_rr": 0.0,
            "feb_mar_rr": 0.0,
            "jan_amount": current_value,
            "march_required": 0.0,
            "march_amount": 0.0,
            "milestone_pct": None,
        })
        return _finalise_cell(cell)

    projected = _z(fy_target) * max(0.0, float(target_pct))
    scen_rr = (projected - _z(current_value)) / max(MONTHS_REMAINING, 1)
    cell.update({
        "scen_rr": scen_rr,
        "feb_mar_rr": scen_rr,
        "march_required": projected,
        "jan_amount": _z(current_value) + scen_rr * MONTHS_JUL_JAN,
        "march_amount": projected,
        "milestone_pct": max(0.0, float(target_pct)),
    })
    return _finalise_cell(cell)


def calculate_scenario_10_grid(
    grid: pd.DataFrame,
    params: Dict[str, Any],
) -> pd.DataFrame:
    """Evaluate Scenario 10 entirely on the FINAL-derived grid."""
    rows: List[Dict[str, Any]] = []

    for row in grid.to_dict("records"):
        role = str(row.get("GridRole", "Channel"))
        sales = row["Sales"]
        asset = row["Asset"]
        current_value = _z(row.get("ytd_ach"))
        fy_target = _z(row.get("fy_target"))

        if role == "Adjustment":
            target_pct = 0.0
        elif role == "Market":
            current_pct = safe_div(current_value, fy_target) or 0.0
            target_pct = _scenario10_market_target_for(
                params,
                sales,
                asset,
                str(row.get("MarketType", "Unspecified")),
                current_pct,
            )
        else:
            planning_channel = _scenario10_management_channel(row)
            current_pct = safe_div(current_value, fy_target) or 0.0
            target_pct = _scenario10_target_for(
                params,
                sales,
                asset,
                planning_channel,
                str(row.get("MarketType", "Unspecified")),
                current_pct,
            )

        cell = _scenario10_compute_cell(
            fy_target,
            current_value,
            target_pct,
            role=role,
        )
        cell.update({
            "GridRole": role,
            "Vertical": row.get("Vertical", ""),
            "Segment": row.get("Segment", "FINAL"),
            "Channel": row.get("Channel", ""),
            "MarketType": row.get("MarketType", "Unspecified"),
            "Scenario10Channel": row.get("Scenario10Channel", ""),
            "Scenario10MarketType": _scenario10_market_bucket(row.get("MarketType", "Unspecified")),
            "Sales": sales,
            "Asset": asset,
            "scenario_10_target_pct": target_pct,
            "source_sheet": "FINAL",
        })
        rows.append(cell)

    return pd.DataFrame(rows)


def build_scenario_10_asset_summary(
    model: "ScenarioModel",
    sales: str,
) -> Tuple[pd.DataFrame, Dict[str, str]]:
    rows: List[Dict[str, Any]] = []
    for asset in ASSETS:
        cell = model.cell(sales, asset=asset)
        current_pct = _num(cell.get("current_march_pct"))
        scenario_pct = _num(cell.get("march_pct"))
        change_pts = None if current_pct is None or scenario_pct is None else scenario_pct - current_pct
        relative_change = (
            None if current_pct in (None, 0) or scenario_pct is None
            else scenario_pct / current_pct - 1.0
        )
        rows.append({
            "Asset Class": asset,
            "FY27 Budget": cell.get("fy_target"),
            "Current (FINAL YTD)": cell.get("current_march"),
            "Current %": current_pct,
            "Simulation Projected Number": cell.get("march_amount"),
            "Simulation Achievement %": scenario_pct,
            "Change vs Current": change_pts,
            "Relative Change %": relative_change,
            "Required Remaining-Month RR": cell.get("scen_rr"),
        })
    formats = {
        "Asset Class": "txt",
        "FY27 Budget": "cr",
        "Current (FINAL YTD)": "cr",
        "Current %": "pct",
        "Simulation Projected Number": "cr",
        "Simulation Achievement %": "pct",
        "Change vs Current": "pts",
        "Relative Change %": "pct_signed",
        "Required Remaining-Month RR": "cr",
    }
    return pd.DataFrame(rows), formats


def build_scenario_10_channel_detail(
    model: "ScenarioModel",
    sales: str,
) -> Tuple[pd.DataFrame, Dict[str, str]]:
    rows: List[Dict[str, Any]] = []
    frame = model.scenario_grid
    if frame is None or frame.empty:
        return pd.DataFrame(), {}

    channel_frame = frame.loc[
        frame.get("GridRole", pd.Series("Channel", index=frame.index)).astype(str) == "Channel"
    ]
    for asset in ASSETS:
        for channel in S10_PLANNING_CHANNELS:
            subset = channel_frame.loc[
                (channel_frame["Sales"] == sales)
                & (channel_frame["Asset"] == asset)
                & (channel_frame["Scenario10Channel"] == channel)
            ]
            if subset.empty:
                continue
            cell = summarize_cells(subset)
            current_pct = _num(cell.get("current_march_pct"))
            scenario_pct = _num(cell.get("march_pct"))
            change_pts = None if current_pct is None or scenario_pct is None else scenario_pct - current_pct
            rows.append({
                "Asset Class": asset,
                "Channel": S10_CHANNEL_LABELS.get(channel, channel),
                "FY27 Budget": cell.get("fy_target"),
                "Current (FINAL YTD)": cell.get("current_march"),
                "Current %": current_pct,
                "Simulation Projected Number": cell.get("march_amount"),
                "Simulation Achievement %": scenario_pct,
                "Change vs Current": change_pts,
                "Required Remaining-Month RR": cell.get("scen_rr"),
            })
    formats = {
        "Asset Class": "txt", "Channel": "txt",
        "FY27 Budget": "cr", "Current (FINAL YTD)": "cr",
        "Current %": "pct", "Simulation Projected Number": "cr",
        "Simulation Achievement %": "pct", "Change vs Current": "pts",
        "Required Remaining-Month RR": "cr",
    }
    return pd.DataFrame(rows), formats


def build_scenario_10_retail_location_detail(
    model: "ScenarioModel",
    sales: str,
) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """T2/T6/T30/B30/EM source and scenario values straight from FINAL."""
    rows: List[Dict[str, Any]] = []
    frame = model.scenario_grid
    if frame is None or frame.empty:
        return pd.DataFrame(), {}
    market_frame = frame.loc[
        frame.get("GridRole", pd.Series("", index=frame.index)).astype(str) == "Market"
    ]
    if market_frame.empty:
        return pd.DataFrame(), {}

    for asset in ASSETS:
        for location in FINAL_MARKET_ROWS:
            subset = market_frame.loc[
                (market_frame["Sales"] == sales)
                & (market_frame["Asset"] == asset)
                & (market_frame["Scenario10MarketType"] == location)
            ]
            if subset.empty:
                continue
            cell = summarize_cells(subset)
            current_pct = _num(cell.get("current_march_pct"))
            scenario_pct = _num(cell.get("march_pct"))
            rows.append({
                "Asset Class": asset,
                "Market Type · Retail + VRM + DHNI": location,
                "FY27 Budget": cell.get("fy_target"),
                "Current (FINAL YTD)": cell.get("current_march"),
                "Current %": current_pct,
                "Simulation Projected Number": cell.get("march_amount"),
                "Simulation Achievement %": scenario_pct,
                "Change vs Current": (
                    None if current_pct is None or scenario_pct is None
                    else scenario_pct - current_pct
                ),
            })
    formats = {
        "Asset Class": "txt", "Market Type · Retail + VRM + DHNI": "txt",
        "FY27 Budget": "cr", "Current (FINAL YTD)": "cr",
        "Current %": "pct", "Simulation Projected Number": "cr",
        "Simulation Achievement %": "pct", "Change vs Current": "pts",
    }
    return pd.DataFrame(rows), formats


# =============================================================================
# 8. SCENARIO MODEL - ONE INTERFACE FOR EVERY VIEW
# =============================================================================
 
class ScenarioModel:
    """Evaluates the selected scenario for any slice of the business."""
 
    def __init__(self, scenario_id: int, grid: pd.DataFrame, params: Dict[str, Any]):
        self.scenario_id = scenario_id
        self.meta = SCENARIOS[scenario_id]
        self.grid = grid
        self.params = params
        self.multipliers = scenario_multipliers(grid, scenario_id, params)
        self._cache: Dict[Tuple, Dict[str, Any]] = {}
        if scenario_id == 7:
            self.scenario_grid = None
        elif scenario_id == 8:
            self.scenario_grid = calculate_scenario_8_grid(grid, params)
        elif scenario_id == 9:
            self.scenario_grid = calculate_scenario_9_grid(grid, params)
        elif scenario_id == 10:
            self.scenario_grid = calculate_scenario_10_grid(grid, params)
        else:
            self.scenario_grid = SCENARIO_FUNCTIONS[scenario_id](grid, params)
 
    # -- core accessor --------------------------------------------------------
    def cell(
        self,
        sales: str,
        asset: Optional[str] = None,
        vertical: Optional[str] = None,
        segment: Optional[str] = None,
        channel: Optional[str] = None,
    ) -> Dict[str, Any]:
        key = (sales, asset, vertical, segment, channel)
        if key in self._cache:
            return self._cache[key]
 
        if self.scenario_id == 7:
            subset = filter_grid(self.grid, sales=sales, asset=asset,
                                 vertical=vertical, segment=segment, channel=channel)
            cell = calculate_scenario_7(
                subset["fy_target"].sum(),
                subset["ytd_target"].sum(),
                subset["ytd_ach"].sum(),
                self.params,
            )
        else:
            frame = self.scenario_grid
            mask = frame["Sales"] == sales
            if self.scenario_id == 10 and "GridRole" in frame.columns:
                # Main Scenario-10 totals use FINAL channel rows only, plus the
                # tiny published-Overall rounding adjustment. Market rows are an
                # alternate cut and must never be double counted here.
                mask &= frame["GridRole"].astype(str).isin(["Channel", "Adjustment"])
            if asset is not None:
                mask &= frame["Asset"] == asset
            if vertical is not None:
                mask &= frame["Vertical"] == vertical
            if segment is not None:
                mask &= frame["Segment"] == segment
            if channel is not None and "Channel" in frame.columns:
                mask &= frame["Channel"] == channel
            cell = summarize_cells(frame.loc[mask])
 
        self._cache[key] = cell
        return cell
 
    # -- convenience views ----------------------------------------------------
    def assets(self, sales: str, **filters: Any) -> Dict[str, Dict[str, Any]]:
        return {asset: self.cell(sales, asset=asset, **filters) for asset in ASSETS}
 
    def baseline(self, sales: str, **filters: Any) -> Dict[str, Any]:
        if self.scenario_id == 10:
            frame = self.grid
            if frame is None or frame.empty:
                return _scenario10_direct_stats(0.0, 0.0)
            mask = frame["Sales"] == sales
            if "GridRole" in frame.columns:
                mask &= frame["GridRole"].astype(str).isin(["Channel", "Adjustment"])
            if filters.get("asset") is not None:
                mask &= frame["Asset"] == filters.get("asset")
            if filters.get("vertical") is not None:
                mask &= frame["Vertical"] == filters.get("vertical")
            if filters.get("segment") is not None:
                mask &= frame["Segment"] == filters.get("segment")
            if filters.get("channel") is not None and "Channel" in frame.columns:
                mask &= frame["Channel"] == filters.get("channel")
            subset = frame.loc[mask]
            return _scenario10_direct_stats(
                subset["fy_target"].sum() if not subset.empty else 0.0,
                subset["ytd_ach"].sum() if not subset.empty else 0.0,
            )
        return summarize_current(self.grid, sales=sales, **filters)
 
    def implied_milestones(self, sales: str) -> Dict[str, Optional[float]]:
        return {
            asset: _multiplier_for(self.multipliers, sales, asset, "*")
            for asset in ASSETS
        }
 
    def available_segments(self) -> List[str]:
        present = set(self.grid["Segment"].unique())
        return [s for s in SEGMENT_ORDER if s in present]
 
    def available_verticals(self) -> List[str]:
        present = set(self.grid["Vertical"].unique())
        return [v for v in VERTICALS if v in present]
 
    def available_channels(self) -> List[str]:
        if "Channel" not in self.grid.columns:
            return []
        present = set(self.grid["Channel"].unique())
        return [c for c in CHANNELS if c in present]
 
 
# =============================================================================
# 9. REVENUE ENGINE
# =============================================================================
 
def calculate_revenue(cells_by_asset: Dict[str, Dict[str, Any]], field: str) -> Dict[str, Any]:
    """Asset-class revenue from a set of scenario cells. No blended rate."""
    by_asset: Dict[str, Optional[float]] = {}
    sales_by_asset: Dict[str, Optional[float]] = {}
    total = 0.0
    for asset in ASSETS:
        amount = _num(cells_by_asset.get(asset, {}).get(field))
        sales_by_asset[asset] = amount
        revenue = None if amount is None else amount * REVENUE_RATE[asset]
        by_asset[asset] = revenue
        total += 0.0 if revenue is None else revenue
    return {"by_asset": by_asset, "sales_by_asset": sales_by_asset, "total": total}
 
 
def calculate_baseline_revenue(cells_by_asset: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    """Baseline revenue on the current-run-rate March projection (scenario independent)."""
    return calculate_revenue(cells_by_asset, "current_march")
 
 
def calculate_incremental_revenue(
    scenario_revenue: Dict[str, Any],
    baseline_revenue: Dict[str, Any],
) -> Dict[str, Any]:
    incremental = {
        asset: (
            None
            if scenario_revenue["by_asset"].get(asset) is None
            or baseline_revenue["by_asset"].get(asset) is None
            else scenario_revenue["by_asset"][asset] - baseline_revenue["by_asset"][asset]
        )
        for asset in ASSETS
    }
    total = scenario_revenue["total"] - baseline_revenue["total"]
    uplift = safe_div(scenario_revenue["total"], baseline_revenue["total"])
    return {
        "by_asset": incremental,
        "total": total,
        "uplift_pct": None if uplift is None else uplift - 1.0,
        "contribution": {
            asset: safe_div(scenario_revenue["by_asset"].get(asset), scenario_revenue["total"])
            for asset in ASSETS
        },
    }
 
 
def revenue_bundle(model: ScenarioModel, basis: str, **filters: Any) -> Dict[str, Any]:
    """Baseline / scenario / incremental revenue for any slice of the business."""
    cells = model.assets(basis, **filters)
    baseline = calculate_baseline_revenue(cells)
    scenario = calculate_revenue(cells, "march_amount")
    january = calculate_revenue(cells, "jan_amount")
    incremental = calculate_incremental_revenue(scenario, baseline)
    return {
        "cells": cells,
        "baseline": baseline,
        "scenario": scenario,
        "january": january,
        "incremental": incremental,
    }
 
 
# =============================================================================
# 10. FINAL <-> MODEL BRIDGE
# =============================================================================
 
def _model_metric_baseline(model: "ScenarioModel", sales: str, label: str) -> Dict[str, Any]:
    if label == "Overall":
        return model.baseline(sales)
    if label in ASSETS:
        return model.baseline(sales, asset=label)
    if label in VERTICALS:
        return model.baseline(sales, vertical=label)
    if model.scenario_id == 10 and label == "Insti":
        return model.baseline(sales, vertical="Insti")
    # Digital is intentionally outside Scenario 10.
    return {}
 
 
def _model_metric_cell(model: "ScenarioModel", sales: str, label: str) -> Optional[Dict[str, Any]]:
    if label == "Overall":
        return model.cell(sales)
    if label in ASSETS:
        return model.cell(sales, asset=label)
    if label in VERTICALS:
        return model.cell(sales, vertical=label)
    if model.scenario_id == 10 and label == "Insti":
        return model.cell(sales, vertical="Insti")
    # Digital stays visible as FINAL current data but has no Scenario-10 outcome.
    return None
 
 
def final_sales_metrics(
    final_metrics: Dict[str, Any],
    model: "ScenarioModel",
    sales: str,
) -> pd.DataFrame:
    """
    Use FINAL Target/YTD as the visible source of truth.
    Fill modelled Overall/asset/vertical rows from the scenario data only if
    the FINAL parser could not locate them.
    """
    parsed = final_metrics.get(sales)
    if isinstance(parsed, pd.DataFrame) and not parsed.empty:
        frame = parsed.copy()
    else:
        frame = pd.DataFrame()
 
    months_done = int(final_metrics.get("months_done", MONTHS_COMPLETED))
    needed = ["Overall", *ASSETS, *VERTICALS]
 
    fallback_rows: List[Dict[str, Any]] = []
    for label in needed:
        if not frame.empty and label in frame.index:
            continue
        base = _model_metric_baseline(model, sales, label)
        if not base:
            continue
        fallback_rows.append(
            {
                "Metric": label,
                "FY27 Target": base.get("fy_target"),
                "YTD": base.get("ytd_ach"),
            }
        )
 
    if fallback_rows:
        fallback = _augment_final_runrate(pd.DataFrame(fallback_rows), months_done).set_index("Metric")
        if frame.empty:
            frame = fallback
        else:
            frame = pd.concat([frame, fallback], axis=0)
 
    if frame.empty:
        return frame
 
    # Preserve management ordering and avoid accidental duplicate rows.
    frame = frame.loc[~frame.index.duplicated(keep="first")].copy()
    order = [label for label in FINAL_METRIC_ROWS if label in frame.index]
    frame = frame.loc[order].copy()

    # Always recalculate run-rate metrics from FINAL Target + YTD before display.
    # This deliberately overrides any older Target/12 workbook/dashboard value.
    # For the current workbook: months_done=3, so required RR = (Target - YTD) / 9.
    frame_for_calc = frame.reset_index()
    first_col = frame_for_calc.columns[0]
    if first_col != "Metric":
        frame_for_calc = frame_for_calc.rename(columns={first_col: "Metric"})
    frame = _augment_final_runrate(frame_for_calc, months_done).set_index("Metric")
    return frame
 
 
def build_final_scenario_comparison(
    final_metrics: Dict[str, Any],
    model: "ScenarioModel",
    sales: str,
) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Scenario comparison expressed on the same Target/YTD/run-rate metrics as FINAL."""
    current = final_sales_metrics(final_metrics, model, sales)
    rows: List[Dict[str, Any]] = []
 
    for label in current.index.tolist():
        source = current.loc[label]
        cell = _model_metric_cell(model, sales, label)
 
        final_target = _num(source.get("FY27 Target"))
        current_rr = _num(source.get("Current RR"))
        if model.scenario_id == 10:
            # Scenario 10's CURRENT column is exactly FINAL YTD/current. It is
            # not the annualised Apr-Jun run-rate projection used elsewhere.
            current_projection = _num(source.get("YTD"))
            current_pct = safe_div(current_projection, final_target)
        else:
            current_projection = _num(source.get("Estimated FY @ Current RR"))
            current_pct = _num(source.get("Projected FY %"))
 
        model_base = _model_metric_baseline(model, sales, label)
        model_target = _num(model_base.get("fy_target")) if model_base else None
 
        scenario_pct = _num(cell.get("march_pct")) if cell is not None else None
 
        scenario_amount = None
        scenario_rr = None
        rr_change = None
        delta_pp = None
 
        if cell is not None:
            # Anchor the scenario outcome to the FINAL FY27 target so the
            # management comparison uses one common metric base.
            scenario_amount = (
                final_target * scenario_pct
                if final_target is not None and scenario_pct is not None
                else _num(cell.get("march_amount"))
            )
 
            scenario_rr = _num(cell.get("scen_rr"))
            if (
                scenario_rr is not None
                and final_target is not None
                and model_target is not None
                and model_target != 0
            ):
                scenario_rr = scenario_rr * final_target / model_target
 
            if current_rr is not None and scenario_rr is not None and current_rr != 0:
                rr_change = scenario_rr / current_rr - 1.0
 
            if current_pct is not None and scenario_pct is not None:
                delta_pp = scenario_pct - current_pct
 
        rows.append(
            {
                "Metric": label,
                "FY27 Target": final_target,
                "YTD": _num(source.get("YTD")),
                "Current RR": current_rr,
                "Required RR to Target": _num(source.get("Required RR to Target")),
                "Current FY Estimate": current_projection,
                "Current Projected %": current_pct,
                "Scenario / Required RR": scenario_rr,
                "Run Rate Change %": rr_change,
                "Scenario March Estimate": scenario_amount,
                "Scenario March %": scenario_pct,
                "Scenario \u0394 pp": delta_pp,
            }
        )
 
    frame = pd.DataFrame(rows)
    formats = {
        "Metric": "txt",
        "FY27 Target": "cr",
        "YTD": "cr",
        "Current RR": "cr",
        "Required RR to Target": "cr",
        "Current FY Estimate": "cr",
        "Current Projected %": "pct",
        "Scenario / Required RR": "cr",
        "Run Rate Change %": "pct_signed",
        "Scenario March Estimate": "cr",
        "Scenario March %": "pct",
        "Scenario Δ pp": "pts",
    }

    if model.scenario_id == 10 and not frame.empty:
        frame = frame.rename(columns={
            "FY27 Target": "FY27 Budget",
            "Current FY Estimate": "Current (FINAL YTD)",
            "Current Projected %": "Current %",
            "Scenario / Required RR": "Required Remaining-Month RR",
            "Scenario March Estimate": "Simulation Projected Number",
            "Scenario March %": "Simulation Achievement %",
            "Scenario Δ pp": "Change vs Current",
        })
        formats = {
            "Metric": "txt",
            "FY27 Budget": "cr",
            "YTD": "cr",
            "Current RR": "cr",
            "Required RR to Target": "cr",
            "Current (FINAL YTD)": "cr",
            "Current %": "pct",
            "Required Remaining-Month RR": "cr",
            "Run Rate Change %": "pct_signed",
            "Simulation Projected Number": "cr",
            "Simulation Achievement %": "pct",
            "Change vs Current": "pts",
        }
    return frame, formats
 
 
# =============================================================================
# 11. TABLE BUILDERS
# =============================================================================
 
def build_current_overview(model: ScenarioModel) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Baseline table - identical for every scenario."""
    rows = []
    for sales in SALES_TYPES:
        base = model.baseline(sales)
        rows.append({
            "Sales": SALES_LABEL[sales],
            "FY Target": base["fy_target"],
            "YTD June Target": base["ytd_target"],
            "YTD June Achievement": base["ytd_ach"],
            "Target Achieved %": base["ytd_ach_pct"],
            "FY Target Completed %": base["fy_completed_pct"],
            "Current Run Rate": base["current_rr"],
            "Current March Projection": base["current_march"],
            "Current March Projection %": base["current_march_pct"],
        })
    formats = {
        "Sales": "txt", "FY Target": "cr", "YTD June Target": "cr",
        "YTD June Achievement": "cr", "Target Achieved %": "pct",
        "FY Target Completed %": "pct", "Current Run Rate": "cr",
        "Current March Projection": "cr", "Current March Projection %": "pct",
    }
    return pd.DataFrame(rows), formats
 
 
def summarize_scenario(model: ScenarioModel, sales: str) -> Dict[str, Any]:
    """Headline current-versus-scenario numbers for one sales basis."""
    cell = model.cell(sales)
    return {
        "Sales": SALES_LABEL[sales],
        "Current Run Rate": cell["current_rr"],
        "Scenario Run Rate": cell["scen_rr"],
        "Run Rate Change %": cell["rr_change_pct"],
        "Jan Achievement": cell["jan_amount"],
        "Jan Achievement %": cell["jan_pct"],
        "Feb-Mar Run Rate": cell["feb_mar_rr"],
        "Current March Projection %": cell["current_march_pct"],
        "Scenario March Achievement": cell["march_amount"],
        "Scenario March Achievement %": cell["march_pct"],
        "Incremental Sales": cell["incremental_sales"],
    }
 
 
def build_comparison(model: ScenarioModel) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Current versus selected scenario, for Gross Sales and Net Sales."""
    rows = [summarize_scenario(model, sales) for sales in SALES_TYPES]
    formats = {
        "Sales": "txt", "Current Run Rate": "cr", "Scenario Run Rate": "cr",
        "Run Rate Change %": "pct_signed", "Jan Achievement": "cr",
        "Jan Achievement %": "pct", "Feb-Mar Run Rate": "cr",
        "Current March Projection %": "pct", "Scenario March Achievement": "cr",
        "Scenario March Achievement %": "pct", "Incremental Sales": "cr_signed",
    }
    return pd.DataFrame(rows), formats
 
 
def build_revenue_impact(model: ScenarioModel, basis: str) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Current and expected revenue by asset class - never a blended rate."""
    bundle = revenue_bundle(model, basis)
    rows = []
    for asset in ASSETS:
        rows.append({
            "Asset Class": asset,
            "Current Revenue": bundle["baseline"]["by_asset"][asset],
            "Expected Revenue": bundle["scenario"]["by_asset"][asset],
        })
    rows.append({
        "Asset Class": "Total",
        "Current Revenue": bundle["baseline"]["total"],
        "Expected Revenue": bundle["scenario"]["total"],
    })
    formats = {
        "Asset Class": "txt",
        "Current Revenue": "cr1",
        "Expected Revenue": "cr1",
    }
    return pd.DataFrame(rows), formats
 
 
def build_vertical_summary(model: ScenarioModel) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Retail, DHNI and VRM, for Gross Sales and Net Sales."""
    rows = []
    for vertical in model.available_verticals():
        for sales in SALES_TYPES:
            cell = model.cell(sales, vertical=vertical)
            bundle = revenue_bundle(model, sales, vertical=vertical)
            rows.append({
                "Vertical": vertical,
                "Sales": SALES_LABEL[sales],
                "FY Target": cell["fy_target"],
                "YTD Achievement": cell["ytd_ach"],
                "Target Achieved %": cell["ytd_ach_pct"],
                "Current Run Rate": cell["current_rr"],
                "Scenario Run Rate": cell["scen_rr"],
                "Run Rate Change %": cell["rr_change_pct"],
                "Current March Projection %": cell["current_march_pct"],
                "Scenario Milestone %": cell["milestone_pct"],
                "Scenario March Projection %": cell["march_pct"],
                "Scenario Revenue": bundle["scenario"]["total"],
                "Incremental Revenue": bundle["incremental"]["total"],
            })
    formats = {
        "Vertical": "txt", "Sales": "txt", "FY Target": "cr", "YTD Achievement": "cr",
        "Target Achieved %": "pct", "Current Run Rate": "cr", "Scenario Run Rate": "cr",
        "Run Rate Change %": "pct_signed", "Current March Projection %": "pct",
        "Scenario Milestone %": "pct", "Scenario March Projection %": "pct",
        "Scenario Revenue": "cr1", "Incremental Revenue": "cr1_signed",
    }
    return pd.DataFrame(rows), formats
 
 
def build_asset_breakdown(model: ScenarioModel, sales: str) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Equity / Debt / Liquid split within each vertical."""
    rows = []
    include_dip = model.scenario_id in (3, 7)
    for vertical in model.available_verticals():
        for asset in ASSETS:
            cell = model.cell(sales, asset=asset, vertical=vertical)
            revenue = _z(cell["march_amount"]) * REVENUE_RATE[asset]
            baseline_revenue = _z(cell["current_march"]) * REVENUE_RATE[asset]
            row = {
                "Vertical": vertical,
                "Asset": asset,
                "FY Target": cell["fy_target"],
                "YTD Achievement": cell["ytd_ach"],
                "Target Achieved %": cell["ytd_ach_pct"],
                "Current Run Rate": cell["current_rr"],
                "Scenario Run Rate": cell["scen_rr"],
                "Run Rate Change %": cell["rr_change_pct"],
            }
            if include_dip:
                row["Feb-Mar Run Rate"] = cell["feb_mar_rr"]
            row.update({
                "Current March Projection %": cell["current_march_pct"],
                "Scenario Milestone %": cell["milestone_pct"],
                "Scenario March Projection %": cell["march_pct"],
                "Scenario Revenue": revenue,
                "Incremental Revenue": revenue - baseline_revenue,
            })
            rows.append(row)
    formats = {
        "Vertical": "txt", "Asset": "txt", "FY Target": "cr", "YTD Achievement": "cr",
        "Target Achieved %": "pct", "Current Run Rate": "cr", "Scenario Run Rate": "cr",
        "Run Rate Change %": "pct_signed", "Feb-Mar Run Rate": "cr",
        "Current March Projection %": "pct", "Scenario Milestone %": "pct",
        "Scenario March Projection %": "pct", "Scenario Revenue": "cr1",
        "Incremental Revenue": "cr1_signed",
    }
    return pd.DataFrame(rows), formats
 
 
def build_segment_scenario_analysis(
    model: ScenarioModel, sales: str
) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Scenario 6 - differentiated performance by business segment."""
    rows = []
    for segment in model.available_segments():
        cell = model.cell(sales, segment=segment)
        rows.append({
            "Segment": segment,
            "FY Target": cell["fy_target"],
            "YTD Achievement": cell["ytd_ach"],
            "Current Run Rate": cell["current_rr"],
            "Current March Projection": cell["current_march"],
            "Scenario Achievement %": cell["milestone_pct"],
            "Scenario Target Amount": cell["march_required"],
            "Scenario Required Run Rate": cell["scen_rr"],
            "Run Rate Uplift %": cell["rr_change_pct"],
            "Incremental Amount": cell["incremental_sales"],
        })
    overall = model.cell(sales)
    rows.append({
        "Segment": "Overall",
        "FY Target": overall["fy_target"],
        "YTD Achievement": overall["ytd_ach"],
        "Current Run Rate": overall["current_rr"],
        "Current March Projection": overall["current_march"],
        "Scenario Achievement %": overall["march_pct"],
        "Scenario Target Amount": overall["march_required"],
        "Scenario Required Run Rate": overall["scen_rr"],
        "Run Rate Uplift %": overall["rr_change_pct"],
        "Incremental Amount": overall["incremental_sales"],
    })
    formats = {
        "Segment": "txt", "FY Target": "cr", "YTD Achievement": "cr",
        "Current Run Rate": "cr", "Current March Projection": "cr",
        "Scenario Achievement %": "pct", "Scenario Target Amount": "cr",
        "Scenario Required Run Rate": "cr", "Run Rate Uplift %": "pct_signed",
        "Incremental Amount": "cr_signed",
    }
    return pd.DataFrame(rows), formats
 
 
def build_momentum_analysis(cell: Dict[str, Any]) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Scenario 7 - month-by-month momentum trajectory."""
    trajectory = cell.get("trajectory") or []
    current_rr = _num(cell.get("current_rr"))
    growth = cell.get("momentum_g")
    leakage = _z(cell.get("leakage"))
    fy_target = cell.get("fy_target")
    cumulative = _z(cell.get("ytd_ach"))
 
    rows = []
    previous = current_rr
    for position, month in enumerate(FUTURE_MONTHS):
        run_rate = trajectory[position] if position < len(trajectory) else None
        cumulative += _z(run_rate)
        if position < MONTHS_JUL_JAN:
            mom = growth if growth is not None else safe_div(run_rate, previous)
            if growth is None and mom is not None:
                mom = mom - 1.0
            phase = "Momentum build-up"
        else:
            mom = -leakage
            phase = "Feb-Mar leakage"
        rows.append({
            "Month": month,
            "Phase": phase,
            "Current Run Rate": current_rr,
            "Required Scenario Run Rate": run_rate,
            "MoM Growth": mom,
            "Cumulative Achievement": cumulative,
            "Achievement %": safe_div(cumulative, fy_target),
        })
        previous = run_rate
    formats = {
        "Month": "txt", "Phase": "txt", "Current Run Rate": "cr",
        "Required Scenario Run Rate": "cr", "MoM Growth": "pct_signed",
        "Cumulative Achievement": "cr", "Achievement %": "pct",
    }
    return pd.DataFrame(rows), formats
 
 
def build_momentum_by_group(
    model: ScenarioModel, sales: str, dimension: str
) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Scenario 7 applied independently to asset classes or to verticals."""
    if dimension == "asset":
        keys = [("Asset", asset, {"asset": asset}) for asset in ASSETS]
        first_column = "Asset"
    else:
        keys = [("Vertical", v, {"vertical": v}) for v in model.available_verticals()]
        first_column = "Vertical"
 
    rows = []
    for _, name, filters in keys:
        cell = model.cell(sales, **filters)
        rows.append({
            first_column: name,
            "Current Run Rate": cell["current_rr"],
            "Required MoM Momentum": cell["momentum_g"],
            "January Achievement": cell["jan_amount"],
            "January Achievement %": cell["jan_pct"],
            "Feb-Mar Leakage": cell.get("leakage"),
            "March Achievement": cell["march_amount"],
            "March Achievement %": cell["march_pct"],
            "Headroom / Shortfall": cell["headroom_amt"],
        })
    formats = {
        first_column: "txt", "Current Run Rate": "cr",
        "Required MoM Momentum": "pct_signed", "January Achievement": "cr",
        "January Achievement %": "pct", "Feb-Mar Leakage": "pct",
        "March Achievement": "cr", "March Achievement %": "pct",
        "Headroom / Shortfall": "cr_signed",
    }
    return pd.DataFrame(rows), formats
 
 
def build_monthly_revenue(model: ScenarioModel, basis: str) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Scenario 7 - monthly revenue implied by the momentum trajectory."""
    trajectories = {
        asset: (model.cell(basis, asset=asset).get("trajectory") or [])
        for asset in ASSETS
    }
    rows = []
    for position, month in enumerate(FUTURE_MONTHS):
        row = {"Month": month}
        total = 0.0
        for asset in ASSETS:
            series = trajectories[asset]
            sales_amount = series[position] if position < len(series) else 0.0
            revenue = _z(sales_amount) * REVENUE_RATE[asset]
            row[f"{asset} Revenue"] = revenue
            total += revenue
        row["Total Revenue"] = total
        rows.append(row)
    formats = {"Month": "txt", "Total Revenue": "cr1"}
    for asset in ASSETS:
        formats[f"{asset} Revenue"] = "cr1"
    return pd.DataFrame(rows), formats
 
 
def build_leakage_sensitivity(
    model: ScenarioModel, basis: str
) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Scenario 7 - how the March outcome moves with the leakage assumption."""
    subset = filter_grid(model.grid, sales=basis)
    fy_target = subset["fy_target"].sum()
    ytd_target = subset["ytd_target"].sum()
    ytd_ach = subset["ytd_ach"].sum()
    rows = []
    for leakage in (0.0, 0.10, 0.20, 0.30):
        params = dict(model.params)
        params["leakage"] = leakage
        cell = calculate_scenario_7(fy_target, ytd_target, ytd_ach, params)
        rows.append({
            "Feb-Mar Leakage": leakage,
            "Required MoM Momentum": cell["momentum_g"],
            "January Achievement %": cell["jan_pct"],
            "March Achievement %": cell["march_pct"],
            "Headroom / Shortfall": cell["headroom_amt"],
        })
    formats = {
        "Feb-Mar Leakage": "pct", "Required MoM Momentum": "pct_signed",
        "January Achievement %": "pct", "March Achievement %": "pct",
        "Headroom / Shortfall": "cr_signed",
    }
    return pd.DataFrame(rows), formats
 
 
def build_channel_scenario_analysis(
    model: "ScenarioModel", basis: str
) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Scenario 8/9 - one row per planning channel."""
    rows = []
    frame = model.scenario_grid
    if frame is None or "Channel" not in getattr(frame, "columns", []):
        return pd.DataFrame(), {}
    for channel in CHANNELS:
        subset = frame[(frame["Sales"] == basis) & (frame["Channel"] == channel)]
        if subset.empty:
            continue
        cell = summarize_cells(subset)
        growth, jan_target, mar_target, leakage = _s8_channel_params(model.params, channel)
        if model.scenario_id == 9 and "optimized_growth" in subset.columns:
            solved = subset["optimized_growth"].dropna()
            if not solved.empty:
                growth = float(solved.max())
        rows.append({
            "Channel": channel, "MoM Growth": growth,
            "Jan 2027 Target": jan_target, "Jan Achievement": cell.get("jan_pct"),
            "Jan Gap / Headroom": cell.get("jan_buffer"),
            "Mar 2027 Target": mar_target, "Mar Achievement": cell.get("march_pct"),
            "Mar Gap / Headroom": cell.get("headroom_amt"),
            "Current Run Rate": cell.get("current_rr"),
            "Jan Exit Run Rate": cell.get("scen_rr"),
            "March Incremental Sales": cell.get("incremental_sales"),
        })
    formats = {
        "Channel": "txt", "MoM Growth": "pct_signed", "Jan 2027 Target": "pct",
        "Jan Achievement": "pct", "Jan Gap / Headroom": "cr_signed",
        "Mar 2027 Target": "pct", "Mar Achievement": "pct", "Mar Gap / Headroom": "cr_signed",
        "Current Run Rate": "cr", "Jan Exit Run Rate": "cr",
        "March Incremental Sales": "cr_signed",
    }
    return pd.DataFrame(rows), formats
 
 
def build_scenario_guide(model: ScenarioModel, basis: str) -> pd.DataFrame:
    rows = []
    for scenario_id in SCENARIO_ORDER:
        meta = SCENARIOS[scenario_id]
        rows.append({
            "Scenario": meta["label"],
            "Description": meta["explanation"],
            "Milestone": meta["milestone"],
            "Selected": "Yes" if scenario_id == model.scenario_id else "",
        })
    rows.append({
        "Scenario": "Revenue methodology",
        "Description": (
            "Revenue is estimated at asset-class level using 60 bps for Equity, 20 bps for Debt "
            f"and 10 bps for Liquid, applied to {SALES_LABEL[basis]} only so that Gross Sales and "
            "Net Sales revenue are never double counted."
        ),
        "Milestone": "",
        "Selected": "",
    })
    rows.append({
        "Scenario": "Timeline assumption",
        "Description": (
            "April, May and June are complete. Three months completed, nine months remaining "
            "(July-January is seven months, February-March is two months). The current run rate "
            "is YTD achievement divided by three."
        ),
        "Milestone": "",
        "Selected": "",
    })
    if model.scenario_id == 3:
        rows.append({
            "Scenario": "Scenario 3 setting",
            "Description": f"Feb-Mar run-rate dip: {fmt_pct(model.params.get('dip', S3_DEFAULT_DIP))}",
            "Milestone": "", "Selected": "",
        })
    if model.scenario_id == 7:
        rows.append({
            "Scenario": "Scenario 7 settings",
            "Description": (
                f"January target: {fmt_pct(model.params.get('jan_target', S7_DEFAULT_JAN_TARGET))} · "
                f"March target: {fmt_pct(model.params.get('mar_target', S7_DEFAULT_MAR_TARGET))} · "
                f"Feb-Mar leakage: {fmt_pct(model.params.get('leakage', S7_DEFAULT_LEAKAGE))}"
            ),
            "Milestone": "", "Selected": "",
        })
    if model.scenario_id == 10:
        rows.append({
            "Scenario": "Scenario 10 settings",
            "Description": (
                "Retail, DHNI, VRM and Insti are editable inside Equity, Debt and Liquid; Digital is excluded. "
                "Current and FY27 Budget are read directly from FINAL. Projected Number = Simulation % × FY27 Budget. "
                "T2/T6/T30/B30/EM are also read from FINAL when that market matrix is present."
            ),
            "Milestone": "March 2027",
            "Selected": "",
        })
    return pd.DataFrame(rows)
 
 
def scenario_default_params(scenario_id: int) -> Dict[str, Any]:
    """Editable defaults for the selected scenario."""
    params: Dict[str, Any] = {
        "dip": S3_DEFAULT_DIP,
        "jan_target": S7_DEFAULT_JAN_TARGET,
        "mar_target": S7_DEFAULT_MAR_TARGET,
        "leakage": S7_DEFAULT_LEAKAGE,
        "channel_growth": dict(S8_DEFAULT_GROWTH),
        "channel_jan_target": dict(S8_DEFAULT_JAN_TARGET),
        "channel_mar_target": dict(S8_DEFAULT_MAR_TARGET),
        "optimizer_target": 1.20,
        "channel_mapping": {},
        "asset_vertical_targets": {},
        "retail_location_targets": {},
        "market_location_factors": {},
    }
    if scenario_id == 1:
        params["runrate_uplift"] = S1_RUNRATE_UPLIFT
    elif scenario_id == 2:
        params["overall_target"] = S2_OVERALL_TARGET
        params["equity_target"] = S2_EQUITY_TARGET
    elif scenario_id == 3:
        params["target_pct"] = S3_TARGET
    elif scenario_id == 4:
        params["target_pct"] = S4_TARGET
    elif scenario_id == 5:
        params["overall_target"] = S5_OVERALL_TARGET
        params["equity_target"] = S5_EQUITY_TARGET
    elif scenario_id == 6:
        params["segment_targets"] = dict(S6_SEGMENT_TARGETS)
    return params
 
 
def build_all_scenario_matrix(
    filtered_grid: pd.DataFrame,
    selected_scenario_id: int,
    selected_params: Dict[str, Any],
    sales_key: str,
    asset: str = "All",
) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Calculate Scenarios 1-10 on the same selected cut."""
    rows: List[Dict[str, Any]] = []
    formats = {
        "Scenario": "txt", "Strategy": "txt", "FY Target": "cr",
        "Current YTD": "cr", "Current March Projection": "cr",
        "Scenario March Estimate": "cr", "Scenario March %": "pct",
        "Required Run Rate": "cr", "Run Rate Change %": "pct_signed",
        "Incremental Sales": "cr_signed", "Headroom / Gap": "cr_signed",
    }
    if filtered_grid.empty:
        return pd.DataFrame(), formats
 
    for sid in SCENARIO_ORDER:
        params = scenario_default_params(sid)
        if sid == selected_scenario_id:
            # The live controls are authoritative for the selected scenario.
            params.update(selected_params)
        try:
            model = ScenarioModel(sid, filtered_grid, params)
            cell = model.cell(sales_key, asset=None if asset == "All" else asset)
            live_strategy = _active_scenario_copy(model)["name"] if sid == selected_scenario_id else SCENARIOS[sid]["name"]
            rows.append({
                "Scenario": f"{sid:02d} · {SCENARIOS[sid]['short']}",
                "Strategy": live_strategy,
                "FY Target": cell.get("fy_target"),
                "Current YTD": cell.get("ytd_ach"),
                "Current March Projection": cell.get("current_march"),
                "Scenario March Estimate": cell.get("march_amount"),
                "Scenario March %": cell.get("march_pct"),
                "Required Run Rate": cell.get("scen_rr"),
                "Run Rate Change %": cell.get("rr_change_pct"),
                "Incremental Sales": cell.get("incremental_sales"),
                "Headroom / Gap": cell.get("headroom_amt"),
            })
        except Exception:  # pragma: no cover - one scenario must never break the view
            rows.append({
                "Scenario": f"{sid:02d} · {SCENARIOS[sid]['short']}",
                "Strategy": SCENARIOS[sid]["name"],
            })
    return pd.DataFrame(rows), formats
 
 
# =============================================================================
# 12. EXCEL EXPORT
# =============================================================================
 
def _round_frame(frame: pd.DataFrame) -> pd.DataFrame:
    out = frame.copy()
    for column in out.columns:
        if pd.api.types.is_numeric_dtype(out[column]):
            out[column] = out[column].astype(float).round(4)
    return out
 
 
def make_export_excel(model: ScenarioModel, basis: str) -> bytes:
    """Build the management export workbook for the selected scenario."""
    sheets: List[Tuple[str, pd.DataFrame]] = []
    sheets.append(("Scenario Guide", build_scenario_guide(model, basis)))
    sheets.append(("Current Baseline", build_current_overview(model)[0]))
    sheets.append(("Current vs Scenario", build_comparison(model)[0]))
    sheets.append(("Revenue Impact", build_revenue_impact(model, basis)[0]))
    sheets.append(("Retail-DHNI-VRM Summary", build_vertical_summary(model)[0]))
    sheets.append(("Gross Sales Breakdown", build_asset_breakdown(model, "GS")[0]))
    sheets.append(("Net Sales Breakdown", build_asset_breakdown(model, "NS")[0]))
 
    segment_model = model if model.scenario_id == 6 else ScenarioModel(6, model.grid, model.params)
    segment_frames = []
    for sales in SALES_TYPES:
        frame = build_segment_scenario_analysis(segment_model, sales)[0]
        frame.insert(0, "Sales", SALES_LABEL[sales])
        segment_frames.append(frame)
    sheets.append(("Scenario 6 Segments", pd.concat(segment_frames, ignore_index=True)))
 
    momentum_model = model if model.scenario_id == 7 else ScenarioModel(7, model.grid, model.params)
    momentum_frames = []
    for sales in SALES_TYPES:
        overall = momentum_model.cell(sales)
        frame = build_momentum_analysis(overall)[0]
        frame.insert(0, "Sales", SALES_LABEL[sales])
        momentum_frames.append(frame)
    for sales in SALES_TYPES:
        for dimension in ("asset", "vertical"):
            frame = build_momentum_by_group(momentum_model, sales, dimension)[0]
            frame.insert(0, "Sales", SALES_LABEL[sales])
            momentum_frames.append(frame)
    sheets.append(("Scenario 7 Momentum", pd.concat(momentum_frames, ignore_index=True)))
    sheets.append(("S7 Monthly Revenue", build_monthly_revenue(momentum_model, basis)[0]))
 
    if model.scenario_id in (8, 9):
        channel_frame = build_channel_scenario_analysis(model, basis)[0]
        if not channel_frame.empty:
            sheets.append((f"Scenario {model.scenario_id} Channels", channel_frame))


    if model.scenario_id == 10:
        for sales in SALES_TYPES:
            asset_summary = build_scenario_10_asset_summary(model, sales)[0]
            channel_detail = build_scenario_10_channel_detail(model, sales)[0]
            location_detail = build_scenario_10_retail_location_detail(model, sales)[0]
            sheets.append((f"S10 {sales} Asset Rollup", asset_summary))
            sheets.append((f"S10 {sales} Asset Channel", channel_detail))
            if not location_detail.empty:
                sheets.append((f"S10 {sales} Retail Mkt", location_detail))
 
    all_scenarios = build_all_scenario_matrix(model.grid, model.scenario_id, model.params, basis)[0]
    if not all_scenarios.empty:
        sheets.append(("All Scenarios", all_scenarios))
 
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for name, frame in sheets:
            _round_frame(frame).to_excel(writer, sheet_name=name[:31], index=False)
        _style_workbook(writer)
    return buffer.getvalue()
 
 
def _style_workbook(writer: Any) -> None:
    """Bold headers, frozen top row and sensible column widths."""
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:  # pragma: no cover - openpyxl always present in the stack
        return
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="12141A")
    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            longest = 0
            letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                longest = max(longest, min(len(value), 60))
            worksheet.column_dimensions[letter].width = max(12, min(longest + 3, 62))
 
 
# =============================================================================
# 13. RM PERFORMANCE ANALYTICS (segmentation & contribution)
# =============================================================================
 
ACHIEVEMENT_BANDS: List[Tuple[str, float, Optional[float]]] = [
    ("100% and above", 1.00, None),
    ("90% - 100%", 0.90, 1.00),
    ("75% - 90%", 0.75, 0.90),
    ("50% - 75%", 0.50, 0.75),
    ("30% - 50%", 0.30, 0.50),
    ("Less than 30%", float("-inf"), 0.30),
]
ACHIEVEMENT_BAND_ORDER: List[str] = [item[0] for item in ACHIEVEMENT_BANDS]
 
 
def achievement_band(value: Any) -> str:
    """Map YTD-target achievement to the management bands used for RMs."""
    ratio = _num(value)
    # Undefined achievement is treated and displayed as 0.
    ratio = 0.0 if ratio is None else ratio
 
    if ratio >= 1.00:
        return "100% and above"
    if ratio >= 0.90:
        return "90% - 100%"
    if ratio >= 0.75:
        return "75% - 90%"
    if ratio >= 0.50:
        return "50% - 75%"
    if ratio >= 0.30:
        return "30% - 50%"
    return "Less than 30%"
 
 
def _rm_identity_columns(records: pd.DataFrame) -> List[str]:
    preferred = [
        "Employee Name", "Emp Code", "ADID", "ZONE", "REGION",
        "EM City", "MKT TYPE", "Type", "Status",
    ]
    return [column for column in preferred if column in records.columns]
 
 
def build_rm_performance_detail(
    records: pd.DataFrame,
    vertical: str,
    sales: str,
    final_target: Optional[float] = None,
) -> pd.DataFrame:
    """
    Build one row per RM.
 
    Banding is based on overall YTD achievement / overall YTD target across
    Equity + Debt + Liquid. Run-rate projection annualises the first three
    completed months.
    """
    subset = records.loc[records["Vertical"] == vertical].copy()
    if subset.empty:
        return pd.DataFrame()
 
    identity = _rm_identity_columns(subset)
    out = subset[identity].copy()
 
    fy_cols = [f"{sales}_{asset}_fy" for asset in ASSETS]
    ytd_target_cols = [f"{sales}_{asset}_ytd_tgt" for asset in ASSETS]
    ach_cols = [f"{sales}_{asset}_ach" for asset in ASSETS]
 
    for columns in (fy_cols, ytd_target_cols, ach_cols):
        for column in columns:
            if column not in subset.columns:
                subset[column] = 0.0
 
    fy_matrix = subset[fy_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    ytd_target_matrix = subset[ytd_target_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    ach_matrix = subset[ach_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
 
    out["FY Target"] = fy_matrix.sum(axis=1)
    out["YTD Target"] = ytd_target_matrix.sum(axis=1)
    out["YTD Achievement"] = ach_matrix.sum(axis=1)
 
    out["YTD Achievement %"] = np.where(
        out["YTD Target"] > 0,
        out["YTD Achievement"] / out["YTD Target"],
        0.0,
    )
 
    out["Achievement Category"] = out["YTD Achievement %"].map(achievement_band)
    out["Current Run Rate"] = out["YTD Achievement"] / max(MONTHS_COMPLETED, 1)
    out["Estimated FY @ Current RR"] = out["Current Run Rate"] * 12.0
    out["Projected FY Achievement %"] = np.where(
        out["FY Target"] > 0,
        out["Estimated FY @ Current RR"] / out["FY Target"],
        0.0,
    )
 
    denominator = _num(final_target)
    if denominator is None or denominator <= 0:
        denominator = _num(out["FY Target"].sum()) or 0.0
 
    out["Contribution to Overall Target %"] = np.where(
        denominator > 0,
        out["Estimated FY @ Current RR"] / denominator,
        0.0,
    )
 
    # Asset-level achieved percentages are useful when drilling into an RM.
    for asset in ASSETS:
        fy = pd.to_numeric(subset[f"{sales}_{asset}_fy"], errors="coerce").fillna(0.0)
        ytd_target = pd.to_numeric(subset[f"{sales}_{asset}_ytd_tgt"], errors="coerce").fillna(0.0)
        ach = pd.to_numeric(subset[f"{sales}_{asset}_ach"], errors="coerce").fillna(0.0)
 
        out[f"{asset} YTD %"] = np.where(ytd_target > 0, ach / ytd_target, 0.0)
        out[f"{asset} Current RR"] = ach / max(MONTHS_COMPLETED, 1)
        out[f"{asset} Projected FY %"] = np.where(
            fy > 0,
            (ach / max(MONTHS_COMPLETED, 1) * 12.0) / fy,
            0.0,
        )
 
    out = out.sort_values(
        ["YTD Achievement %", "YTD Achievement"],
        ascending=[False, False],
    ).reset_index(drop=True)
 
    return out
 
 
def build_category_contribution(
    detail: pd.DataFrame,
    final_target: Optional[float] = None,
) -> pd.DataFrame:
    """
    Aggregate RM bands and quantify how many percentage points each band is
    projected to contribute to the FINAL FY target at the current run rate.
    """
    if detail.empty:
        return pd.DataFrame()
 
    target_denominator = _num(final_target)
    if target_denominator is None or target_denominator <= 0:
        target_denominator = _num(detail["FY Target"].sum()) or 0.0
 
    total_projected = _num(detail["Estimated FY @ Current RR"].sum()) or 0.0
 
    work = detail.copy()
    work["_rm"] = 1
 
    grouped = (
        work.groupby("Achievement Category", dropna=False)
        .agg(
            **{
                "RM Count": ("_rm", "sum"),
                "FY Target": ("FY Target", "sum"),
                "YTD Target": ("YTD Target", "sum"),
                "YTD Achievement": ("YTD Achievement", "sum"),
                "Current Run Rate": ("Current Run Rate", "sum"),
                "Estimated FY @ Current RR": ("Estimated FY @ Current RR", "sum"),
            }
        )
        .reindex(ACHIEVEMENT_BAND_ORDER, fill_value=0)
        .reset_index()
    )
 
    grouped["Current YTD Achievement %"] = np.where(
        grouped["YTD Target"] > 0,
        grouped["YTD Achievement"] / grouped["YTD Target"],
        0.0,
    )
    grouped["Category Projected FY %"] = np.where(
        grouped["FY Target"] > 0,
        grouped["Estimated FY @ Current RR"] / grouped["FY Target"],
        0.0,
    )
    grouped["Contribution to Overall Target %"] = np.where(
        target_denominator > 0,
        grouped["Estimated FY @ Current RR"] / target_denominator,
        0.0,
    )
    grouped["Share of Projected Sales %"] = np.where(
        total_projected != 0,
        grouped["Estimated FY @ Current RR"] / total_projected,
        0.0,
    )
 
    return grouped
 
 
def _final_vertical_target(
    final_metrics: Dict[str, Any],
    sales: str,
    vertical: str,
) -> Optional[float]:
    frame = final_metrics.get(sales)
    if not isinstance(frame, pd.DataFrame) or frame.empty or vertical not in frame.index:
        return None
    return _num(frame.loc[vertical].get("FY27 Target"))
 
 
def _final_vertical_ytd(
    final_metrics: Dict[str, Any],
    sales: str,
    vertical: str,
) -> Optional[float]:
    frame = final_metrics.get(sales)
    if not isinstance(frame, pd.DataFrame) or frame.empty or vertical not in frame.index:
        return None
    return _num(frame.loc[vertical].get("YTD"))
 
 
def make_rm_segmentation_export(
    records: pd.DataFrame,
    final_metrics: Dict[str, Any],
) -> bytes:
    """Downloadable workbook covering every vertical and both sales bases."""
    output = io.BytesIO()
 
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for vertical in VERTICALS:
            for sales in SALES_TYPES:
                final_target = _final_vertical_target(final_metrics, sales, vertical)
                detail = build_rm_performance_detail(records, vertical, sales, final_target)
                contribution = build_category_contribution(detail, final_target)
                stars = (
                    detail.sort_values(
                        ["YTD Achievement %", "YTD Achievement"],
                        ascending=[False, False],
                    ).head(10).copy()
                    if not detail.empty else pd.DataFrame()
                )
 
                prefix = f"{vertical}-{sales}"
                if detail.empty:
                    detail = pd.DataFrame({"Note": [f"No {vertical} records for {SALES_LABEL[sales]}."]})
                if contribution.empty:
                    contribution = pd.DataFrame({"Note": ["No banded contribution available."]})
                if stars.empty:
                    stars = pd.DataFrame({"Note": ["No ranked RMs available."]})
 
                detail.to_excel(writer, sheet_name=f"{prefix}-RM"[:31], index=False)
                contribution.to_excel(writer, sheet_name=f"{prefix}-Bands"[:31], index=False)
                stars.to_excel(writer, sheet_name=f"{prefix}-Stars"[:31], index=False)
 
        try:
            from openpyxl.styles import Font
        except Exception:  # pragma: no cover
            Font = None
 
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            if Font is not None:
                for cell in ws[1]:
                    cell.font = Font(bold=True)
            for cells in ws.columns:
                width = max((len(str(cell.value or "")) for cell in cells[:150]), default=10)
                ws.column_dimensions[cells[0].column_letter].width = min(max(width + 2, 12), 36)
 
    output.seek(0)
    return output.getvalue()
 
 
# =============================================================================
# 14. DESIGN SYSTEM - LIQUID GLASS THEME, COMPONENTS & PLOTLY TEMPLATE
# =============================================================================
 
INK = "#F5F5F7"
INK_SOFT = "#A1A4AD"
INK_MUTED = "#6D7079"
GOLD = "#D8B76A"
GOLD_SOFT = "#C9AA65"
GREEN = "#63D99A"
RED = "#FF6B6B"
GRID_LINE = "rgba(255,255,255,0.07)"
FONT_STACK = 'Inter, -apple-system, BlinkMacSystemFont, "SF Pro Display", "Helvetica Neue", sans-serif'
 
GLASS_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
 
:root {
    --background: #050608;
    --glass-1: rgba(255,255,255,0.075);
    --glass-2: rgba(255,255,255,0.045);
    --glass-3: rgba(255,255,255,0.025);
    --border: rgba(255,255,255,0.11);
    --border-strong: rgba(255,255,255,0.18);
    --text: #F5F5F7;
    --secondary: #A1A4AD;
    --tertiary: #6D7079;
    --gold: #D8B76A;
    --gold-soft: #C9AA65;
    --gold-glow: rgba(216,183,106,0.16);
    --green: #63D99A;
    --red: #FF6B6B;
    --radius-lg: 24px;
    --radius-md: 18px;
    --radius-sm: 12px;
}
 
/* ---------- canvas ---------- */
.stApp, [data-testid="stAppViewContainer"] {
    background:
        radial-gradient(circle at 10% 0%, rgba(255,255,255,0.035), transparent 30%),
        radial-gradient(circle at 90% 8%, rgba(214,179,106,0.055), transparent 26%),
        radial-gradient(circle at 50% 100%, rgba(90,90,120,0.045), transparent 36%),
        #050608;
    color: var(--text);
    font-family: Inter, -apple-system, BlinkMacSystemFont, "SF Pro Display", "Helvetica Neue", sans-serif;
}
[data-testid="stHeader"] { background: transparent; }
#MainMenu, footer { visibility: hidden; }
 
.block-container, [data-testid="stMainBlockContainer"] {
    padding-top: 1.4rem;
    padding-bottom: 4rem;
    max-width: 1560px;
}
 
.stApp, .stApp p, .stApp span, .stApp li, .stApp label,
.stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5 {
    color: var(--text);
    font-family: Inter, -apple-system, BlinkMacSystemFont, "SF Pro Display", "Helvetica Neue", sans-serif;
    -webkit-font-smoothing: antialiased;
}
a, a:visited { color: var(--gold-soft); }
 
/* ---------- glass surfaces ---------- */
.glass-card, .glass-panel, .glass-kpi {
    background: linear-gradient(135deg, rgba(255,255,255,0.085), rgba(255,255,255,0.035));
    border: 1px solid var(--border);
    border-radius: var(--radius-lg);
    backdrop-filter: blur(24px) saturate(140%);
    -webkit-backdrop-filter: blur(24px) saturate(140%);
    box-shadow: 0 20px 60px rgba(0,0,0,0.28), inset 0 1px 0 rgba(255,255,255,0.08);
    transition: border-color 180ms ease, transform 180ms ease, box-shadow 180ms ease;
}
.glass-card:hover, .glass-kpi:hover {
    border-color: var(--border-strong);
    transform: translateY(-1px);
}
.glass-panel {
    background: linear-gradient(135deg, rgba(255,255,255,0.05), rgba(255,255,255,0.022));
    border-radius: var(--radius-md);
    box-shadow: 0 14px 40px rgba(0,0,0,0.22), inset 0 1px 0 rgba(255,255,255,0.05);
    padding: 18px 20px;
}
.glass-card { padding: 22px 24px; }
.glass-kpi { padding: 16px 18px; border-radius: var(--radius-md); }
 
/* ---------- executive header ---------- */
.exec-header {
    display: flex;
    align-items: flex-end;
    justify-content: space-between;
    gap: 24px;
    flex-wrap: wrap;
    padding: 6px 2px 18px 2px;
    border-bottom: 1px solid var(--border);
    margin-bottom: 26px;
}
.exec-mark {
    font-size: 0.66rem;
    letter-spacing: 0.30em;
    text-transform: uppercase;
    color: var(--gold);
    font-weight: 600;
    margin-bottom: 10px;
}
.exec-title {
    font-size: 2.15rem;
    font-weight: 600;
    letter-spacing: -0.028em;
    line-height: 1.05;
    margin: 0;
}
.exec-sub {
    color: var(--secondary);
    font-size: 0.9rem;
    font-weight: 400;
    margin-top: 8px;
    letter-spacing: 0.005em;
}
.exec-status { display: flex; gap: 10px; flex-wrap: wrap; align-items: stretch; }
.status-chip {
    background: var(--glass-2);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 9px 14px;
    min-width: 104px;
    backdrop-filter: blur(18px);
    -webkit-backdrop-filter: blur(18px);
}
.status-chip .k {
    font-size: 0.6rem;
    letter-spacing: 0.16em;
    text-transform: uppercase;
    color: var(--tertiary);
    font-weight: 600;
}
.status-chip .v {
    font-size: 0.92rem;
    font-weight: 600;
    margin-top: 3px;
    color: var(--text);
}
.status-live .v { color: var(--green); }
.status-dot {
    display: inline-block;
    width: 7px; height: 7px;
    border-radius: 50%;
    background: var(--green);
    margin-right: 7px;
    box-shadow: 0 0 0 3px rgba(99,217,154,0.14);
}
 
/* ---------- section headers ---------- */
.glass-section {
    display: flex;
    align-items: baseline;
    gap: 14px;
    margin: 40px 0 6px 0;
    padding-top: 6px;
}
.glass-section .idx {
    font-size: 0.7rem;
    font-weight: 600;
    color: var(--gold);
    letter-spacing: 0.18em;
    font-variant-numeric: tabular-nums;
}
.glass-section .ttl {
    font-size: 1.22rem;
    font-weight: 600;
    letter-spacing: -0.015em;
    color: var(--text);
}
.glass-section .sub {
    font-size: 0.82rem;
    color: var(--tertiary);
    font-weight: 400;
}
.section-rule {
    height: 1px;
    background: linear-gradient(90deg, rgba(216,183,106,0.42), rgba(255,255,255,0.05) 42%, transparent);
    margin: 10px 0 18px 0;
}
 
/* ---------- metric typography ---------- */
.metric-label {
    font-size: 0.63rem;
    letter-spacing: 0.17em;
    text-transform: uppercase;
    color: var(--tertiary);
    font-weight: 600;
}
.metric-hero {
    font-size: 2.5rem;
    font-weight: 600;
    letter-spacing: -0.035em;
    line-height: 1.03;
    margin: 10px 0 2px 0;
    font-variant-numeric: tabular-nums;
}
.metric-hero.gold { color: var(--gold); }
.metric-value {
    font-size: 1.42rem;
    font-weight: 600;
    letter-spacing: -0.02em;
    margin-top: 6px;
    font-variant-numeric: tabular-nums;
}
.metric-secondary {
    font-size: 0.78rem;
    color: var(--secondary);
    margin-top: 5px;
    line-height: 1.5;
}
.metric-delta {
    font-size: 0.78rem;
    font-weight: 600;
    margin-top: 6px;
    font-variant-numeric: tabular-nums;
}
.metric-delta.pos { color: var(--green); }
.metric-delta.neg { color: var(--red); }
.metric-delta.flat { color: var(--secondary); font-weight: 500; }
.metric-delta.gold { color: var(--gold); }
 
.kpi-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    gap: 12px;
    margin: 4px 0 8px 0;
}
.hero-grid {
    display: grid;
    grid-template-columns: 1.15fr 1fr 1fr;
    gap: 16px;
    margin-bottom: 8px;
}
.hero-grid.no-aum {
    grid-template-columns: repeat(2, minmax(0, 1fr));
}
.trio-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
    gap: 14px;
}
 
/* ---------- text-overlap safety (layout/colors unchanged) ---------- */
.glass-card, .glass-panel, .glass-kpi,
.kpi-grid > *, .hero-grid > *, .trio-grid > *,
.kpi-head, .kpi-row, .progress-legend, .exec-header > *, .exec-status > * {
    min-width: 0;
}
.metric-label, .metric-hero, .metric-value, .metric-secondary, .metric-delta,
.kpi-row .k, .kpi-row .v, .scenario-hero .title, .scenario-hero .thesis,
.scenario-hero .detail, .scenario-hero .milestone, .glass-note, .glass-callout,
.exec-title, .exec-sub, .status-chip .v {
    overflow-wrap: anywhere;
    word-break: normal;
}
.kpi-row .k { flex: 1 1 auto; }
.kpi-row .v { flex: 0 1 52%; text-align: right; }
.progress-legend { gap: 10px; flex-wrap: wrap; }
.progress-legend span:last-child { margin-left: auto; text-align: right; }
div[role="radiogroup"] > label { max-width: 100%; }
div[role="radiogroup"] > label div[data-testid="stMarkdownContainer"] p {
    white-space: normal !important;
    overflow-wrap: anywhere;
}
.stTabs [data-baseweb="tab"] {
    min-width: 0;
    white-space: normal;
    text-align: center;
}

/* ---------- Streamlit text/icon overlap hardening ---------- */
/*
   IMPORTANT: On some corporate machines the Material Symbols web-font is
   blocked. Streamlit then paints the icon name itself (for example
   "keyboard_arrow_right") inside sidebar expanders. The long fallback text
   sits on top of labels such as "Segment mapping" / "Channel mapping".

   This patch does not redesign the sidebar. It removes only the broken icon
   glyph and gives the expander label its own protected column.
*/
[data-testid="stExpander"] summary {
    min-width: 0 !important;
}

/* Catch current and older Streamlit Material-icon wrappers. */
[data-testid="stExpander"] summary [data-testid="stExpanderToggleIcon"],
[data-testid="stExpander"] summary [data-testid="stIconMaterial"],
[data-testid="stExpander"] summary [aria-hidden="true"],
[data-testid="stExpander"] summary span[class*="material-symbol"],
[data-testid="stExpander"] summary span[class*="material-icon"] {
    font-size: 0 !important;
    line-height: 0 !important;
    letter-spacing: 0 !important;
    text-indent: -9999px !important;
    color: transparent !important;
    overflow: hidden !important;
    white-space: nowrap !important;
    max-width: 18px !important;
}

/* Strong fallback for the exact sidebar overlap shown in the screenshot.
   Streamlit places the expander toggle before the label. If the toggle loses
   its icon font, constrain that first control to a tiny fixed slot so its
   literal fallback text can never cross into the label. */
[data-testid="stSidebar"] [data-testid="stExpander"] summary {
    display: grid !important;
    grid-template-columns: 18px minmax(0, 1fr) !important;
    align-items: center !important;
    column-gap: 8px !important;
    width: 100% !important;
    overflow: hidden !important;
}
[data-testid="stSidebar"] [data-testid="stExpander"] summary > * {
    min-width: 0 !important;
}
[data-testid="stSidebar"] [data-testid="stExpander"] summary > :first-child {
    width: 18px !important;
    min-width: 18px !important;
    max-width: 18px !important;
    height: 18px !important;
    overflow: hidden !important;
    font-size: 0 !important;
    line-height: 0 !important;
    letter-spacing: 0 !important;
    text-indent: -9999px !important;
    color: transparent !important;
    white-space: nowrap !important;
}
[data-testid="stSidebar"] [data-testid="stExpander"] summary > :first-child * {
    font-size: 0 !important;
    line-height: 0 !important;
    letter-spacing: 0 !important;
    text-indent: -9999px !important;
    color: transparent !important;
    overflow: hidden !important;
    white-space: nowrap !important;
}
/* Replace the missing Material glyph with a font-independent chevron. */
[data-testid="stSidebar"] [data-testid="stExpander"] summary > :first-child::after {
    content: "›";
    display: block !important;
    width: 18px !important;
    height: 18px !important;
    text-align: center !important;
    text-indent: 0 !important;
    font-family: Arial, sans-serif !important;
    font-size: 17px !important;
    font-weight: 600 !important;
    line-height: 18px !important;
    color: var(--secondary) !important;
    transform-origin: 50% 50%;
}
[data-testid="stSidebar"] [data-testid="stExpander"] details[open] summary > :first-child::after {
    transform: rotate(90deg);
}

/* Keep the actual expander title in its own column and allow wrapping. */
[data-testid="stSidebar"] [data-testid="stExpander"] summary p,
[data-testid="stSidebar"] [data-testid="stExpander"] summary div[data-testid="stMarkdownContainer"] {
    min-width: 0 !important;
    margin: 0 !important;
    white-space: normal !important;
    overflow-wrap: anywhere !important;
    word-break: normal !important;
    line-height: 1.3 !important;
}

[data-testid="stSidebar"] [data-testid="stWidgetLabel"],
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"],
[data-testid="stSidebar"] [data-testid="stCaptionContainer"],
[data-testid="stMetricLabel"],
[data-testid="stMetricValue"] {
    min-width: 0 !important;
}

[data-testid="stSidebar"] [data-testid="stWidgetLabel"] p,
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p,
[data-testid="stSidebar"] [data-testid="stCaptionContainer"] p,
[data-testid="stMetricLabel"] p,
[data-testid="stMetricValue"] > div {
    white-space: normal !important;
    overflow-wrap: anywhere !important;
    word-break: normal !important;
    line-height: 1.3 !important;
}

.stTabs [data-baseweb="tab-list"] {
    flex-wrap: wrap !important;
}

/* ---------- Current Performance + asset contribution ---------- */
.sales-kpi-card {
    position: relative;
}
/* Gross and Net Sales use different accents while keeping the same card layout. */
.sales-kpi-card.gross-sales-card {
    border-color: rgba(105, 166, 255, 0.34);
    background: linear-gradient(135deg, rgba(77, 143, 238, 0.13), rgba(255,255,255,0.035));
}
.sales-kpi-card.gross-sales-card .metric-hero { color: #8EC2FF; }
.sales-kpi-card.gross-sales-card .kpi-tag {
    color: #9BC8FF;
    border-color: rgba(105, 166, 255, 0.30);
}
.sales-kpi-card.gross-sales-card .asset-contrib-details > summary {
    border-color: rgba(105, 166, 255, 0.48);
    background: rgba(105, 166, 255, 0.08);
}
.sales-kpi-card.gross-sales-card .asset-contrib-details > summary::after,
.sales-kpi-card.gross-sales-card .asset-contrib-share { color: #8EC2FF; }

.sales-kpi-card.net-sales-card {
    border-color: rgba(99, 217, 154, 0.34);
    background: linear-gradient(135deg, rgba(69, 176, 119, 0.13), rgba(255,255,255,0.035));
}
.sales-kpi-card.net-sales-card .metric-hero { color: #79E0AA; }
.sales-kpi-card.net-sales-card .kpi-tag {
    color: #79E0AA;
    border-color: rgba(99, 217, 154, 0.30);
}
.sales-kpi-card.net-sales-card .asset-contrib-details > summary {
    border-color: rgba(99, 217, 154, 0.48);
    background: rgba(99, 217, 154, 0.08);
}
.sales-kpi-card.net-sales-card .asset-contrib-details > summary::after,
.sales-kpi-card.net-sales-card .asset-contrib-share { color: #79E0AA; }

.metric-hero-line {
    display: flex;
    align-items: center;
    gap: 10px;
    min-width: 0;
}
.metric-hero-line .metric-hero {
    min-width: 0;
}
.asset-contrib-details {
    position: relative;
    margin: 0;
    padding: 0;
}
.asset-contrib-details > summary {
    list-style: none;
    width: 27px;
    height: 27px;
    min-width: 27px;
    border: 1px solid rgba(216,183,106,0.48);
    border-radius: 50%;
    background: rgba(216,183,106,0.08);
    cursor: pointer;
    display: grid;
    place-items: center;
    user-select: none;
    transition: background 160ms ease, border-color 160ms ease, transform 160ms ease;
}
.asset-contrib-details > summary::-webkit-details-marker { display: none; }
.asset-contrib-details > summary::marker { content: ""; }
.asset-contrib-details > summary::after {
    content: "+";
    color: var(--gold);
    font-family: Arial, sans-serif;
    font-size: 19px;
    font-weight: 400;
    line-height: 1;
    transform: translateY(-1px);
}
.asset-contrib-details[open] > summary::after { content: "−"; }
.asset-contrib-details > summary:hover {
    background: rgba(216,183,106,0.15);
    border-color: rgba(216,183,106,0.72);
    transform: translateY(-1px);
}
.asset-contrib-panel {
    position: absolute;
    z-index: 50;
    top: 36px;
    right: 0;
    width: min(310px, 72vw);
    padding: 13px 14px;
    border: 1px solid var(--border-strong);
    border-radius: 14px;
    background: rgba(13,14,18,0.97);
    box-shadow: 0 18px 48px rgba(0,0,0,0.48), inset 0 1px 0 rgba(255,255,255,0.06);
    backdrop-filter: blur(22px);
    -webkit-backdrop-filter: blur(22px);
}
.asset-contrib-title {
    color: var(--secondary);
    font-size: 0.65rem;
    font-weight: 600;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    margin-bottom: 7px;
}
.asset-contrib-row {
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 12px;
    padding: 8px 0;
    border-top: 1px solid rgba(255,255,255,0.055);
}
.asset-contrib-name {
    color: var(--secondary);
    font-size: 0.8rem;
}
.asset-contrib-values {
    display: flex;
    align-items: baseline;
    justify-content: flex-end;
    gap: 18px;
    min-width: 0;
}
.asset-contrib-amount {
    color: var(--text);
    font-size: 0.82rem;
    font-weight: 600;
    font-variant-numeric: tabular-nums;
}
.asset-contrib-share {
    color: var(--gold);
    font-size: 0.74rem;
    font-weight: 600;
    min-width: 58px;
    text-align: right;
    padding-right: 2px;
    font-variant-numeric: tabular-nums;
}
.asset-contrib-empty {
    color: var(--secondary);
    font-size: 0.76rem;
    line-height: 1.45;
}

/* ---------- kpi card internals ---------- */
.kpi-head {
    display: flex;
    justify-content: space-between;
    align-items: center;
    gap: 10px;
}
.kpi-tag {
    font-size: 0.6rem;
    letter-spacing: 0.14em;
    text-transform: uppercase;
    color: var(--tertiary);
    border: 1px solid var(--border);
    border-radius: 999px;
    padding: 3px 9px;
}
.kpi-rows { margin-top: 16px; border-top: 1px solid rgba(255,255,255,0.07); }
.kpi-row {
    display: flex;
    justify-content: space-between;
    align-items: baseline;
    gap: 12px;
    padding: 8px 0;
    border-bottom: 1px solid rgba(255,255,255,0.045);
    font-size: 0.83rem;
}
.kpi-row:last-child { border-bottom: none; }
.kpi-row .k { color: var(--secondary); }
.kpi-row .v { font-weight: 600; font-variant-numeric: tabular-nums; }
.kpi-row .v.gold { color: var(--gold); }
.kpi-row .v.pos { color: var(--green); }
.kpi-row .v.neg { color: var(--red); }
 
/* ---------- progress ---------- */
.progress {
    position: relative;
    height: 6px;
    border-radius: 999px;
    background: rgba(255,255,255,0.08);
    margin: 16px 0 8px 0;
    overflow: visible;
}
.progress-fill {
    position: absolute;
    left: 0; top: 0; bottom: 0;
    border-radius: 999px;
    background: linear-gradient(90deg, rgba(216,183,106,0.55), var(--gold));
    box-shadow: 0 0 18px rgba(216,183,106,0.28);
}
.progress-fill.neutral {
    background: linear-gradient(90deg, rgba(255,255,255,0.22), rgba(255,255,255,0.42));
    box-shadow: none;
}
.progress-marker {
    position: absolute;
    top: -4px;
    width: 2px;
    height: 14px;
    background: var(--text);
    opacity: 0.85;
    border-radius: 2px;
}
.progress-legend {
    display: flex;
    justify-content: space-between;
    font-size: 0.72rem;
    color: var(--tertiary);
    font-variant-numeric: tabular-nums;
}
 
/* ---------- notes, callouts, pills ---------- */
.glass-note {
    color: var(--tertiary);
    font-size: 0.78rem;
    line-height: 1.6;
    margin: 8px 0 14px 0;
}
.glass-callout {
    background: var(--glass-3);
    border: 1px solid var(--border);
    border-left: 2px solid var(--gold);
    border-radius: var(--radius-sm);
    padding: 13px 16px;
    font-size: 0.85rem;
    line-height: 1.62;
    color: var(--secondary);
    margin: 10px 0 14px 0;
}
.glass-callout b, .glass-callout strong { color: var(--text); font-weight: 600; }
.glass-callout.ok { border-left-color: var(--green); }
.glass-callout.warn { border-left-color: var(--red); }
.tag-ok { color: var(--green); font-weight: 600; letter-spacing: 0.04em; }
.tag-warn { color: var(--red); font-weight: 600; letter-spacing: 0.04em; }
.inline-pill {
    display: inline-block;
    border: 1px solid var(--border);
    background: var(--glass-3);
    color: var(--secondary);
    border-radius: 999px;
    padding: 4px 11px;
    font-size: 0.7rem;
    font-weight: 500;
    margin: 3px 5px 3px 0;
}
.inline-pill.gold { color: var(--gold); border-color: rgba(216,183,106,0.35); }
 

/* ---------- scenario numeric emphasis ---------- */
.scenario-kpi-grid .metric-value {
    font-size: 2.20rem;
    line-height: 1.03;
    font-weight: 750;
    letter-spacing: -0.025em;
}
.trio-grid .stage-card .metric-hero {
    font-size: 2.35rem;
    line-height: 1.02;
}
.trio-grid .stage-card .kpi-row .v {
    font-size: 1.24rem;
    line-height: 1.15;
    font-weight: 750;
}
.scenario-table .glass-table td.amount {
    font-size: 1.15rem;
    line-height: 1.2;
    font-weight: 750;
    letter-spacing: -0.015em;
}
.revenue-table .glass-table tbody tr.total td {
    font-size: 1.12rem;
    font-weight: 800;
}
.revenue-table .glass-table tbody tr.total td.amount {
    font-size: 1.34rem;
    line-height: 1.15;
}
.revenue-compare-grid {
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 14px;
    align-items: stretch;
    margin: 4px 0 14px 0;
}
.revenue-current-card { border-color: rgba(142,194,255,0.34); }
.revenue-expected-card { border-color: rgba(121,224,170,0.38); }
.revenue-current-card .revenue-compare-amount { color: #8EC2FF; }
.revenue-expected-card .revenue-compare-amount { color: #79E0AA; }
.revenue-compare-amount {
    font-size: 2.70rem;
    font-weight: 750;
    letter-spacing: -0.035em;
    line-height: 1.05;
    margin-top: 10px;
    font-variant-numeric: tabular-nums;
}
.revenue-bridge-card {
    min-width: 0;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    text-align: center;
    padding: 16px 10px;
}
.revenue-bridge-delta {
    font-size: 1.55rem;
    font-weight: 700;
    font-variant-numeric: tabular-nums;
}
@media (max-width: 800px) {
    .revenue-compare-grid { grid-template-columns: 1fr; }
    .revenue-bridge-card { padding: 4px 10px; }
}

/* ---------- scenario hero ---------- */
.scenario-hero {
    background:
        radial-gradient(circle at 88% -30%, rgba(216,183,106,0.16), transparent 55%),
        linear-gradient(135deg, rgba(255,255,255,0.085), rgba(255,255,255,0.03));
    border: 1px solid rgba(216,183,106,0.24);
    border-radius: var(--radius-lg);
    padding: 26px 28px;
    backdrop-filter: blur(26px) saturate(140%);
    -webkit-backdrop-filter: blur(26px) saturate(140%);
    box-shadow: 0 24px 70px rgba(0,0,0,0.34), inset 0 1px 0 rgba(255,255,255,0.09);
    margin-bottom: 16px;
}
.scenario-hero .eyebrow {
    font-size: 0.64rem;
    letter-spacing: 0.24em;
    text-transform: uppercase;
    color: var(--gold);
    font-weight: 600;
}
.scenario-hero .title {
    font-size: 1.72rem;
    font-weight: 600;
    letter-spacing: -0.028em;
    margin: 10px 0 6px 0;
}
.scenario-hero .thesis {
    font-size: 1.0rem;
    color: var(--text);
    opacity: 0.9;
    font-weight: 400;
    margin-bottom: 8px;
}
.scenario-hero .detail {
    font-size: 0.84rem;
    color: var(--secondary);
    line-height: 1.62;
    max-width: 940px;
}
.scenario-hero .milestone {
    margin-top: 14px;
    font-size: 0.76rem;
    color: var(--gold-soft);
    letter-spacing: 0.02em;
}
.stage-card { padding: 18px 20px; }
.stage-card .stage {
    font-size: 0.62rem;
    letter-spacing: 0.2em;
    text-transform: uppercase;
    color: var(--tertiary);
    font-weight: 600;
}
.stage-card.now { border-top: 2px solid rgba(255,255,255,0.22); }
.stage-card.jan { border-top: 2px solid rgba(216,183,106,0.55); }
.stage-card.mar { border-top: 2px solid rgba(99,217,154,0.45); }
 
/* ---------- glass tables ---------- */
.glass-table-wrap {
    border: 1px solid var(--border);
    border-radius: var(--radius-md);
    background: linear-gradient(135deg, rgba(255,255,255,0.045), rgba(255,255,255,0.018));
    backdrop-filter: blur(18px);
    -webkit-backdrop-filter: blur(18px);
    overflow: auto;
    margin: 6px 0 14px 0;
    max-height: 560px;
}
table.glass-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 0.8rem;
    font-variant-numeric: tabular-nums;
}
table.glass-table thead th {
    position: sticky;
    top: 0;
    z-index: 2;
    background: rgba(16,17,22,0.94);
    backdrop-filter: blur(14px);
    color: var(--tertiary);
    font-size: 0.62rem;
    letter-spacing: 0.13em;
    text-transform: uppercase;
    font-weight: 600;
    text-align: left;
    padding: 11px 14px;
    border-bottom: 1px solid var(--border);
    white-space: nowrap;
}
table.glass-table thead th.num { text-align: right; }
table.glass-table tbody td {
    padding: 10px 14px;
    border-bottom: 1px solid rgba(255,255,255,0.045);
    color: var(--text);
    white-space: nowrap;
}
table.glass-table tbody td.num { text-align: right; }
table.glass-table tbody td.pos { color: var(--green); }
table.glass-table tbody td.neg { color: var(--red); }
table.glass-table tbody tr:hover td { background: rgba(255,255,255,0.035); }
table.glass-table tbody tr.total td { font-weight: 600; color: var(--gold); }
table.glass-table tbody tr:last-child td { border-bottom: none; }
 
/* ---------- FINAL workbook reference ---------- */
.final-sheet-scroll {
    overflow: auto;
    max-height: 70vh;
    border: 1px solid var(--border);
    border-radius: var(--radius-md);
    background: rgba(255,255,255,0.02);
    padding: 4px;
}
table.final-sheet-table {
    border-collapse: collapse;
    width: max-content;
    min-width: 100%;
    font-size: 11.5px;
    color: var(--text);
}
table.final-sheet-table td {
    border: 1px solid rgba(255,255,255,0.07);
    padding: 5px 8px;
    min-width: 74px;
    white-space: nowrap;
    vertical-align: middle;
}
td.final-spacer { height: 7px; border: none !important; }
 
/* ---------- streamlit widgets ---------- */
div[role="radiogroup"] { gap: 8px !important; flex-wrap: wrap; }
div[role="radiogroup"] > label {
    background: var(--glass-2);
    border: 1px solid var(--border);
    border-radius: 999px;
    padding: 8px 15px;
    margin: 0 !important;
    transition: all 160ms ease;
    backdrop-filter: blur(14px);
    -webkit-backdrop-filter: blur(14px);
}
div[role="radiogroup"] > label > div:first-child { display: none !important; }
div[role="radiogroup"] > label div[data-testid="stMarkdownContainer"] p {
    font-size: 0.78rem !important;
    color: var(--secondary) !important;
    font-weight: 500;
    margin: 0 !important;
}
div[role="radiogroup"] > label:hover { border-color: var(--border-strong); }
div[role="radiogroup"] > label:has(input:checked) {
    border-color: rgba(216,183,106,0.55);
    background: linear-gradient(135deg, rgba(216,183,106,0.16), rgba(255,255,255,0.05));
    box-shadow: 0 8px 26px rgba(216,183,106,0.14), inset 0 1px 0 rgba(255,255,255,0.10);
}
div[role="radiogroup"] > label:has(input:checked) div[data-testid="stMarkdownContainer"] p {
    color: var(--text) !important;
    font-weight: 600;
}
 
[data-baseweb="select"] > div, [data-baseweb="input"], .stTextInput input, .stNumberInput input {
    background: var(--glass-2) !important;
    border-color: var(--border) !important;
    color: var(--text) !important;
    border-radius: 12px !important;
}
[data-baseweb="select"] svg { fill: var(--secondary); }
[data-baseweb="popover"] div[role="listbox"], [data-baseweb="menu"] {
    background: rgba(16,17,22,0.97) !important;
    border: 1px solid var(--border) !important;
    border-radius: 14px !important;
}
[data-baseweb="menu"] li:hover { background: rgba(255,255,255,0.06) !important; }
[data-baseweb="tag"] {
    background: rgba(216,183,106,0.16) !important;
    border: 1px solid rgba(216,183,106,0.3) !important;
    color: var(--text) !important;
}
 
.stSlider [data-baseweb="slider"] div[role="slider"] {
    background: var(--gold) !important;
    box-shadow: 0 0 0 4px rgba(216,183,106,0.16) !important;
    transition: transform 140ms ease;
}
.stSlider [data-baseweb="slider"] div[role="slider"]:active { transform: scale(1.1); }
.stSlider [data-testid="stTickBar"], .stSlider [data-testid="stTickBarMin"],
.stSlider [data-testid="stTickBarMax"] { color: var(--tertiary) !important; }
 
.stButton > button, .stDownloadButton > button, .stFormSubmitButton > button {
    background: linear-gradient(135deg, rgba(216,183,106,0.9), rgba(201,170,101,0.78)) !important;
    color: #10120F !important;
    border: 1px solid rgba(216,183,106,0.5) !important;
    border-radius: 14px !important;
    font-weight: 600 !important;
    letter-spacing: 0.01em;
    padding: 8px 18px !important;
    transition: transform 160ms ease, box-shadow 160ms ease;
    box-shadow: 0 8px 24px rgba(216,183,106,0.16);
}
.stButton > button:hover, .stDownloadButton > button:hover {
    transform: translateY(-1px);
    box-shadow: 0 14px 34px rgba(216,183,106,0.24);
}
.stButton > button[kind="secondary"] {
    background: var(--glass-2) !important;
    color: var(--text) !important;
    border: 1px solid var(--border) !important;
    box-shadow: none;
}
 
.stTabs [data-baseweb="tab-list"] {
    gap: 6px;
    border-bottom: 1px solid var(--border);
    background: transparent;
}
.stTabs [data-baseweb="tab"] {
    background: var(--glass-3);
    border: 1px solid transparent;
    border-radius: 12px 12px 0 0;
    color: var(--tertiary);
    padding: 8px 16px;
    font-size: 0.8rem;
}
.stTabs [aria-selected="true"] {
    color: var(--text) !important;
    background: var(--glass-1);
    border-color: var(--border);
    border-bottom-color: transparent;
    font-weight: 600;
}
.stTabs [data-baseweb="tab-highlight"] { background: var(--gold) !important; }
 
[data-testid="stExpander"] {
    border: 1px solid var(--border) !important;
    border-radius: var(--radius-md) !important;
    background: var(--glass-3) !important;
    backdrop-filter: blur(16px);
    -webkit-backdrop-filter: blur(16px);
}
[data-testid="stExpander"] summary { color: var(--secondary) !important; font-size: 0.84rem; }
[data-testid="stExpander"] summary:hover { color: var(--text) !important; }
 
[data-testid="stSidebar"] {
    background: rgba(10,11,14,0.86) !important;
    border-right: 1px solid var(--border);
    backdrop-filter: blur(26px) saturate(140%);
    -webkit-backdrop-filter: blur(26px) saturate(140%);
}
[data-testid="stSidebar"] * { color: var(--text); }
[data-testid="stSidebar"] [data-testid="stCaptionContainer"] p { color: var(--tertiary) !important; }
.sidebar-mark {
    font-size: 0.6rem;
    letter-spacing: 0.26em;
    text-transform: uppercase;
    color: var(--gold);
    font-weight: 600;
    margin-bottom: 4px;
}
.sidebar-title {
    font-size: 0.62rem;
    letter-spacing: 0.16em;
    text-transform: uppercase;
    color: var(--tertiary);
    font-weight: 600;
    margin: 18px 0 6px 0;
}
 
[data-testid="stDataFrame"] {
    border: 1px solid var(--border);
    border-radius: var(--radius-md);
    overflow: hidden;
}
[data-testid="stFileUploaderDropzone"] {
    background: var(--glass-2) !important;
    border: 1px dashed rgba(255,255,255,0.18) !important;
    border-radius: var(--radius-md) !important;
    color: var(--secondary) !important;
}
[data-testid="stFileUploaderDropzone"]:hover { border-color: rgba(216,183,106,0.45) !important; }
 
[data-testid="stAlert"] {
    background: var(--glass-2) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--radius-sm) !important;
    color: var(--text) !important;
}
hr { border-color: var(--border) !important; }
[data-testid="stMetricValue"] { color: var(--text) !important; }
 
/* ---------- responsive ---------- */
@media (max-width: 1180px) {
    .hero-grid { grid-template-columns: 1fr 1fr; }
    .exec-title { font-size: 1.85rem; }
}
@media (max-width: 760px) {
    .hero-grid, .hero-grid.no-aum { grid-template-columns: 1fr; }
    .kpi-grid { grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); }
    .metric-hero { font-size: 2.05rem; }
    .exec-header { flex-direction: column; align-items: flex-start; }
    div[role="radiogroup"] { flex-wrap: nowrap; overflow-x: auto; padding-bottom: 6px; }
}
@media (prefers-reduced-motion: reduce) {
    * { animation-duration: 0.01ms !important; transition-duration: 0.01ms !important; }
    .glass-card:hover, .glass-kpi:hover, .stButton > button:hover { transform: none; }
}
</style>
"""
 
 
def inject_theme() -> None:
    """Inject the liquid-glass design system exactly once per run."""
    st.markdown(GLASS_CSS, unsafe_allow_html=True)
 
 
# --- Small HTML helpers -------------------------------------------------------
 
def _pct_width(value: Any, cap: float = 100.0) -> float:
    v = _num(value)
    if v is None:
        return 0.0
    return float(min(max(v * 100.0, 0.0), cap))
 
 
def _tone_for(value: Any, invert: bool = False) -> str:
    v = _num(value)
    if v is None:
        return "flat"
    if abs(v) < 1e-12:
        return "flat"
    positive = v > 0
    if invert:
        positive = not positive
    return "pos" if positive else "neg"
 
 
def section_header(index: str, title: str, subtitle: str = "") -> None:
    sub = f"<span class='sub'>{escape(subtitle)}</span>" if subtitle else ""
    st.markdown(
        f"<div class='glass-section'><span class='idx'>{escape(index)}</span>"
        f"<span class='ttl'>{escape(title)}</span>{sub}</div>"
        "<div class='section-rule'></div>",
        unsafe_allow_html=True,
    )
 
 
def glass_note(text: str) -> None:
    st.markdown(f"<div class='glass-note'>{text}</div>", unsafe_allow_html=True)
 
 
def glass_callout(text: str, tone: str = "") -> None:
    css = "glass-callout" + (f" {tone}" if tone else "")
    st.markdown(f"<div class='{css}'>{text}</div>", unsafe_allow_html=True)
 
 
def progress_html(
    achieved_pct: Any,
    marker_pct: Any = None,
    left_label: str = "",
    right_label: str = "",
    neutral: bool = False,
) -> str:
    fill = _pct_width(achieved_pct)
    marker = _pct_width(marker_pct) if marker_pct is not None else None
    marker_html = (
        f"<div class='progress-marker' style='left:{marker:.1f}%'></div>"
        if marker is not None else ""
    )
    fill_class = "progress-fill neutral" if neutral else "progress-fill"
    legend = ""
    if left_label or right_label:
        legend = (
            "<div class='progress-legend'>"
            f"<span>{escape(left_label)}</span><span>{escape(right_label)}</span></div>"
        )
    return (
        f"<div class='progress'><div class='{fill_class}' style='width:{fill:.1f}%'></div>"
        f"{marker_html}</div>{legend}"
    )
 
 
def kpi_tile_html(
    label: str,
    value: str,
    delta: str = "",
    delta_tone: str = "flat",
    secondary: str = "",
) -> str:
    delta_html = (
        f"<div class='metric-delta {delta_tone}'>{escape(delta)}</div>" if delta else ""
    )
    secondary_html = (
        f"<div class='metric-secondary'>{escape(secondary)}</div>" if secondary else ""
    )
    return (
        "<div class='glass-kpi'>"
        f"<div class='metric-label'>{escape(label)}</div>"
        f"<div class='metric-value'>{escape(value)}</div>"
        f"{delta_html}{secondary_html}</div>"
    )
 
 
def kpi_strip(tiles: Sequence[Dict[str, str]], css_class: str = "") -> None:
    """Render a responsive row of glass KPI tiles from one markdown call."""
    if not tiles:
        return
    cards = "".join(
        kpi_tile_html(
            tile.get("label", ""),
            tile.get("value", NA_TEXT),
            tile.get("delta", ""),
            tile.get("tone", "flat"),
            tile.get("secondary", ""),
        )
        for tile in tiles
    )
    classes = "kpi-grid" + (f" {css_class.strip()}" if css_class.strip() else "")
    st.markdown(f"<div class='{classes}'>{cards}</div>", unsafe_allow_html=True)
 
 
def _dataframe_kwargs() -> Dict[str, Any]:
    try:
        parameters = inspect.signature(st.dataframe).parameters
    except (TypeError, ValueError):
        return {}
    if "width" in parameters and parameters["width"].default == "stretch":
        return {}
    if "use_container_width" in parameters:
        return {"use_container_width": True}
    return {}
 
 
def render_glass_table(
    frame: pd.DataFrame,
    formats: Optional[Dict[str, str]] = None,
    total_rows: Sequence[str] = (),
    max_html_rows: int = 240,
    empty_message: str = "No rows for this selection.",
    css_class: str = "",
) -> None:
    """Compact glass table: sticky header, right-aligned numbers, signed colour."""
    if frame is None or frame.empty:
        glass_note(escape(empty_message))
        return
 
    if len(frame) > max_html_rows:  # very large frames stay virtualised
        display = format_table(frame, formats) if formats else frame
        try:
            st.dataframe(display, hide_index=True, **_dataframe_kwargs())
        except TypeError:  # pragma: no cover - very old Streamlit
            st.dataframe(display)
        return
 
    formats = formats or {}
    display = format_table(frame, formats) if formats else frame.astype(str)
 
    header_cells = []
    for column in display.columns:
        kind = formats.get(column, "txt")
        css = " class='num'" if kind in NUMERIC_FORMATS else ""
        header_cells.append(f"<th{css}>{escape(str(column))}</th>")
 
    first_column = display.columns[0] if len(display.columns) else None
    body_rows = []
    for position in range(len(display)):
        raw_first = str(frame.iloc[position][first_column]) if first_column is not None else ""
        row_class = " class='total'" if raw_first in total_rows else ""
        cells = []
        for column in display.columns:
            kind = formats.get(column, "txt")
            classes = []
            if kind in NUMERIC_FORMATS:
                classes.append("num")
            if kind in AMOUNT_FORMATS:
                classes.append("amount")
            if kind in SIGNED_FORMATS:
                classes.append(_tone_for(frame.iloc[position][column]))
            css = f" class='{' '.join(classes)}'" if classes else ""
            cells.append(f"<td{css}>{escape(str(display.iloc[position][column]))}</td>")
        body_rows.append(f"<tr{row_class}>{''.join(cells)}</tr>")
 
    wrapper_classes = "glass-table-wrap" + (f" {css_class.strip()}" if css_class.strip() else "")
    st.markdown(
        f"<div class='{wrapper_classes}'><table class='glass-table'>"
        f"<thead><tr>{''.join(header_cells)}</tr></thead>"
        f"<tbody>{''.join(body_rows)}</tbody></table></div>",
        unsafe_allow_html=True,
    )
 
 
def reset_stale_selection(key: str, options: Sequence[Any]) -> None:
    """Drop a stored widget value that is no longer offered, so nothing errors."""
    if key in st.session_state and st.session_state[key] not in options:
        st.session_state.pop(key, None)
 
 
def rerun() -> None:
    handler = getattr(st, "rerun", None) or getattr(st, "experimental_rerun", None)
    if handler is not None:
        handler()
 
 
# =============================================================================
# 15. EXECUTIVE COMPONENTS - CURRENT REALITY
# =============================================================================
 
def render_apple_header(
    final_metrics: Dict[str, Any],
    records: pd.DataFrame,
    page_label: str,
) -> None:
    """01 · Minimal executive header with workbook status on the right."""
    months_done = int(final_metrics.get("months_done", MONTHS_COMPLETED))
    sheet = str(final_metrics.get("sheet_name") or "FINAL")
    rm_count = len(records)
 
    st.markdown(
        "<div class='exec-header'>"
        "<div>"
        f"<div class='exec-mark'>{escape(page_label)}</div>"
        f"<div class='exec-title'>{escape(APP_TITLE)}</div>"
        f"<div class='exec-sub'>{escape(APP_SUBTITLE)}</div>"
        "</div>"
        "<div class='exec-status'>"
        "<div class='status-chip status-live'><div class='k'>Workbook</div>"
        f"<div class='v'><span class='status-dot'></span>{escape(sheet)}</div></div>"
        "<div class='status-chip'><div class='k'>Data period</div>"
        "<div class='v'>Apr–Jun FY27</div></div>"
        "<div class='status-chip'><div class='k'>Months done</div>"
        f"<div class='v'>{months_done} of 12</div></div>"
        "<div class='status-chip'><div class='k'>RMs in scope</div>"
        f"<div class='v'>{rm_count:,}</div></div>"
        "</div></div>",
        unsafe_allow_html=True,
    )
 
 
def _kpi_row_html(label: str, value: str, css: str = "") -> str:
    value_class = f"v {css}".strip()
    return (
        f"<div class='kpi-row'><span class='k'>{escape(label)}</span>"
        f"<span class='{value_class}'>{escape(value)}</span></div>"
    )
 
 
def render_aum_hero(aum: pd.DataFrame) -> str:
    """AUM card shown only in Current Performance Metrics."""
    if not isinstance(aum, pd.DataFrame) or aum.empty or "Overall" not in aum.index:
        return (
            "<div class='glass-card'><div class='metric-label'>Assets under management</div>"
            "<div class='metric-hero'>—</div>"
            "<div class='metric-secondary'>AUM could not be located on the FINAL sheet.</div></div>"
        )

    row = aum.loc["Overall"]
    achievement = _num(row.get("Achievement %"))
    gap = _num(row.get("Gap to Target"))

    rows = "".join([
        _kpi_row_html("Target", fmt_cr(row.get("Target"))),
        _kpi_row_html(
            "Gap to target",
            fmt_cr_signed(None if gap is None else -gap),
            "neg" if (gap or 0) > 0 else "pos",
        ),
        _kpi_row_html("Achieved", fmt_pct(achievement), "gold"),
    ])

    return (
        "<div class='glass-card'>"
        "<div class='kpi-head'><span class='metric-label'>Assets under management</span>"
        "<span class='kpi-tag'>FINAL</span></div>"
        f"<div class='metric-hero gold'>{escape(fmt_cr(row.get('Current')))}</div>"
        "<div class='metric-label'>Current AUM</div>"
        + progress_html(
            achievement,
            marker_pct=1.0,
            left_label=f"{fmt_pct(achievement)} of target",
            right_label=f"target {fmt_cr(row.get('Target'))}",
        )
        + f"<div class='kpi-rows'>{rows}</div></div>"
    )


def _asset_contribution_html(frame: pd.DataFrame) -> str:
    """Expandable YTD contribution of Equity, Debt and Liquid to the sales total."""
    if not isinstance(frame, pd.DataFrame) or frame.empty:
        return (
            "<div class='asset-contrib-empty'>"
            "Asset-class contribution could not be located on the FINAL sheet."
            "</div>"
        )

    total = None
    if "Overall" in frame.index:
        total = _num(frame.loc["Overall"].get("YTD"))

    asset_values: List[Tuple[str, Optional[float]]] = []
    for asset_name in FINAL_ASSET_ROWS:
        value = None
        if asset_name in frame.index:
            value = _num(frame.loc[asset_name].get("YTD"))
        asset_values.append((asset_name, value))

    # If FINAL does not expose an Overall YTD value, fall back to the visible
    # asset-class rows so the contribution percentages can still be calculated.
    if total is None:
        present_values = [value for _, value in asset_values if value is not None]
        total = sum(present_values) if present_values else None

    contribution_rows: List[str] = []
    for asset_name, value in asset_values:
        share = None
        if value is not None and total is not None and total != 0:
            share = value / total
        contribution_rows.append(
            "<div class='asset-contrib-row'>"
            f"<span class='asset-contrib-name'>{escape(asset_name)}</span>"
            "<span class='asset-contrib-values'>"
            f"<span class='asset-contrib-amount'>{escape(fmt_cr(value))}</span>"
            f"<span class='asset-contrib-share'>{escape(fmt_pct(share))}</span>"
            "</span></div>"
        )

    return (
        "<div class='asset-contrib-title'>YTD asset-class contribution</div>"
        + "".join(contribution_rows)
    )


def render_sales_kpi_card(title: str, row: pd.Series, frame: pd.DataFrame) -> str:
    """Gross / Net sales card with an inline + control for asset contribution."""
    sales_card_class = "gross-sales-card" if title.strip().lower().startswith("gross") else "net-sales-card"
    if row is None or len(row) == 0:
        return (
            f"<div class='glass-card sales-kpi-card {sales_card_class}'><div class='metric-label'>{escape(title)}</div>"
            "<div class='metric-hero'>—</div>"
            "<div class='metric-secondary'>Metrics could not be located on the FINAL sheet.</div></div>"
        )
 
    achievement = _num(row.get("Achievement %"))
    projected = _num(row.get("Projected FY %"))
    current_rr = _num(row.get("Current RR"))

    # Recompute here as a final UI safeguard: remaining target / 9 remaining months.
    # Example Net Sales: (58,699.46 - 17,852.62) / 9 = 4,538.54 Cr/month.
    fy_target = _num(row.get("FY27 Target"))
    ytd_value = _num(row.get("YTD"))
    required_rr = None
    if fy_target is not None and ytd_value is not None:
        required_rr = (fy_target - ytd_value) / max(MONTHS_REMAINING, 1)
    pace_gap = None
    if current_rr is not None and required_rr is not None and required_rr != 0:
        pace_gap = current_rr / required_rr - 1.0
 
    rows = "".join([
        _kpi_row_html("FY27 target", fmt_cr(row.get("FY27 Target"))),
        _kpi_row_html("Current run rate", fmt_cr(current_rr)),
        _kpi_row_html("Required run rate", fmt_cr(required_rr), "gold"),
        _kpi_row_html(
            "Pace vs required",
            fmt_pct_signed(pace_gap),
            _tone_for(pace_gap),
        ),
        _kpi_row_html(
            "Projected FY",
            fmt_pct(projected),
            "pos" if (projected or 0) >= 1 else "neg",
        ),
    ])

    contribution = _asset_contribution_html(frame)
 
    return (
        f"<div class='glass-card sales-kpi-card {sales_card_class}'>"
        f"<div class='kpi-head'><span class='metric-label'>{escape(title)}</span>"
        "<span class='kpi-tag'>FY27</span></div>"
        "<div class='metric-hero-line'>"
        f"<div class='metric-hero'>{escape(fmt_cr(row.get('YTD')))}</div>"
        "<details class='asset-contrib-details'>"
        f"<summary title='Show {escape(title)} asset-class contribution' "
        f"aria-label='Show {escape(title)} asset-class contribution'></summary>"
        f"<div class='asset-contrib-panel'>{contribution}</div>"
        "</details></div>"
        "<div class='metric-label'>Year to date</div>"
        + progress_html(
            achievement,
            marker_pct=projected,
            left_label=f"{fmt_pct(achievement)} achieved",
            right_label=f"projected {fmt_pct(projected)}",
        )
        + f"<div class='kpi-rows'>{rows}</div></div>"
    )
 
 
def _overall_row(frame: Any) -> pd.Series:
    if isinstance(frame, pd.DataFrame) and not frame.empty and "Overall" in frame.index:
        return frame.loc["Overall"]
    return pd.Series(dtype=float)
 
 
def render_current_performance(final_metrics: Dict[str, Any], model: ScenarioModel) -> None:
    """02 · Current Performance Metrics · FINAL - AUM, Gross and Net at one glance."""
    section_header(
        "02",
        "Current Performance Metrics · FINAL",
        "Where the business stands before any scenario is applied",
    )

    gs = final_sales_metrics(final_metrics, model, "GS")
    ns = final_sales_metrics(final_metrics, model, "NS")
    aum = final_metrics.get("AUM")
    aum_frame = aum if isinstance(aum, pd.DataFrame) else pd.DataFrame()

    # Current Performance Metrics must always contain these three cards in this
    # order: AUM, Gross Sales, Net Sales. AUM remains excluded from the separate
    # Business Drivers section, as requested.
    aum_card = render_aum_hero(aum_frame)
    gross_card = render_sales_kpi_card("Gross Sales", _overall_row(gs), gs)
    net_card = render_sales_kpi_card("Net Sales", _overall_row(ns), ns)
    cards = aum_card + gross_card + net_card
    st.markdown(f"<div class='hero-grid'>{cards}</div>", unsafe_allow_html=True)

    gs_row, ns_row = _overall_row(gs), _overall_row(ns)
    aum_row = _overall_row(aum_frame)
    tiles = [
        {
            "label": "AUM gap to target",
            "value": fmt_cr(aum_row.get("Gap to Target")),
            "secondary": "Target less current AUM",
        },
        {
            "label": "Gross sales shortfall",
            "value": fmt_cr(
                None if _num(gs_row.get("FY27 Target")) is None
                else _z(gs_row.get("FY27 Target")) - _z(gs_row.get("YTD"))
            ),
            "secondary": "Still to book by March 2027",
        },
        {
            "label": "Net sales shortfall",
            "value": fmt_cr(
                None if _num(ns_row.get("FY27 Target")) is None
                else _z(ns_row.get("FY27 Target")) - _z(ns_row.get("YTD"))
            ),
            "secondary": "Still to book by March 2027",
        },
        {
            "label": "Net projected FY",
            "value": fmt_pct(ns_row.get("Projected FY %")),
            "delta": fmt_pts(
                None if _num(ns_row.get("Projected FY %")) is None
                else _num(ns_row.get("Projected FY %")) - 1.0
            ),
            "tone": _tone_for(
                None if _num(ns_row.get("Projected FY %")) is None
                else _num(ns_row.get("Projected FY %")) - 1.0
            ),
            "secondary": "At the current run rate",
        },
    ]
    kpi_strip(tiles)
    glass_note(
        "AUM, Gross Sales and Net Sales are read directly from the workbook's "
        "<b>FINAL</b> sheet. Achievement, run rate and projected FY are derived from those "
        "same Target and YTD values."
    )

def render_business_driver_selector(
    final_metrics: Dict[str, Any],
    model: ScenarioModel,
) -> None:
    """03 · Business drivers - shown only by Asset Class and Channel."""
    section_header(
        "03",
        "Business drivers",
        "Performance is segregated only by asset class and channel",
    )

    st.markdown("<div class='metric-label'>Sales basis</div>", unsafe_allow_html=True)
    basis_label = st.radio(
        "Sales basis",
        [SALES_LABEL["GS"], SALES_LABEL["NS"]],
        index=1,
        horizontal=True,
        key="driver_basis",
        label_visibility="collapsed",
    )
    basis = "GS" if basis_label == SALES_LABEL["GS"] else "NS"
    st.session_state["display_basis"] = basis

    frame = final_sales_metrics(final_metrics, model, basis)
    if frame.empty:
        glass_note("Driver metrics could not be located on the FINAL sheet.")
        return

    overall = _overall_row(frame)
    projected = _num(overall.get("Projected FY %"))
    kpi_strip([
        {"label": "FY27 target", "value": fmt_cr(overall.get("FY27 Target"))},
        {
            "label": "YTD",
            "value": fmt_cr(overall.get("YTD")),
            "delta": fmt_pct(overall.get("Achievement %")),
            "tone": "gold",
            "secondary": "of FY27 target booked",
        },
        {"label": "Current run rate", "value": fmt_cr(overall.get("Current RR"))},
        {
            "label": "Required run rate",
            "value": fmt_cr(overall.get("Required RR to Target")),
            "secondary": "(Target − YTD) ÷ 9 remaining months",
        },
        {
            "label": "Projected FY",
            "value": fmt_pct(projected),
            "delta": fmt_pts(None if projected is None else projected - 1.0),
            "tone": _tone_for(None if projected is None else projected - 1.0),
        },
    ])

    display_columns = [
        "FY27 Target", "YTD", "Achievement %", "Current RR",
        "Required RR to Target", "Estimated FY @ Current RR", "Projected FY %",
    ]
    formats = {
        "Scope": "txt", "FY27 Target": "cr", "YTD": "cr", "Achievement %": "pct",
        "Current RR": "cr", "Required RR to Target": "cr",
        "Estimated FY @ Current RR": "cr", "Projected FY %": "pct",
    }

    asset_frame = frame.loc[[i for i in FINAL_ASSET_ROWS if i in frame.index]]
    channel_frame = frame.loc[[i for i in FINAL_CHANNEL_ROWS if i in frame.index]]

    tabs = st.tabs(["Asset class", "Channel"])
    with tabs[0]:
        if asset_frame.empty:
            glass_note("Asset-class metrics could not be located on the FINAL sheet.")
        else:
            render_glass_table(
                asset_frame.reset_index().rename(columns={"Metric": "Scope"})[["Scope", *display_columns]],
                formats,
            )
    with tabs[1]:
        if channel_frame.empty:
            glass_note("Channel metrics could not be located on the FINAL sheet.")
        else:
            render_glass_table(
                channel_frame.reset_index().rename(columns={"Metric": "Scope"})[["Scope", *display_columns]],
                formats,
            )

# =============================================================================
# 16. ANALYSIS SCOPE & CURRENT RUN RATE
# =============================================================================
 
LOCATION_FIELD_ALIASES = ["MKT TYPE", "Market Type", "Mkt Type"]
 
 
def _location_column(records: pd.DataFrame) -> Optional[str]:
    for c in LOCATION_FIELD_ALIASES:
        if c in records.columns:
            return c
    return None
 
 
def _location_options(records: pd.DataFrame, channel: str) -> List[str]:
    col = _location_column(records)
    if col is None:
        return ["All"]
    work = records
    if channel != "All" and "Vertical" in work.columns:
        work = work.loc[work["Vertical"] == channel]
    values = sorted({str(v).strip() for v in work[col].dropna().tolist() if str(v).strip()})
    # Preserve the workbook's own location cuts, including T2/T6/T30/B30/EM.
    return ["All"] + values
 
 
def apply_management_cuts(records: pd.DataFrame, channel: str, location: str) -> pd.DataFrame:
    out = records.copy()
    if channel != "All" and "Vertical" in out.columns:
        out = out.loc[out["Vertical"] == channel].copy()
    col = _location_column(out)
    if location != "All" and col is not None:
        out = out.loc[out[col].astype(str).str.strip() == location].copy()
    return out
 
 
def render_analysis_scope(records: pd.DataFrame) -> Tuple[str, str, str, pd.DataFrame]:
    """
    Scope controls that recalculate the analytical engine.

    Channel and location remain as existing filters. Asset-class results are
    shown together downstream instead of forcing an Equity / Debt / Liquid drill-down.
    """
    st.markdown("<div class='metric-label'>Analysis scope</div>", unsafe_allow_html=True)
    left, middle = st.columns([1, 1])

    with left:
        present = set(records.get("Vertical", pd.Series(dtype=str)).astype(str))
        channel_options = ["All"] + [v for v in VERTICALS if v in present]
        reset_stale_selection("scope_channel", channel_options)
        channel = st.selectbox("Channel", channel_options, index=0, key="scope_channel")
    with middle:
        location_options = _location_options(records, channel)
        reset_stale_selection("scope_location", location_options)
        location = st.selectbox(
            "Location / market type", location_options, index=0, key="scope_location"
        )

    # Asset classes stay together; no Equity / Debt / Liquid selector.
    asset = "All"
    filtered = apply_management_cuts(records, channel, location)

    active = [x for x in (
        channel if channel != "All" else None,
        location if location != "All" else None,
    ) if x]
    scope_text = " · ".join(active) if active else "All business"
    st.markdown(
        f"<div class='glass-note'>Scope <span class='inline-pill gold'>{escape(scope_text)}</span>"
        f"<span class='inline-pill'>{len(filtered):,} RMs</span>"
        "Every scenario below is recalculated on this population.</div>",
        unsafe_allow_html=True,
    )

    if filtered.empty:
        glass_callout(
            "No RM records match this scope. Widen the channel or location selection "
            "to bring the calculation engine back online.",
            tone="warn",
        )
    return channel, location, asset, filtered

def render_current_runrate(grid: pd.DataFrame, basis: str, asset: str) -> None:
    """04 · Current pace, required pace and projected finish. Graph removed."""
    section_header(
        "04",
        "Current run rate & target gap",
        "The pace the business is running at, against the pace the target needs",
    )

    if grid.empty:
        glass_note("No records in scope, so the current run rate cannot be calculated.")
        return

    cell = summarize_current(grid, sales=basis, asset=None if asset == "All" else asset)
    required_rr = (
        _z(cell.get("fy_target")) - _z(cell.get("ytd_ach"))
    ) / max(MONTHS_REMAINING, 1)
    current_rr = _num(cell.get("current_rr"))
    pace_gap = None
    if current_rr is not None and required_rr:
        pace_gap = current_rr / required_rr - 1.0
    projected = _num(cell.get("current_march_pct"))
    shortfall = _z(cell.get("fy_target")) - _z(cell.get("current_march"))

    kpi_strip([
        {"label": "FY target", "value": fmt_cr(cell.get("fy_target")),
         "secondary": f"{SALES_LABEL[basis]} · RM calculation sheets"},
        {"label": "YTD achievement", "value": fmt_cr(cell.get("ytd_ach")),
         "delta": fmt_pct(cell.get("ytd_ach_pct")), "tone": "gold",
         "secondary": "against the YTD June target"},
        {"label": "Current run rate", "value": fmt_cr(current_rr),
         "secondary": "YTD ÷ 3 completed months"},
        {"label": "Required run rate", "value": fmt_cr(required_rr),
         "delta": fmt_pct_signed(pace_gap), "tone": _tone_for(pace_gap),
         "secondary": "remaining FY target ÷ remaining months"},
        {"label": "Projected March", "value": fmt_pct(projected),
         "delta": fmt_pts(None if projected is None else projected - 1.0),
         "tone": _tone_for(None if projected is None else projected - 1.0),
         "secondary": fmt_cr(cell.get("current_march"))},
        {"label": "Gap at current pace", "value": fmt_cr(shortfall),
         "tone": _tone_for(-shortfall),
         "secondary": "FY target less projected March"},
    ])

    glass_note(
        "The current run rate is the completed Apr–Jun achievement divided by three. "
        "The required run rate is the remaining FY target (Target − YTD) divided by the 9 remaining months."
    )

# =============================================================================
# 17. SCENARIO PLANNING COMPONENTS
# =============================================================================
 
def render_scenario_navigator() -> int:
    """05 · Scenario planning header; selection now lives only in the sidebar."""
    section_header(
        "05",
        "Scenario planning",
        "What changes if the organisation changes the trajectory",
    )
    scenario_id = int(st.session_state.get("scenario_id", 1))
    if scenario_id not in SCENARIO_ORDER:
        scenario_id = 1
        st.session_state["scenario_id"] = scenario_id
    return scenario_id
 
 
def _pct_input(label: str, default_fraction: float, key: str, min_value: float = 0.0, max_value: float = 300.0, step: float = 1.0) -> float:
    """Editable percentage field; returns a fraction used by the scenario engine."""
    value = st.number_input(
        label,
        min_value=float(min_value),
        max_value=float(max_value),
        value=float(default_fraction * 100.0),
        step=float(step),
        format="%.1f",
        key=key,
    )
    return float(value) / 100.0



def _scenario10_slug(value: Any) -> str:
    raw = str(value).strip().casefold()
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in raw)
    return "_".join(part for part in cleaned.split("_") if part) or "all"


def _scenario10_scope_prefix() -> str:
    # Scenario 10 is portfolio-level and FINAL-sourced; RM scope filters do not
    # change its source numbers.
    return "s10_final"


def _scenario10_current_pct(
    grid: pd.DataFrame,
    sales: str,
    asset: str,
    vertical: Optional[str] = None,
    market_type: Optional[str] = None,
) -> float:
    if market_type is not None:
        return _scenario10_market_current_pct(grid, sales, asset, market_type)
    stats = _scenario10_current_stats(
        grid,
        sales,
        asset,
        planning_channel=vertical,
    )
    return _scenario10_safe_current_pct(stats)


def _scenario10_current_number(
    grid: pd.DataFrame,
    sales: str,
    asset: str,
    vertical: Optional[str] = None,
    market_type: Optional[str] = None,
) -> float:
    if market_type is not None:
        stats = _scenario10_market_stats(grid, sales, asset, market_type)
    else:
        stats = _scenario10_current_stats(grid, sales, asset, planning_channel=vertical)
    return _z(stats.get("current_march"))


def _scenario10_fy_target(
    grid: pd.DataFrame,
    sales: str,
    asset: str,
    vertical: Optional[str] = None,
    market_type: Optional[str] = None,
) -> float:
    if market_type is not None:
        return _scenario10_market_fy_target(grid, sales, asset, market_type)
    subset = _scenario10_subset(
        grid,
        sales=sales,
        asset=asset,
        planning_channel=vertical,
    )
    return float(subset["fy_target"].sum()) if not subset.empty else 0.0


def _scenario10_factor_key(prefix: str, asset: str, scope: str) -> str:
    return f"{prefix}_factor_{_scenario10_slug(asset)}_{_scenario10_slug(scope)}"


def _scenario10_get_factor(prefix: str, asset: str, scope: str) -> float:
    value = _num(st.session_state.get(_scenario10_factor_key(prefix, asset, scope), 1.0))
    return max(0.0, 1.0 if value is None else float(value))


def _scenario10_set_factor(prefix: str, asset: str, scope: str, factor: float) -> None:
    st.session_state[_scenario10_factor_key(prefix, asset, scope)] = max(0.0, float(factor))


def _scenario10_linked_target(
    current_pct: float,
    factor: float,
    fallback_pct: float = 0.0,
) -> float:
    if current_pct > 0:
        return max(0.0, current_pct * max(0.0, factor))
    return max(float(fallback_pct), 0.0)


def _scenario10_prepare_basis_switch(prefix: str, edit_basis: str) -> None:
    previous = st.session_state.get(f"{prefix}_last_basis")
    if previous != edit_basis:
        st.session_state[f"{prefix}_last_basis"] = edit_basis


def _scenario10_channel_input(
    grid: pd.DataFrame,
    prefix: str,
    edit_basis: str,
    asset: str,
    channel: str,
) -> Tuple[float, float, float]:
    """Edit FINAL achievement %, link the opposite sales basis by relative change."""
    other_basis = "GS" if edit_basis == "NS" else "NS"
    current_primary = _scenario10_current_pct(grid, edit_basis, asset, channel)
    current_other = _scenario10_current_pct(grid, other_basis, asset, channel)
    factor = _scenario10_get_factor(prefix, asset, channel)
    default_primary = (
        current_primary * factor if current_primary > 0 else max(current_primary, 0.0)
    )

    widget_key = (
        f"{prefix}_widget_{edit_basis.lower()}_{_scenario10_slug(asset)}_"
        f"{_scenario10_slug(channel)}"
    )
    if widget_key not in st.session_state:
        st.session_state[widget_key] = max(default_primary, 0.0) * 100.0

    label = S10_CHANNEL_LABELS.get(channel, channel)
    value = st.number_input(
        f"{label} simulation %",
        min_value=0.0,
        value=float(st.session_state[widget_key]),
        step=1.0,
        format="%.1f",
        key=widget_key,
        help=(
            f"Current is read directly from FINAL: Current ÷ FY27 Budget = "
            f"{current_primary:.1%}. Projected Number = this simulation % × FY27 Budget."
        ),
    ) / 100.0

    if current_primary > 0:
        factor = max(0.0, value / current_primary)
    else:
        factor = 1.0
    _scenario10_set_factor(prefix, asset, channel, factor)

    linked = (
        max(0.0, current_other * factor)
        if current_other > 0 and current_primary > 0
        else max(0.0, value)
    )
    return value, linked, factor


def _scenario10_market_location_input(
    grid: pd.DataFrame,
    prefix: str,
    edit_basis: str,
    asset: str,
    location: str,
) -> Tuple[float, float, float]:
    """Edit a FINAL T2/T6/T30/B30/EM achievement percentage."""
    other_basis = "GS" if edit_basis == "NS" else "NS"
    current_primary = _scenario10_market_current_pct(grid, edit_basis, asset, location)
    current_other = _scenario10_market_current_pct(grid, other_basis, asset, location)
    scope = f"Market::{location}"
    factor = _scenario10_get_factor(prefix, asset, scope)
    default_primary = current_primary * factor if current_primary > 0 else max(current_primary, 0.0)

    widget_key = (
        f"{prefix}_widget_{edit_basis.lower()}_{_scenario10_slug(asset)}_market_"
        f"{_scenario10_slug(location)}"
    )
    if widget_key not in st.session_state:
        st.session_state[widget_key] = max(default_primary, 0.0) * 100.0

    value = st.number_input(
        f"{location} simulation %",
        min_value=0.0,
        value=float(st.session_state[widget_key]),
        step=1.0,
        format="%.1f",
        key=widget_key,
        help=(
            f"{location} Current and FY27 Budget are read from FINAL. "
            "Projected Number = simulation % × FY27 Budget."
        ),
    ) / 100.0

    if current_primary > 0:
        factor = max(0.0, value / current_primary)
    else:
        factor = 1.0
    _scenario10_set_factor(prefix, asset, scope, factor)

    linked = (
        max(0.0, current_other * factor)
        if current_other > 0 and current_primary > 0
        else max(0.0, value)
    )
    return value, linked, factor


def _scenario10_build_target_maps(
    grid: pd.DataFrame,
    prefix: str,
    edit_basis: str,
    primary_channel_targets: Dict[str, Dict[str, float]],
) -> Tuple[
    Dict[str, Dict[str, Dict[str, float]]],
    Dict[str, Dict[str, Dict[str, float]]],
    Dict[str, Dict[str, Dict[str, float]]],
]:
    """Build channel and market simulation percentages for both NS and GS."""
    channel_targets: Dict[str, Dict[str, Dict[str, float]]] = {
        sales: {asset: {} for asset in ASSETS} for sales in SALES_TYPES
    }
    market_targets: Dict[str, Dict[str, Dict[str, float]]] = {
        sales: {asset: {} for asset in ASSETS} for sales in SALES_TYPES
    }

    for asset in ASSETS:
        for channel in S10_PLANNING_CHANNELS:
            factor = _scenario10_get_factor(prefix, asset, channel)
            for sales in SALES_TYPES:
                current = _scenario10_current_pct(grid, sales, asset, channel)
                if sales == edit_basis:
                    value = primary_channel_targets.get(asset, {}).get(channel, current)
                else:
                    primary_current = _scenario10_current_pct(grid, edit_basis, asset, channel)
                    primary_value = primary_channel_targets.get(asset, {}).get(channel, primary_current)
                    value = (
                        current * factor
                        if current > 0 and primary_current > 0
                        else primary_value
                    )
                channel_targets[sales][asset][channel] = max(0.0, float(value))

        for location in _scenario10_retail_locations(grid):
            factor = _scenario10_get_factor(prefix, asset, f"Market::{location}")
            for sales in SALES_TYPES:
                current = _scenario10_market_current_pct(grid, sales, asset, location)
                market_targets[sales][asset][location] = max(0.0, current * factor if current > 0 else current)

    # Kept for backward-compatible function shape; Scenario 10 no longer applies
    # market factors to channel rows because FINAL does not provide a channel × market cross-tab.
    return channel_targets, {}, market_targets


def _scenario10_preview_asset_pct(
    grid: pd.DataFrame,
    sales: str,
    asset: str,
    vertical_targets: Dict[str, Dict[str, Dict[str, float]]],
    market_factors: Dict[str, Dict[str, Dict[str, float]]],
) -> float:
    subset = _scenario10_subset(grid, sales=sales, asset=asset)
    if subset.empty:
        return 0.0
    target_total = float(subset["fy_target"].sum())
    if target_total == 0:
        return 0.0
    projected = 0.0
    for row in subset.to_dict("records"):
        channel = _scenario10_management_channel(row)
        pct = vertical_targets.get(sales, {}).get(asset, {}).get(channel, 0.0)
        projected += _z(row.get("fy_target")) * pct
    return projected / target_total


def render_scenario_controls(
    scenario_id: int,
    base_params: Dict[str, Any],
    grid: Optional[pd.DataFrame] = None,
    final_metrics: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Editable percentage assumptions for every scenario."""
    params = dict(base_params)

    with st.expander("Scenario assumptions · editable", expanded=True):
        if scenario_id == 1:
            params["runrate_uplift"] = _pct_input(
                "Run-rate uplift %", float(base_params.get("runrate_uplift", S1_RUNRATE_UPLIFT)),
                "s1_uplift", 0.0, 200.0,
            )

        elif scenario_id == 2:
            columns = st.columns(2)
            with columns[0]:
                params["overall_target"] = _pct_input(
                    "Overall target by January %", float(base_params.get("overall_target", S2_OVERALL_TARGET)),
                    "s2_overall", 0.0, 250.0,
                )
            with columns[1]:
                params["equity_target"] = _pct_input(
                    "Equity target by January %", float(base_params.get("equity_target", S2_EQUITY_TARGET)),
                    "s2_equity", 0.0, 300.0,
                )

        elif scenario_id == 3:
            columns = st.columns(2)
            with columns[0]:
                params["target_pct"] = _pct_input(
                    "January target achievement %", float(base_params.get("target_pct", S3_TARGET)),
                    "s3_target", 0.0, 300.0,
                )
            with columns[1]:
                params["dip"] = _pct_input(
                    "February–March run-rate dip %", float(base_params.get("dip", S3_DEFAULT_DIP)),
                    "s3_dip", 0.0, 100.0,
                )

        elif scenario_id == 4:
            params["target_pct"] = _pct_input(
                "March target achievement %", float(base_params.get("target_pct", S4_TARGET)),
                "s4_target", 0.0, 300.0,
            )

        elif scenario_id == 5:
            columns = st.columns(2)
            with columns[0]:
                params["overall_target"] = _pct_input(
                    "Overall March target %", float(base_params.get("overall_target", S5_OVERALL_TARGET)),
                    "s5_overall", 0.0, 300.0,
                )
            with columns[1]:
                params["equity_target"] = _pct_input(
                    "Equity March target %", float(base_params.get("equity_target", S5_EQUITY_TARGET)),
                    "s5_equity", 0.0, 350.0,
                )

        elif scenario_id == 6:
            defaults = dict(base_params.get("segment_targets", S6_SEGMENT_TARGETS))
            columns = st.columns(3)
            edited: Dict[str, float] = {}
            for idx, segment in enumerate(SEGMENT_ORDER):
                with columns[idx]:
                    edited[segment] = _pct_input(
                        f"{segment} March target %",
                        float(defaults.get(segment, S6_SEGMENT_TARGETS.get(segment, 1.0))),
                        f"s6_{segment.lower().replace(' ', '_')}", 0.0, 400.0,
                    )
            params["segment_targets"] = edited

        elif scenario_id == 7:
            columns = st.columns(3)
            with columns[0]:
                params["jan_target"] = _pct_input(
                    "January achievement target %", float(base_params.get("jan_target", S7_DEFAULT_JAN_TARGET)),
                    "s7_jan", 0.0, 300.0,
                )
            with columns[1]:
                params["mar_target"] = _pct_input(
                    "March achievement target %", float(base_params.get("mar_target", S7_DEFAULT_MAR_TARGET)),
                    "s7_mar", 0.0, 300.0,
                )
            with columns[2]:
                params["leakage"] = _pct_input(
                    "February–March leakage %", float(base_params.get("leakage", S7_DEFAULT_LEAKAGE)),
                    "s7_leak", 0.0, 100.0,
                )

        elif scenario_id == 8:
            params["leakage"] = _pct_input(
                "February–March leakage %", float(base_params.get("leakage", S8_DEFAULT_LEAKAGE)),
                "s8_leakage", 0.0, 100.0,
            )
            growth = dict(base_params.get("channel_growth", S8_DEFAULT_GROWTH))
            jan_targets = dict(base_params.get("channel_jan_target", S8_DEFAULT_JAN_TARGET))
            mar_targets = dict(base_params.get("channel_mar_target", S8_DEFAULT_MAR_TARGET))
            st.markdown("<div class='metric-secondary'>Monthly growth % · January target % · March target %</div>", unsafe_allow_html=True)
            for channel in CHANNELS:
                st.markdown(f"<div class='metric-label' style='margin-top:12px'>{escape(channel)}</div>", unsafe_allow_html=True)
                columns = st.columns(3)
                with columns[0]:
                    growth[channel] = _pct_input(
                        "Monthly growth %", float(growth.get(channel, S8_DEFAULT_GROWTH.get(channel, 0.05))),
                        f"s8_g_{channel}", -50.0, 200.0,
                    )
                with columns[1]:
                    jan_targets[channel] = _pct_input(
                        "January target %", float(jan_targets.get(channel, S8_DEFAULT_JAN_TARGET.get(channel, 1.0))),
                        f"s8_j_{channel}", 0.0, 400.0,
                    )
                with columns[2]:
                    mar_targets[channel] = _pct_input(
                        "March target %", float(mar_targets.get(channel, S8_DEFAULT_MAR_TARGET.get(channel, 1.0))),
                        f"s8_m_{channel}", 0.0, 400.0,
                    )
            params["channel_growth"] = growth
            params["channel_jan_target"] = jan_targets
            params["channel_mar_target"] = mar_targets

        elif scenario_id == 9:
            columns = st.columns(3)
            with columns[0]:
                jan_target = _pct_input("Portfolio January milestone %", 1.00, "s9_jan_target", 0.0, 300.0)
            with columns[1]:
                ambition = _pct_input(
                    "Portfolio March ambition %", float(base_params.get("optimizer_target", 1.20)),
                    "s9_target", 0.0, 400.0,
                )
            with columns[2]:
                params["leakage"] = _pct_input(
                    "February–March leakage %", float(base_params.get("leakage", S8_DEFAULT_LEAKAGE)),
                    "s9_leakage", 0.0, 100.0,
                )
            params["optimizer_target"] = ambition
            params["channel_mar_target"] = {c: ambition for c in CHANNELS}
            params["channel_jan_target"] = {c: jan_target for c in CHANNELS}

        elif scenario_id == 10:
            if grid is None or grid.empty:
                glass_callout(
                    "<b>Scenario 10 cannot start:</b> the detailed GS/NS matrix could not be read "
                    "from FINAL. Scenario 10 does not fall back to the RM calculation sheets.",
                    tone="warn",
                )
                params["asset_vertical_targets"] = {}
                params["retail_location_targets"] = {}
                params["market_location_factors"] = {}
            else:
                prefix = _scenario10_scope_prefix()
                edit_label = st.radio(
                    "Edit sales basis",
                    [SALES_LABEL["NS"], SALES_LABEL["GS"]],
                    index=0,
                    horizontal=True,
                    key=f"{prefix}_basis_selector",
                    help=(
                        "Scenario 10 reads Current and FY27 Budget only from FINAL. "
                        "Edit one basis; the other basis follows the same relative change when a positive current ratio exists."
                    ),
                )
                edit_basis = "NS" if edit_label == SALES_LABEL["NS"] else "GS"
                other_basis = "GS" if edit_basis == "NS" else "NS"
                _scenario10_prepare_basis_switch(prefix, edit_basis)

                glass_callout(
                    "<b>Scenario 10 · FINAL-sheet simulation · EX-DIGITAL</b><br>"
                    "Current = the YTD/current number printed in FINAL. FY27 Budget = the Target printed in FINAL. "
                    "The app does <b>not</b> annualise YTD for this scenario."
                )
                glass_note(
                    "Calculation used everywhere in Scenario 10: "
                    "<b>Simulation Projected Number = Simulation Achievement % × FY27 Budget</b>. "
                    "Current % = FINAL Current ÷ FINAL FY27 Budget. Digital is excluded from the roll-up."
                )

                primary_channel_targets: Dict[str, Dict[str, float]] = {}
                locations = _scenario10_retail_locations(grid)

                for asset in ASSETS:
                    st.markdown(
                        f"<div class='glass-panel' style='margin-top:16px'>"
                        f"<div class='metric-label'>{escape(asset)} · FINAL channel simulation</div>"
                        f"<div class='metric-secondary'>Editable: {escape(SALES_LABEL[edit_basis])} · "
                        f"Linked: {escape(SALES_LABEL[other_basis])}</div></div>",
                        unsafe_allow_html=True,
                    )

                    primary_for_asset: Dict[str, float] = {}
                    channel_groups = [["Retail", "DHNI", "VRM"], ["Institutional"]]

                    for channel_group in channel_groups:
                        columns = st.columns(len(channel_group))
                        for idx, channel in enumerate(channel_group):
                            with columns[idx]:
                                primary, linked, factor = _scenario10_channel_input(
                                    grid, prefix, edit_basis, asset, channel
                                )
                                primary_for_asset[channel] = primary

                                current_number = _scenario10_current_number(
                                    grid, edit_basis, asset, channel
                                )
                                budget = _scenario10_fy_target(
                                    grid, edit_basis, asset, channel
                                )
                                current_pct = _scenario10_current_pct(
                                    grid, edit_basis, asset, channel
                                )
                                projected_number = budget * primary
                                st.markdown(
                                    "<div class='glass-note'>"
                                    f"<b>FINAL Current:</b> {escape(fmt_cr(current_number))}<br>"
                                    f"<b>FY27 Budget:</b> {escape(fmt_cr(budget))}<br>"
                                    f"Current %: {escape(fmt_pct(current_pct))}<br>"
                                    f"Simulation projected: <b>{escape(fmt_cr(projected_number))}</b><br>"
                                    f"{escape(SALES_LABEL[other_basis])} linked: {escape(fmt_pct(linked))}"
                                    "</div>",
                                    unsafe_allow_html=True,
                                )

                    primary_channel_targets[asset] = primary_for_asset

                    if locations:
                        st.markdown(
                            "<div class='metric-label' style='margin-top:14px'>"
                            "Market type · FINAL source · Retail + VRM + DHNI cut</div>",
                            unsafe_allow_html=True,
                        )
                        selected_location = st.selectbox(
                            "Market type",
                            locations,
                            index=0,
                            key=f"{prefix}_location_select_{_scenario10_slug(asset)}",
                            help="B30 includes B30 Select; T30 includes T30 Ext.",
                        )
                        loc_primary, loc_linked, loc_factor = _scenario10_market_location_input(
                            grid,
                            prefix,
                            edit_basis,
                            asset,
                            selected_location,
                        )
                        loc_budget = _scenario10_market_fy_target(
                            grid, edit_basis, asset, selected_location
                        )
                        loc_current = _scenario10_current_number(
                            grid, edit_basis, asset, market_type=selected_location
                        )
                        loc_current_pct = _scenario10_market_current_pct(
                            grid, edit_basis, asset, selected_location
                        )
                        st.markdown(
                            "<div class='glass-note'>"
                            f"<b>{escape(selected_location)}</b> · FINAL Current {escape(fmt_cr(loc_current))} · "
                            f"FY27 Budget {escape(fmt_cr(loc_budget))}<br>"
                            f"Current % {escape(fmt_pct(loc_current_pct))} · "
                            f"Simulation {escape(fmt_pct(loc_primary))} · "
                            f"Projected Number <b>{escape(fmt_cr(loc_budget * loc_primary))}</b><br>"
                            f"{escape(SALES_LABEL[other_basis])} linked {escape(fmt_pct(loc_linked))}"
                            "</div>",
                            unsafe_allow_html=True,
                        )
                    else:
                        glass_note(
                            "T2 / T6 / T30 / B30 / EM values were not found as a repeated Target/YTD matrix in FINAL. "
                            "Scenario 10 will not manufacture that market split from the RM sheets."
                        )

                vertical_targets, market_factors, market_targets = _scenario10_build_target_maps(
                    grid, prefix, edit_basis, primary_channel_targets
                )
                params["asset_vertical_targets"] = vertical_targets
                params["market_location_factors"] = market_factors
                params["retail_location_targets"] = market_targets
                params["scenario10_edit_basis"] = edit_basis
                params["scenario10_source"] = "FINAL only"
                params["scenario10_excludes_digital"] = True

                st.markdown(
                    "<div class='metric-secondary' style='margin-top:18px'>"
                    "Live Equity / Debt / Liquid roll-up from FINAL budgets and the simulation percentages"
                    "</div>",
                    unsafe_allow_html=True,
                )
                preview_cards: List[Dict[str, Any]] = []
                for asset in ASSETS:
                    primary_pct = _scenario10_preview_asset_pct(
                        grid, edit_basis, asset, vertical_targets, market_factors
                    )
                    linked_pct = _scenario10_preview_asset_pct(
                        grid, other_basis, asset, vertical_targets, market_factors
                    )
                    current_primary = _scenario10_current_pct(grid, edit_basis, asset)
                    target_amount = _scenario10_fy_target(grid, edit_basis, asset)
                    projected_amount = target_amount * primary_pct
                    preview_cards.append({
                        "label": asset,
                        "value": fmt_pct(primary_pct),
                        "delta": fmt_pts(primary_pct - current_primary),
                        "tone": _tone_for(primary_pct - current_primary),
                        "secondary": (
                            f"{fmt_cr(projected_amount)} projected ÷ {fmt_cr(target_amount)} budget · "
                            f"{SALES_LABEL[other_basis]} linked {fmt_pct(linked_pct)}"
                        ),
                    })
                kpi_strip(preview_cards)

        glass_note(
            "All percentages in the selected scenario are editable. Changing a value recalculates "
            "required run rates, expected sales and revenue immediately."
        )

    return params


def _active_scenario_assumption_text(model: ScenarioModel) -> str:
    """Dynamic assumption summary so edited percentages are visible beside the scenario."""
    p = model.params
    sid = model.scenario_id
    if sid == 1:
        return f"Run-rate uplift: {fmt_pct(p.get('runrate_uplift', S1_RUNRATE_UPLIFT))}"
    if sid == 2:
        return f"January overall: {fmt_pct(p.get('overall_target', S2_OVERALL_TARGET))} · January Equity: {fmt_pct(p.get('equity_target', S2_EQUITY_TARGET))}"
    if sid == 3:
        return f"January target: {fmt_pct(p.get('target_pct', S3_TARGET))} · Feb–Mar dip: {fmt_pct(p.get('dip', S3_DEFAULT_DIP))}"
    if sid == 4:
        return f"March target: {fmt_pct(p.get('target_pct', S4_TARGET))}"
    if sid == 5:
        return f"March overall: {fmt_pct(p.get('overall_target', S5_OVERALL_TARGET))} · March Equity: {fmt_pct(p.get('equity_target', S5_EQUITY_TARGET))}"
    if sid == 6:
        targets = dict(p.get('segment_targets', S6_SEGMENT_TARGETS))
        return " · ".join(f"{s}: {fmt_pct(targets.get(s))}" for s in SEGMENT_ORDER)
    if sid == 7:
        return f"January: {fmt_pct(p.get('jan_target'))} · March: {fmt_pct(p.get('mar_target'))} · Leakage: {fmt_pct(p.get('leakage'))}"
    if sid == 8:
        return f"Channel assumptions editable · Leakage: {fmt_pct(p.get('leakage'))}"
    if sid == 9:
        jan = next(iter(p.get('channel_jan_target', {}).values()), None)
        return f"January milestone: {fmt_pct(jan)} · March ambition: {fmt_pct(p.get('optimizer_target'))} · Leakage: {fmt_pct(p.get('leakage'))}"
    if sid == 10:
        return "FINAL-only simulation · Current = FINAL YTD · Projected Number = Simulation % × FY27 Budget · Digital excluded"
    return ""


def _active_scenario_copy(model: ScenarioModel) -> Dict[str, str]:
    """Return scenario title/thesis/explanation using the live editable percentages."""
    p = model.params
    sid = model.scenario_id

    if sid == 1:
        uplift = fmt_pct(p.get("runrate_uplift", S1_RUNRATE_UPLIFT))
        return {
            "name": f"+{uplift} Run-Rate Push",
            "thesis": f"Lift the current pace by {uplift} and carry that pace through the remaining nine months.",
            "explanation": f"Increase the current Apr–Jun monthly run rate by {uplift} from July onward and measure the resulting March achievement.",
        }
    if sid == 2:
        overall = fmt_pct(p.get("overall_target", S2_OVERALL_TARGET))
        equity = fmt_pct(p.get("equity_target", S2_EQUITY_TARGET))
        return {
            "name": f"{overall} Overall by January + {equity} Equity",
            "thesis": f"Reach {equity} of the Equity FY target while carrying {overall} of the overall book by January.",
            "explanation": f"Reach {equity} of the Equity FY target and {overall} of the overall FY target by January. The residual requirement is allocated to Debt and Liquid in FY-target proportion.",
        }
    if sid == 3:
        target = fmt_pct(p.get("target_pct", S3_TARGET))
        dip = fmt_pct(p.get("dip", S3_DEFAULT_DIP))
        return {
            "name": f"{target} by January, then {dip} Feb–Mar dip",
            "thesis": f"Reach {target} of the FY target by January, then absorb a {dip} closing run-rate dip.",
            "explanation": f"Reach {target} of the FY target by January, followed by a {dip} February–March run-rate decline.",
        }
    if sid == 4:
        target = fmt_pct(p.get("target_pct", S4_TARGET))
        return {
            "name": f"{target} by March",
            "thesis": f"Hold the required pace for the remaining nine months and close the year at {target}.",
            "explanation": f"Determine the monthly run rate required to finish March at {target} of the FY target.",
        }
    if sid == 5:
        overall = fmt_pct(p.get("overall_target", S5_OVERALL_TARGET))
        equity = fmt_pct(p.get("equity_target", S5_EQUITY_TARGET))
        return {
            "name": f"{equity} Equity + {overall} Overall by March",
            "thesis": f"Push Equity to {equity} while the overall portfolio lands at {overall} by March.",
            "explanation": f"Reach {equity} of the Equity FY target and {overall} of the overall FY target by March, with Debt and Liquid balancing the remaining requirement.",
        }
    if sid == 6:
        targets = dict(p.get("segment_targets", S6_SEGMENT_TARGETS))
        digital = fmt_pct(targets.get("Digital", S6_SEGMENT_TARGETS["Digital"]))
        b30 = fmt_pct(targets.get("Retail B30", S6_SEGMENT_TARGETS["Retail B30"]))
        others = fmt_pct(targets.get("Others", S6_SEGMENT_TARGETS["Others"]))
        return {
            "name": f"Digital {digital} + Retail B30 {b30} + Others {others}",
            "thesis": f"Set differentiated March outcomes: Digital {digital}, Retail B30 {b30}, Others {others}.",
            "explanation": f"Model differentiated performance where Digital achieves {digital}, Retail B30 achieves {b30}, and Others achieve {others} of their respective FY targets.",
        }
    if sid == 7:
        jan = fmt_pct(p.get("jan_target", S7_DEFAULT_JAN_TARGET))
        mar = fmt_pct(p.get("mar_target", S7_DEFAULT_MAR_TARGET))
        leakage = fmt_pct(p.get("leakage", S7_DEFAULT_LEAKAGE))
        return {
            "name": f"Momentum Build-Up · Jan {jan} → Mar {mar}",
            "thesis": f"Build a January buffer to reach {jan}, absorb {leakage} leakage, and protect a {mar} March outcome.",
            "explanation": f"Build progressive month-on-month momentum from July 2026 to reach {jan} by January 2027, absorb {leakage} February–March run-rate leakage, and finish March at the {mar} ambition.",
        }
    if sid == 8:
        leakage = fmt_pct(p.get("leakage", S8_DEFAULT_LEAKAGE))
        return {
            "name": "Channel Growth & Target Simulator",
            "thesis": f"Set channel-by-channel growth and target percentages, with {leakage} February–March leakage.",
            "explanation": f"Independently adjust monthly growth, January target achievement and March target achievement for all nine channels. The current leakage assumption is {leakage}.",
        }
    if sid == 9:
        jan = next(iter(p.get("channel_jan_target", {}).values()), 1.0)
        jan_text = fmt_pct(jan)
        mar_text = fmt_pct(p.get("optimizer_target", 1.0))
        leakage = fmt_pct(p.get("leakage", S8_DEFAULT_LEAKAGE))
        return {
            "name": f"Channel Mix Optimiser · Mar {mar_text}",
            "thesis": f"Find the minimum channel growth needed for a {mar_text} March ambition while protecting a {jan_text} January milestone.",
            "explanation": f"Optimise the channel growth trajectory to achieve {mar_text} by March, preserve the {jan_text} January milestone, and allow for {leakage} February–March leakage.",
        }
    if sid == 10:
        return {
            "name": "Asset × Channel Target Simulator",
            "thesis": (
                "Edit Retail, DHNI, VRM and Insti inside Equity, Debt and Liquid and watch "
                "the ex-Digital Asset-Class and Overall outcomes update immediately."
            ),
            "explanation": (
                "Scenario 10 uses FINAL as the only source for Current and FY27 Budget. Retail, DHNI, VRM and Insti "
                "are simulated inside Equity, Debt and Liquid while Digital is excluded. Projected Number = Simulation % × "
                "FY27 Budget. T2/T6/T30/B30/EM are read from the FINAL market matrix when present; the opposite sales basis "
                "follows the same relative change when a positive current ratio exists."
            ),
        }
    return {
        "name": model.meta.get("name", "Scenario"),
        "thesis": model.meta.get("thesis", ""),
        "explanation": model.meta.get("explanation", ""),
    }


def render_scenario_hero(model: ScenarioModel, basis: str, asset: str) -> Dict[str, Any]:
    """06 · Scenario hero: the thesis, then Current → January → March."""
    meta = model.meta
    live_copy = _active_scenario_copy(model)
    section_header(
        "06",
        f"Scenario {model.scenario_id} · {live_copy['name']}",
        "The selected strategy and what it demands",
    )
 
    st.markdown(
        "<div class='scenario-hero'>"
        f"<div class='eyebrow'>Scenario {model.scenario_id:02d} · {escape(meta['short'])}</div>"
        f"<div class='title'>{escape(live_copy['name'])}</div>"
        f"<div class='thesis'>{escape(live_copy['thesis'])}</div>"
        f"<div class='detail'>{escape(live_copy['explanation'])}</div>"
        f"<div class='milestone'>{escape(_active_scenario_assumption_text(model))}</div>"
        "</div>",
        unsafe_allow_html=True,
    )
 
    cell = model.cell(basis, asset=None if asset == "All" else asset)
 
    now_rows = "".join([
        _kpi_row_html("YTD booked", fmt_cr(cell.get("ytd_ach"))),
        _kpi_row_html("Current run rate", fmt_cr(cell.get("current_rr"))),
        _kpi_row_html("Projected March", fmt_pct(cell.get("current_march_pct"))),
    ])
    jan_rows = "".join([
        _kpi_row_html("Milestone", fmt_pct(cell.get("milestone_pct")), "gold"),
        _kpi_row_html("Required run rate", fmt_cr(cell.get("scen_rr"))),
        _kpi_row_html(
            "Buffer vs milestone",
            fmt_cr_signed(cell.get("jan_buffer")),
            _tone_for(cell.get("jan_buffer")),
        ),
    ])
    march_rows = "".join([
        _kpi_row_html("Scenario achievement", fmt_cr(cell.get("march_amount"))),
        _kpi_row_html("Feb–Mar run rate", fmt_cr(cell.get("feb_mar_rr"))),
        _kpi_row_html(
            "Headroom vs ambition",
            fmt_cr_signed(cell.get("headroom_amt")),
            _tone_for(cell.get("headroom_amt")),
        ),
    ])
 
    st.markdown(
        "<div class='trio-grid'>"
        "<div class='glass-card stage-card now'><div class='stage'>Current trajectory</div>"
        f"<div class='metric-hero'>{escape(fmt_pct(cell.get('current_march_pct')))}</div>"
        "<div class='metric-label'>Projected March at today's pace</div>"
        f"<div class='kpi-rows'>{now_rows}</div></div>"
        "<div class='glass-card stage-card jan'><div class='stage'>January 2027 milestone</div>"
        f"<div class='metric-hero gold'>{escape(fmt_pct(cell.get('jan_pct')))}</div>"
        "<div class='metric-label'>Of FY target by January</div>"
        f"<div class='kpi-rows'>{jan_rows}</div></div>"
        "<div class='glass-card stage-card mar'><div class='stage'>March 2027 outcome</div>"
        f"<div class='metric-hero'>{escape(fmt_pct(cell.get('march_pct')))}</div>"
        "<div class='metric-label'>Of FY target by March</div>"
        f"<div class='kpi-rows'>{march_rows}</div></div>"
        "</div>",
        unsafe_allow_html=True,
    )
    return cell
 
 
def render_scenario_comparison(
    final_metrics: Dict[str, Any],
    model: ScenarioModel,
    basis: str,
    asset: str,
) -> None:
    """The management answer to: what changes if I choose this scenario?"""
    cell = model.cell(basis, asset=None if asset == "All" else asset)
 
    revenue_total = None
    if asset == "All":
        bundle = revenue_bundle(model, REVENUE_BASIS)
        revenue_total = bundle["incremental"]["total"]
    else:
        scenario_revenue = _z(cell.get("march_amount")) * REVENUE_RATE.get(asset, 0.0)
        baseline_revenue = _z(cell.get("current_march")) * REVENUE_RATE.get(asset, 0.0)
        revenue_total = scenario_revenue - baseline_revenue
 
    current_pct = _num(cell.get("current_march_pct"))
    scenario_pct = _num(cell.get("march_pct"))
    delta = None if (current_pct is None or scenario_pct is None) else scenario_pct - current_pct
 
    current_label = "Current FINAL achievement" if model.scenario_id == 10 else "Current projected FY"
    scenario_label = "Simulation achievement" if model.scenario_id == 10 else "Scenario projected FY"
    kpi_strip([
        {"label": current_label, "value": fmt_pct(current_pct),
         "secondary": fmt_cr(cell.get("current_march"))},
        {"label": scenario_label, "value": fmt_pct(scenario_pct),
         "delta": fmt_pts(delta), "tone": _tone_for(delta),
         "secondary": fmt_cr(cell.get("march_amount"))},
        {"label": "Incremental sales", "value": fmt_cr_signed(cell.get("incremental_sales")),
         "tone": _tone_for(cell.get("incremental_sales")),
         "secondary": "versus the current trajectory"},
        {"label": "Required run rate", "value": fmt_cr(cell.get("scen_rr")),
         "delta": fmt_pct_signed(cell.get("rr_change_pct")),
         "tone": _tone_for(cell.get("rr_change_pct")),
         "secondary": "monthly, from July"},
        {"label": "Revenue impact", "value": fmt_cr_signed(revenue_total, 1),
         "tone": _tone_for(revenue_total),
         "secondary": f"on {SALES_LABEL[REVENUE_BASIS]}, asset-class rates"},
    ], css_class="scenario-kpi-grid")
 
    gs_frame, gs_formats = build_final_scenario_comparison(final_metrics, model, "GS")
    ns_frame, ns_formats = build_final_scenario_comparison(final_metrics, model, "NS")
 
    tabs = st.tabs(["Net Sales · current vs scenario", "Gross Sales · current vs scenario"])
    with tabs[0]:
        render_glass_table(ns_frame, ns_formats, total_rows=("Overall",), css_class="scenario-table")
    with tabs[1]:
        render_glass_table(gs_frame, gs_formats, total_rows=("Overall",), css_class="scenario-table")
 
    if model.scenario_id == 10:
        glass_note(
            "Scenario 10 is FINAL-only: Current = FINAL YTD/current, FY27 Budget = FINAL Target, "
            "and Simulation Projected Number = Simulation % × FY27 Budget. Retail, DHNI, VRM and "
            "Insti are simulated; Digital is shown only as source/current information and is excluded "
            "from the Scenario-10 roll-up."
        )
    else:
        glass_note(
            "Scenario outcomes are anchored to the FINAL FY27 targets so the comparison uses one "
            "common base. Retail, DHNI and VRM carry scenario calculations from their detailed "
            "sheets; channels such as Insti and Digital stay visible as current FINAL metrics only."
        )
 
 
# =============================================================================
# 18. SCENARIO TRAJECTORY, REVENUE & DRIVER COMPONENTS
# =============================================================================
 
def scenario_trajectory(model: ScenarioModel, basis: str, asset: str) -> List[float]:
    """Monthly run rates July -> March for any scenario, on the selected cut."""
    asset_filter = None if asset == "All" else asset
    cell = model.cell(basis, asset=asset_filter)
 
    trajectory = cell.get("trajectory")
    if isinstance(trajectory, (list, tuple)) and len(trajectory) == len(FUTURE_MONTHS):
        return [float(_z(v)) for v in trajectory]
 
    frame = model.scenario_grid
    if frame is not None and "trajectory" in getattr(frame, "columns", []):
        mask = frame["Sales"] == basis
        if asset_filter is not None:
            mask &= frame["Asset"] == asset_filter
        totals = [0.0] * len(FUTURE_MONTHS)
        found = False
        for series in frame.loc[mask, "trajectory"]:
            if isinstance(series, (list, tuple)) and len(series) == len(FUTURE_MONTHS):
                found = True
                for position, value in enumerate(series):
                    totals[position] += _z(value)
        if found:
            return totals
 
    scen_rr = _z(cell.get("scen_rr"))
    feb_mar_rr = _z(cell.get("feb_mar_rr")) if cell.get("feb_mar_rr") is not None else scen_rr
    return [scen_rr] * MONTHS_JUL_JAN + [feb_mar_rr] * MONTHS_FEB_MAR
 
 
def trajectory_cell(model: ScenarioModel, basis: str, asset: str) -> Dict[str, Any]:
    """A scenario cell enriched with a trajectory and the right leakage label."""
    cell = dict(model.cell(basis, asset=None if asset == "All" else asset))
    cell["trajectory"] = scenario_trajectory(model, basis, asset)
    if cell.get("leakage") is None:
        if model.scenario_id == 3:
            cell["leakage"] = float(model.params.get("dip", S3_DEFAULT_DIP))
        elif model.scenario_id in (7, 8, 9):
            cell["leakage"] = float(model.params.get("leakage", S7_DEFAULT_LEAKAGE))
        else:
            cell["leakage"] = 0.0
    return cell
 
 
def render_scenario_trajectory(model: ScenarioModel, basis: str, asset: str) -> None:
    """07 · Monthly progression shown as data only; all graphs are removed."""
    section_header(
        "07",
        "Scenario trajectory",
        "Month by month from July 2026 to March 2027",
    )

    cell = trajectory_cell(model, basis, asset)
    frame, formats = build_momentum_analysis(cell)
    render_glass_table(frame, formats, css_class="scenario-table")

    if model.scenario_id == 7:
        render_momentum_detail(model, basis)

def render_momentum_detail(model: ScenarioModel, basis: str) -> None:
    """Scenario 7 specifics: binding milestone, buffer, leakage and sensitivity."""
    cell = model.cell(basis)
    momentum = cell.get("momentum_g")
 
    kpi_strip([
        {"label": "Required MoM momentum",
         "value": fmt_pct(momentum) if momentum is not None else NA_TEXT,
         "tone": "gold",
         "secondary": f"binding milestone: {cell.get('binding')}"},
        {"label": "January achievement", "value": fmt_pct(cell.get("jan_pct")),
         "delta": fmt_cr_signed(cell.get("jan_buffer")),
         "tone": _tone_for(cell.get("jan_buffer")),
         "secondary": "buffer created before leakage"},
        {"label": "Feb–Mar leakage", "value": fmt_pct(cell.get("leakage")),
         "secondary": f"Feb {fmt_cr(cell.get('feb_mar_rr'))} · Mar {fmt_cr(cell.get('march_rr'))}"},
        {"label": "March achievement", "value": fmt_pct(cell.get("march_pct")),
         "delta": fmt_pts(cell.get("headroom_pct")),
         "tone": _tone_for(cell.get("headroom_pct"))},
        {"label": "January exit run rate", "value": fmt_cr(cell.get("scen_rr")),
         "delta": fmt_pct_signed(cell.get("rr_change_pct")),
         "tone": _tone_for(cell.get("rr_change_pct"))},
    ], css_class="scenario-kpi-grid")
 
    if cell.get("feasible"):
        glass_callout(
            "<span class='tag-ok'>Target achievable</span> — the momentum trajectory reaches the "
            f"January milestone and still clears the March ambition after "
            f"{fmt_pct(cell.get('leakage'))} February–March leakage.",
            tone="ok",
        )
    else:
        glass_callout(
            "<span class='tag-warn'>Additional momentum required</span> — March needs a further "
            f"{fmt_cr(cell.get('additional_march_sales'))}, which means lifting the January run "
            f"rate by {fmt_cr(cell.get('additional_jan_rr'))} per month.",
            tone="warn",
        )
    if cell.get("note"):
        glass_note(escape(str(cell["note"])))
 
    momentum_text = fmt_pct(momentum) if momentum is not None else "a flat required"
    outcome = (
        f"{fmt_pct(cell.get('headroom_pct'))} headroom" if _z(cell.get("headroom_amt")) >= 0
        else f"a {fmt_cr(abs(_z(cell.get('headroom_amt'))))} shortfall"
    )
    glass_callout(
        f"<b>What this asks of the organisation:</b> build roughly {momentum_text} month-on-month "
        f"momentum from July through January, lifting the monthly run rate from "
        f"{fmt_cr(cell.get('current_rr'))} to {fmt_cr(cell.get('scen_rr'))} by January to reach the "
        f"{fmt_pct(model.params.get('jan_target', S7_DEFAULT_JAN_TARGET))} milestone. After "
        f"{fmt_pct(cell.get('leakage'))} February–March leakage the trajectory lands at "
        f"{fmt_pct(cell.get('march_pct'))} of the FY target, leaving {outcome}."
    )
 
    tabs = st.tabs(["Asset class", "Retail / DHNI / VRM", "Leakage sensitivity", "Monthly revenue"])
    with tabs[0]:
        for sales in SALES_TYPES:
            st.markdown(
                f"<div class='metric-label'>{SALES_LABEL[sales]}</div>", unsafe_allow_html=True
            )
            frame, formats = build_momentum_by_group(model, sales, "asset")
            render_glass_table(frame, formats, css_class="scenario-table")
    with tabs[1]:
        for sales in SALES_TYPES:
            st.markdown(
                f"<div class='metric-label'>{SALES_LABEL[sales]}</div>", unsafe_allow_html=True
            )
            frame, formats = build_momentum_by_group(model, sales, "vertical")
            render_glass_table(frame, formats, css_class="scenario-table")
    with tabs[2]:
        frame, formats = build_leakage_sensitivity(model, basis)
        render_glass_table(frame, formats, css_class="scenario-table")
        glass_note(
            "Momentum is re-solved at each leakage assumption, so the required July–January "
            "build changes with the February–March pressure."
        )
    with tabs[3]:
        frame, formats = build_monthly_revenue(model, REVENUE_BASIS)
        render_glass_table(frame, formats, css_class="scenario-table")
        january_revenue = calculate_revenue(model.assets(REVENUE_BASIS), "jan_amount")
        march_revenue = calculate_revenue(model.assets(REVENUE_BASIS), "march_amount")
        baseline = calculate_baseline_revenue(model.assets(REVENUE_BASIS))
        glass_note(
            f"January scenario revenue {fmt_cr(january_revenue['total'], 1)} · "
            f"March scenario revenue {fmt_cr(march_revenue['total'], 1)} · baseline "
            f"{fmt_cr(baseline['total'], 1)} · incremental "
            f"{fmt_cr_signed(march_revenue['total'] - baseline['total'], 1)}."
        )
 
 
def render_revenue_impact(model: ScenarioModel) -> Dict[str, Any]:
    """08 · Current-versus-expected revenue only, with asset-class detail."""
    section_header(
        "08",
        "Current revenue vs expected revenue",
        "Current trajectory compared with the selected scenario · Net Sales revenue basis",
    )

    bundle = revenue_bundle(model, REVENUE_BASIS)
    current_revenue = bundle["baseline"]["total"]
    expected_revenue = bundle["scenario"]["total"]

    st.markdown(
        "<div class='revenue-compare-grid'>"
        "<div class='glass-card revenue-current-card'>"
        "<div class='metric-label'>Current revenue</div>"
        f"<div class='revenue-compare-amount'>{escape(fmt_cr(current_revenue, 1))}</div>"
        "<div class='metric-secondary'>Expected by March at the current trajectory</div>"
        "</div>"
        "<div class='glass-card revenue-expected-card'>"
        "<div class='metric-label'>Expected revenue</div>"
        f"<div class='revenue-compare-amount'>{escape(fmt_cr(expected_revenue, 1))}</div>"
        f"<div class='metric-secondary'>Scenario {model.scenario_id} expected by March</div>"
        "</div>"
        "</div>",
        unsafe_allow_html=True,
    )

    frame, formats = build_revenue_impact(model, REVENUE_BASIS)
    render_glass_table(
        frame,
        formats,
        total_rows=("Total",),
        css_class="scenario-table revenue-table",
    )

    glass_note(
        "Revenue is calculated separately by asset class on Net Sales using the configured "
        "Equity, Debt and Liquid revenue rates. The table shows only current and expected revenue."
    )
    return bundle

def render_all_scenarios(
    grid: pd.DataFrame,
    scenario_id: int,
    params: Dict[str, Any],
    basis: str,
    asset: str,
) -> None:
    """Every scenario evaluated on the selected scope, table only."""
    with st.expander("Compare all ten scenarios on this scope", expanded=False):
        frame, formats = build_all_scenario_matrix(grid, scenario_id, params, basis, asset)
        if frame.empty:
            glass_note("No scenario output is available for this scope.")
            return
        render_glass_table(
            frame, formats,
            total_rows=(f"{scenario_id:02d} · {SCENARIOS[scenario_id]['short']}",),
            css_class="scenario-table",
        )
        glass_note(
            "Scenarios 1–10 are calculated on the same scope and sales basis. The selected "
            "scenario uses the live assumptions above; the others use their configured defaults."
        )

# =============================================================================
# 19. SCENARIO-SPECIFIC DRIVERS
# =============================================================================
 
def render_segment_section(model: ScenarioModel, basis: str, counts: Dict[str, int]) -> None:
    """Scenario 6 · differentiated performance by business segment."""
    unmapped = [s for s in ("Digital", "Retail B30") if counts.get(s, 0) == 0]
    if unmapped:
        missing = " and ".join(unmapped)
        glass_callout(
            f"<b>Segment validation:</b> no records could be classified as {missing} from the "
            "workbook metadata (MKT TYPE, Type, ZONE, REGION, EM City). Those records remain in "
            f"Others, so the {missing} uplift is not applied. Point the classification at the "
            "right column under <i>Segment mapping</i> in the sidebar.",
            tone="warn",
        )
 
    segment_targets = dict(model.params.get("segment_targets", S6_SEGMENT_TARGETS))
    present = " · ".join(
        f"{segment} {float(segment_targets.get(segment, S6_SEGMENT_TARGETS[segment])):.1%} of FY target ({counts.get(segment, 0)} RMs)"
        for segment in SEGMENT_ORDER
    )
    glass_note(f"Scenario assumption — {present}.")
 
    tabs = st.tabs([SALES_LABEL["NS"], SALES_LABEL["GS"]])
    for tab, sales in zip(tabs, ["NS", "GS"]):
        with tab:
            frame, formats = build_segment_scenario_analysis(model, sales)
            render_glass_table(frame, formats, total_rows=("Overall",), css_class="scenario-table")
 
    overall = model.cell(basis)
    lines = []
    for segment in model.available_segments():
        cell = model.cell(basis, segment=segment)
        uplift = (
            fmt_pct_signed(cell["rr_change_pct"]) if cell["rr_change_pct"] is not None else NA_TEXT
        )
        lines.append(
            f"<b>{segment}</b> moves from {fmt_pct(cell['current_march_pct'])} to "
            f"{fmt_pct(cell['march_pct'])} of FY target, needing {fmt_cr(cell['scen_rr'])} per month "
            f"({uplift} run-rate uplift)"
        )
    glass_callout(
        f"On {SALES_LABEL[basis]}: " + "; ".join(lines) + ". Overall March achievement moves from "
        f"{fmt_pct(overall['current_march_pct'])} to {fmt_pct(overall['march_pct'])}, an improvement "
        f"of {fmt_cr_signed(overall['incremental_sales'])}."
    )
 
 
def render_channel_simulator(model: ScenarioModel, basis: str) -> None:
    """Scenario 8 · executive channel simulator, table only."""
    frame, formats = build_channel_scenario_analysis(model, basis)
    if frame.empty:
        glass_callout(
            "No mapped channel data is available. Use <i>Channel mapping</i> in the sidebar to "
            "classify the workbook into the nine planning channels.",
            tone="warn",
        )
        return

    jan_gap = _z(frame["Jan Gap / Headroom"].sum())
    mar_gap = _z(frame["Mar Gap / Headroom"].sum())
    incremental = _z(frame["March Incremental Sales"].sum())

    kpi_strip([
        {"label": "January portfolio headroom", "value": fmt_cr_signed(jan_gap),
         "tone": _tone_for(jan_gap), "secondary": "above or below the January target"},
        {"label": "March portfolio headroom", "value": fmt_cr_signed(mar_gap),
         "tone": _tone_for(mar_gap), "secondary": "above or below the March target"},
        {"label": "Incremental March sales", "value": fmt_cr_signed(incremental),
         "tone": _tone_for(incremental), "secondary": "versus the current projection"},
        {"label": "Channels in play", "value": f"{len(frame)} of {len(CHANNELS)}",
         "secondary": "mapped from workbook metadata"},
    ], css_class="scenario-kpi-grid")

    render_glass_table(frame, formats, css_class="scenario-table")

    on_track = (frame["Jan Gap / Headroom"] >= 0).all() and (frame["Mar Gap / Headroom"] >= 0).all()
    if on_track:
        glass_callout(
            "<span class='tag-ok'>All channels on track</span> — the selected growth assumptions "
            "clear both the January and March targets after leakage.",
            tone="ok",
        )
    else:
        misses = frame.loc[
            (frame["Jan Gap / Headroom"] < 0) | (frame["Mar Gap / Headroom"] < 0), "Channel"
        ].tolist()
        glass_callout(
            "<span class='tag-warn'>Channel gap</span> — review " + ", ".join(misses)
            + ". Raise monthly growth or reset the target for those channels.",
            tone="warn",
        )

def render_channel_optimizer(model: ScenarioModel, basis: str) -> None:
    """Scenario 9 · minimum growth required, table only."""
    ambition = float(model.params.get("optimizer_target", 1.20))
    frame, formats = build_channel_scenario_analysis(model, basis)
    if frame.empty:
        glass_callout(
            "No mapped channel data is available. Map the workbook into channels in the sidebar "
            "to run the optimiser.",
            tone="warn",
        )
        return

    solved = [(_num(v) or 0.0) for v in frame["MoM Growth"]]
    hardest = frame.iloc[int(np.argmax(solved))] if solved else None
    portfolio = model.cell(basis)

    kpi_strip([
        {"label": "March ambition", "value": fmt_pct(ambition), "tone": "gold",
         "secondary": "portfolio target set by management"},
        {"label": "Minimum required growth", "value": fmt_pct(max(solved) if solved else None),
         "secondary": f"hardest channel: {hardest['Channel']}" if hardest is not None else ""},
        {"label": "Average required growth",
         "value": fmt_pct(float(np.mean(solved)) if solved else None),
         "secondary": "across mapped channels"},
        {"label": "January milestone", "value": fmt_pct(portfolio.get("jan_pct")),
         "secondary": "editable portfolio January target"},
        {"label": "Feb–Mar leakage",
         "value": fmt_pct(model.params.get("leakage", S8_DEFAULT_LEAKAGE)),
         "secondary": "applied after January"},
    ], css_class="scenario-kpi-grid")

    display = frame.copy()
    display["Portfolio ambition"] = ambition
    display_formats = dict(formats)
    display_formats["Portfolio ambition"] = "pct"
    render_glass_table(display, display_formats, css_class="scenario-table")

    glass_callout(
        "The optimiser solves the minimum compounding trajectory each channel must run from July "
        f"to hold the January milestone and still land the {fmt_pct(ambition)} March ambition after "
        f"{fmt_pct(model.params.get('leakage', S8_DEFAULT_LEAKAGE))} leakage. Channels already at "
        "or above the requirement solve to zero additional growth."
    )


def render_asset_channel_target_simulator(model: ScenarioModel) -> None:
    """Scenario 10 · linked Asset × Channel × Retail-market view, excluding Digital."""
    edit_basis = str(model.params.get("scenario10_edit_basis", "NS"))
    linked_basis = "GS" if edit_basis == "NS" else "NS"

    glass_callout(
        "<b>Scenario 10 roll-up · EX-DIGITAL:</b> Retail, DHNI, VRM and Insti are the calculation "
        "building blocks. Every displayed achievement percentage is Projected Number ÷ FY27 Target. "
        "Current is read directly from FINAL YTD; it is not annualised. The linked Net/Gross Sales "
        "basis follows the same relative change where the current ratio is positive."
    )

    sales_tabs = st.tabs([SALES_LABEL[edit_basis], f"Linked {SALES_LABEL[linked_basis]}"])
    for tab, sales in zip(sales_tabs, (edit_basis, linked_basis)):
        with tab:
            asset_frame, asset_formats = build_scenario_10_asset_summary(model, sales)
            if not asset_frame.empty:
                cards: List[Dict[str, Any]] = []
                for _, row in asset_frame.iterrows():
                    delta = _num(row.get("Change vs Current"))
                    cards.append({
                        "label": str(row.get("Asset Class")),
                        "value": fmt_pct(row.get("Simulation Achievement %")),
                        "delta": fmt_pts(delta),
                        "tone": _tone_for(delta),
                        "secondary": (
                            f"FINAL current {fmt_pct(row.get('Current %'))} · "
                            f"projected {fmt_cr(row.get('Simulation Projected Number'))}"
                        ),
                    })
                kpi_strip(cards)

            st.markdown("<div class='metric-label' style='margin-top:16px'>Equity / Debt / Liquid roll-up</div>", unsafe_allow_html=True)
            render_glass_table(asset_frame, asset_formats, css_class="scenario-table")

            st.markdown("<div class='metric-label' style='margin-top:18px'>Retail / DHNI / VRM / Insti inside each asset · Digital excluded</div>", unsafe_allow_html=True)
            detail_frame, detail_formats = build_scenario_10_channel_detail(model, sales)
            render_glass_table(detail_frame, detail_formats, css_class="scenario-table")

            location_frame, location_formats = build_scenario_10_retail_location_detail(model, sales)
            if not location_frame.empty:
                st.markdown("<div class='metric-label' style='margin-top:18px'>Market type · Retail + VRM + DHNI</div>", unsafe_allow_html=True)
                render_glass_table(location_frame, location_formats, css_class="scenario-table")


def render_scenario_drivers(
    model: ScenarioModel,
    basis: str,
    segment_counts: Dict[str, int],
) -> None:
    """09 · Scenario data segregated only into Channel and Asset Class views."""
    section_header(
        "09",
        "Scenario drivers",
        "Where the required delivery actually sits",
    )

    if model.scenario_id == 6:
        render_segment_section(model, basis, segment_counts)
    elif model.scenario_id == 8:
        render_channel_simulator(model, basis)
    elif model.scenario_id == 9:
        render_channel_optimizer(model, basis)
    elif model.scenario_id == 10:
        render_asset_channel_target_simulator(model)
        return

    tabs = st.tabs(["Channel", "Asset class"])
    with tabs[0]:
        frame, formats = build_vertical_summary(model)
        if frame.empty:
            glass_note("No channel data is available for this scope.")
        else:
            render_glass_table(frame, formats, css_class="scenario-table")

    with tabs[1]:
        rows: List[Dict[str, Any]] = []
        for sales in ("NS", "GS"):
            for asset_name in ASSETS:
                cell = model.cell(sales, asset=asset_name)
                rows.append({
                    "Sales": SALES_LABEL[sales],
                    "Asset Class": asset_name,
                    "FY Target": cell.get("fy_target"),
                    "YTD Achievement": cell.get("ytd_ach"),
                    "Target Achieved %": cell.get("ytd_ach_pct"),
                    "Current Run Rate": cell.get("current_rr"),
                    "Scenario Run Rate": cell.get("scen_rr"),
                    "Run Rate Change %": cell.get("rr_change_pct"),
                    "Current March Projection %": cell.get("current_march_pct"),
                    "Scenario March Projection %": cell.get("march_pct"),
                    "Incremental Sales": cell.get("incremental_sales"),
                })
        asset_frame = pd.DataFrame(rows)
        render_glass_table(
            asset_frame,
            {
                "Sales": "txt", "Asset Class": "txt", "FY Target": "cr",
                "YTD Achievement": "cr", "Target Achieved %": "pct",
                "Current Run Rate": "cr", "Scenario Run Rate": "cr",
                "Run Rate Change %": "pct_signed", "Current March Projection %": "pct",
                "Scenario March Projection %": "pct", "Incremental Sales": "cr_signed",
            },
            css_class="scenario-table",
        )

def render_detail_tables(model: ScenarioModel) -> None:
    """10 · Detailed analytical tables, supporting the cards above."""
    section_header("10", "Detailed analytical tables", "The full numbers behind every card")
 
    tabs = st.tabs(["Current baseline", "Current vs scenario", "Scenario guide"])
    with tabs[0]:
        frame, formats = build_current_overview(model)
        render_glass_table(frame, formats, css_class="scenario-table")
    with tabs[1]:
        frame, formats = build_comparison(model)
        render_glass_table(frame, formats, css_class="scenario-table")
    with tabs[2]:
        guide = build_scenario_guide(model, REVENUE_BASIS)
        render_glass_table(guide, {c: "txt" for c in guide.columns}, css_class="scenario-table")
 
 
def render_final_reference(payload: bytes) -> None:
    """11 · The workbook's own FINAL sheet, kept as the audit surface."""
    section_header("11", "FINAL workbook reference", "The uploaded source of truth, unaltered")
 
    with st.expander("Open the FINAL sheet", expanded=False):
        try:
            st.markdown(build_final_sheet_html(payload), unsafe_allow_html=True)
        except WorkbookError as error:
            st.error(str(error))
        except Exception:  # pragma: no cover - defensive
            st.error("The FINAL sheet could not be rendered from this workbook.")
 
    with st.expander("FINAL sheet as raw data", expanded=False):
        try:
            raw = load_final_sheet_frame(payload)
            st.dataframe(raw, hide_index=True, height=560, **_dataframe_kwargs())
        except Exception:  # pragma: no cover - defensive
            st.error("The FINAL sheet could not be read as raw data.")
 
 
def render_export(model: ScenarioModel, payload: bytes) -> None:
    """12 · Export the selected scenario and the source workbook."""
    section_header("12", "Export", "Take the analysis into the management pack")
 
    left, right = st.columns(2)
    with left:
        try:
            export_payload = make_export_excel(model, REVENUE_BASIS)
            st.download_button(
                f"Download scenario {model.scenario_id} analysis",
                data=export_payload,
                file_name=f"scenario_{model.scenario_id}_analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception:  # pragma: no cover - defensive
            st.warning("The scenario export could not be generated for this scope.")
    with right:
        st.download_button(
            "Download the uploaded workbook",
            data=payload,
            file_name="sales_command_center_source.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
 
    glass_note(
        "The scenario workbook contains the scenario guide, current baseline, current-vs-scenario "
        "comparison, revenue impact, Retail/DHNI/VRM summary, Gross and Net breakdowns, the "
        "Scenario 6 segment view, the Scenario 7 momentum trajectory and monthly revenue, the "
        "channel matrix where relevant, Scenario 10 Asset × Channel detail, and all ten scenarios side by side."
    )
 
 
# =============================================================================
# 20. SIDEBAR - DATA MAPPING & NAVIGATION
# =============================================================================
 
def render_segment_controls(records: pd.DataFrame) -> Dict[str, Any]:
    """Configurable segment classification for Scenario 6."""
    suggestions = identify_segments(records)
    mapping: Dict[str, Any] = dict(st.session_state.get("segment_mapping") or suggestions)
 
    usable_columns = [
        field for field in META_FIELDS
        if field in records.columns and text_column(records, field).ne("").any()
    ]
 
    with st.sidebar.expander("Segment mapping", expanded=False):
        st.caption(
            "Digital, Retail B30 and Others are derived from workbook metadata. "
            "Everything unmatched falls into Others."
        )
        for segment in ("Digital", "Retail B30"):
            options = ["Not mapped"] + usable_columns
            current = mapping.get(segment, {}).get("column", "Not mapped")
            index = options.index(current) if current in options else 0
            column = st.selectbox(
                f"{segment} identified by", options, index=index, key=f"seg_col_{segment}"
            )
            if column == "Not mapped":
                mapping.pop(segment, None)
                continue
            values = sorted({v for v in text_column(records, column) if v.strip()})
            preset = [v for v in mapping.get(segment, {}).get("values", []) if v in values]
            chosen = st.multiselect(
                f"{segment} values", values, default=preset, key=f"seg_vals_{segment}_{column}"
            )
            if chosen:
                mapping[segment] = {"column": column, "values": list(chosen)}
            else:
                mapping.pop(segment, None)
 
    st.session_state["segment_mapping"] = mapping
    return mapping
 
 
def render_channel_controls(records: pd.DataFrame) -> Dict[str, Any]:
    """Channel mapping for the Scenario 8 / 9 planning universe."""
    suggestions = identify_channels(records)
    mapping: Dict[str, Any] = dict(st.session_state.get("channel_mapping") or suggestions)
    usable_columns = [
        f for f in META_FIELDS + ["Vertical"]
        if f in records.columns and text_column(records, f).ne("").any()
    ]
 
    with st.sidebar.expander("Channel mapping", expanded=False):
        st.caption(
            "Map workbook metadata to Digital, VRM, EM, B30, T30, T8, DHNI, Retail and "
            "Institutional. Unmapped rows stay unclassified."
        )
        for channel in CHANNELS:
            options = ["Not mapped"] + usable_columns
            current = mapping.get(channel, {}).get("column", "Not mapped")
            index = options.index(current) if current in options else 0
            column = st.selectbox(channel, options, index=index, key=f"ch_col_{channel}")
            if column == "Not mapped":
                mapping.pop(channel, None)
                continue
            values = sorted({v for v in text_column(records, column) if v.strip()})
            preset = [v for v in mapping.get(channel, {}).get("values", []) if v in values]
            chosen = st.multiselect(
                f"{channel} values", values, default=preset, key=f"ch_vals_{channel}_{column}"
            )
            if chosen:
                mapping[channel] = {"column": column, "values": list(chosen)}
            else:
                mapping.pop(channel, None)
 
    st.session_state["channel_mapping"] = mapping
    return mapping
 
 
def _sidebar_pct(key: str, default_fraction: float) -> str:
    """Format an editable percentage from session state for the scenario dropdown label."""
    raw = st.session_state.get(key, default_fraction * 100.0)
    try:
        return f"{float(raw):.1f}%"
    except (TypeError, ValueError):
        return f"{default_fraction * 100.0:.1f}%"


def _sidebar_scenario_label(sid: int) -> str:
    """Scenario dropdown copy that follows the latest edited percentage assumptions."""
    if sid == 1:
        return f"Scenario 1 · +{_sidebar_pct('s1_uplift', S1_RUNRATE_UPLIFT)} Run-Rate Push"
    if sid == 2:
        return f"Scenario 2 · {_sidebar_pct('s2_overall', S2_OVERALL_TARGET)} Overall by Jan + {_sidebar_pct('s2_equity', S2_EQUITY_TARGET)} Equity"
    if sid == 3:
        return f"Scenario 3 · {_sidebar_pct('s3_target', S3_TARGET)} by Jan · {_sidebar_pct('s3_dip', S3_DEFAULT_DIP)} Dip"
    if sid == 4:
        return f"Scenario 4 · {_sidebar_pct('s4_target', S4_TARGET)} by March"
    if sid == 5:
        return f"Scenario 5 · {_sidebar_pct('s5_equity', S5_EQUITY_TARGET)} Equity + {_sidebar_pct('s5_overall', S5_OVERALL_TARGET)} Overall"
    if sid == 6:
        return f"Scenario 6 · Digital {_sidebar_pct('s6_digital', S6_SEGMENT_TARGETS['Digital'])} + B30 {_sidebar_pct('s6_retail_b30', S6_SEGMENT_TARGETS['Retail B30'])} + Others {_sidebar_pct('s6_others', S6_SEGMENT_TARGETS['Others'])}"
    if sid == 7:
        return f"Scenario 7 · Jan {_sidebar_pct('s7_jan', S7_DEFAULT_JAN_TARGET)} → Mar {_sidebar_pct('s7_mar', S7_DEFAULT_MAR_TARGET)} · Leakage {_sidebar_pct('s7_leak', S7_DEFAULT_LEAKAGE)}"
    if sid == 8:
        return "Scenario 8 · Channel Growth & Target Simulator"
    if sid == 9:
        return f"Scenario 9 · Jan {_sidebar_pct('s9_jan_target', 1.0)} → Mar {_sidebar_pct('s9_target', 1.20)} · Leakage {_sidebar_pct('s9_leakage', S8_DEFAULT_LEAKAGE)}"
    if sid == 10:
        return "Scenario 10 · Asset × Channel Target Simulator"
    return SCENARIOS[sid]["label"]


def render_sidebar(records: pd.DataFrame) -> Tuple[str, Dict[str, Any], Dict[str, Any]]:
    """Quiet utility rail: page, data mapping, assumptions and workbook actions."""
    sidebar = st.sidebar
    sidebar.markdown("<div class='sidebar-mark'>Command Center</div>", unsafe_allow_html=True)
    sidebar.markdown("<div class='sidebar-title'>View</div>", unsafe_allow_html=True)
    page = sidebar.radio(
        "View",
        ["Executive command center", "RM performance", "Bon voyage", "AI Choice"],
        index=0,
        key="application_page_selector",
        label_visibility="collapsed",
    )

    # Scenario selection has a single home: the left sidebar. This prevents
    # duplicate scenario selectors from appearing above or inside comparisons.
    sidebar.markdown("<div class='sidebar-title'>Scenario planning</div>", unsafe_allow_html=True)
    current_scenario = int(st.session_state.get("scenario_id", 1))
    if current_scenario not in SCENARIO_ORDER:
        current_scenario = 1
    selected_scenario_id = sidebar.selectbox(
        "Scenario",
        SCENARIO_ORDER,
        index=SCENARIO_ORDER.index(current_scenario),
        format_func=_sidebar_scenario_label,
        key="sidebar_scenario_selector_v7",
        label_visibility="collapsed",
    )
    st.session_state["scenario_id"] = int(selected_scenario_id)
 
    sidebar.markdown("<div class='sidebar-title'>Data mapping</div>", unsafe_allow_html=True)
    segment_mapping = render_segment_controls(records)
    channel_mapping = render_channel_controls(records)
 
    sidebar.markdown("<div class='sidebar-title'>Assumptions</div>", unsafe_allow_html=True)
    with sidebar.expander("Model assumptions", expanded=False):
        st.caption(
            "Revenue basis: Net Sales only. Rates: Equity 60 bps · Debt 20 bps · Liquid 10 bps. "
            "Timeline: April–June complete, nine months remaining, July–January is seven months "
            "and February–March is two months. The current run rate is YTD ÷ 3."
        )
 
    sidebar.markdown("<div class='sidebar-title'>Workbook</div>", unsafe_allow_html=True)
    if sidebar.button("Use another workbook", key="reset_workbook_button"):
        reset_workbook()
 
    return page, segment_mapping, channel_mapping
 
 
# =============================================================================
# 21. UPLOAD EXPERIENCE
# =============================================================================
 
def render_upload_screen() -> None:
    st.markdown(
        "<div class='exec-header'><div>"
        "<div class='exec-mark'>FY27 · Executive management view</div>"
        f"<div class='exec-title'>{escape(APP_TITLE)}</div>"
        f"<div class='exec-sub'>{escape(APP_SUBTITLE)}</div>"
        "</div></div>",
        unsafe_allow_html=True,
    )
 
    left, right = st.columns([1.35, 1])
    with left:
        st.markdown(
            "<div class='glass-card'>"
            "<div class='metric-label'>Start here</div>"
            "<div class='metric-hero'>Upload the RM scorecard</div>"
            "<div class='metric-secondary'>The workbook stays in this session only. "
            "Nothing is stored once the tab is closed.</div></div>",
            unsafe_allow_html=True,
        )
        uploaded = st.file_uploader(
            "Workbook", type=["xlsx", "xlsm"], label_visibility="collapsed"
        )
    with right:
        st.markdown(
            "<div class='glass-panel'>"
            "<div class='metric-label'>The workbook needs</div>"
            "<div class='metric-secondary'>"
            "<span class='inline-pill gold'>FINAL</span>"
            "<span class='inline-pill'>RM Retail Sales</span>"
            "<span class='inline-pill'>RM DHNI</span>"
            "<span class='inline-pill'>VRM</span><br><br>"
            "FINAL supplies Gross Sales and Net Sales targets. The three RM sheets supply "
            "employee-level targets and achievement used for the Asset Class and Channel views."
            "</div></div>",
            unsafe_allow_html=True,
        )
 
    if uploaded is not None:
        payload = uploaded.getvalue()
        try:
            load_workbook(payload)
            parse_final_dashboard_metrics(payload)
        except WorkbookError as error:
            st.error(str(error))
        except Exception:  # pragma: no cover - never show a traceback
            st.error(
                "That file could not be read as the RM scorecard workbook. Check that it is a "
                "valid Excel file containing RM Retail Sales, RM DHNI, VRM and FINAL."
            )
        else:
            st.session_state["workbook"] = payload
            rerun()
 
 
# =============================================================================
# 22. EXECUTIVE COMMAND CENTER - SCREEN FLOW
# =============================================================================
 
def render_command_center(records: pd.DataFrame, payload: bytes) -> None:
    """The full management journey, in the order an executive reads it."""
    final_metrics = parse_final_dashboard_metrics(payload)
 
    render_apple_header(final_metrics, records, "FY27 · Executive management view")
 
    full_grid = build_base_grid(records)
    base_params = scenario_default_params(1)
    base_model = ScenarioModel(1, full_grid, base_params)
 
    # 02 · Current reality, straight from FINAL - before any scenario exists.
    render_current_performance(final_metrics, base_model)
 
    # 03 · Drivers behind that reality.
    render_business_driver_selector(final_metrics, base_model)
    basis = st.session_state.get("display_basis", "NS")
 
    # 04 · Scope + the pace question.
    section_header(
        "04",
        "Current run rate & target gap",
        "The pace the business is running at, against the pace the target needs",
    )
    channel, location, asset, scoped_records = render_analysis_scope(records)
    scoped_grid = build_base_grid(scoped_records) if not scoped_records.empty else pd.DataFrame()
 
    if scoped_grid.empty:
        render_final_reference(payload)
        return
 
    scoped_cell = summarize_current(
        scoped_grid, sales=basis, asset=None if asset == "All" else asset
    )
    required_rr = (
        _z(scoped_cell.get("fy_target")) - _z(scoped_cell.get("ytd_ach"))
    ) / max(MONTHS_REMAINING, 1)
    current_rr = _num(scoped_cell.get("current_rr"))
    pace_gap = (current_rr / required_rr - 1.0) if (current_rr is not None and required_rr) else None
    projected = _num(scoped_cell.get("current_march_pct"))
    shortfall = _z(scoped_cell.get("fy_target")) - _z(scoped_cell.get("current_march"))
 
    kpi_strip([
        {"label": "FY target", "value": fmt_cr(scoped_cell.get("fy_target")),
         "secondary": f"{SALES_LABEL[basis]} · RM calculation sheets"},
        {"label": "YTD achievement", "value": fmt_cr(scoped_cell.get("ytd_ach")),
         "delta": fmt_pct(scoped_cell.get("ytd_ach_pct")), "tone": "gold",
         "secondary": "against the YTD June target"},
        {"label": "Current run rate", "value": fmt_cr(current_rr),
         "secondary": "YTD ÷ 3 completed months"},
        {"label": "Required run rate", "value": fmt_cr(required_rr),
         "delta": fmt_pct_signed(pace_gap), "tone": _tone_for(pace_gap),
         "secondary": "remaining FY target ÷ remaining months"},
        {"label": "Projected March", "value": fmt_pct(projected),
         "delta": fmt_pts(None if projected is None else projected - 1.0),
         "tone": _tone_for(None if projected is None else projected - 1.0),
         "secondary": fmt_cr(scoped_cell.get("current_march"))},
        {"label": "Gap at current pace", "value": fmt_cr(shortfall),
         "tone": _tone_for(-shortfall),
         "secondary": "FY target less projected March"},
    ])
 
    glass_note(
        "The current run rate is completed Apr–Jun achievement divided by three. The required "
        "run rate is the remaining FY target (Target − YTD) divided by the 9 remaining months."
    )
 
    # 05 · Scenario planning.
    scenario_id = render_scenario_navigator()

    # Scenario 10 is deliberately rebuilt from FINAL only. It does not use the
    # RM-sheet scope grid for Current, FY27 Budget or projected-number math.
    scenario_grid = (
        build_scenario10_final_grid(final_metrics)
        if scenario_id == 10
        else scoped_grid
    )

    params = render_scenario_controls(
        scenario_id,
        scenario_default_params(scenario_id),
        scenario_grid,
        final_metrics,
    )
    params["channel_mapping"] = st.session_state.get("channel_mapping", {})
 
    try:
        model = ScenarioModel(scenario_id, scenario_grid, params)
    except Exception:  # pragma: no cover - defensive
        st.error("This scenario could not be calculated on the current scope. Widen the scope or "
                 "select another scenario.")
        render_final_reference(payload)
        return
 
    # 06 · Selected scenario.
    render_scenario_hero(model, basis, asset)
    render_scenario_comparison(final_metrics, model, basis, asset)
    if scenario_id == 10:
        glass_note(
            "Scenario 10 ignores RM-sheet Channel/Location scope filters by design. Its Current, "
            "FY27 Budget and simulation values are read from FINAL only."
        )
    elif channel != "All" or location != "All":
        glass_note(
            "The scope filter applies to the RM calculation sheets. FINAL targets in the table "
            "above remain the published portfolio targets, so scenario percentages are the "
            "comparable figures while a filter is active."
        )

    if scenario_id != 10:
        render_all_scenarios(scoped_grid, scenario_id, params, basis, asset)
 
    # 07 · Trajectory.
    render_scenario_trajectory(model, basis, asset)
 
    # 08 · Revenue.
    render_revenue_impact(model)
 
    # 09 · Drivers behind the scenario.
    render_scenario_drivers(model, basis, segment_diagnostics(scoped_records))
 
    # 10-12 · Tables, source workbook, export.
    render_detail_tables(model)
    render_final_reference(payload)
    render_export(model, payload)
 
 
# =============================================================================
# 23. RM PERFORMANCE SEGMENTATION PAGE
# =============================================================================
 
def _clean_filter_values(series: pd.Series) -> List[str]:
    """Return sorted non-empty values for RM filter dropdowns."""
    cleaned = series.astype(str).str.replace("\u00a0", " ", regex=False).str.strip()
    cleaned = cleaned[
        cleaned.ne("")
        & cleaned.str.lower().ne("nan")
        & cleaned.str.lower().ne("none")
    ]
    return sorted(cleaned.drop_duplicates().tolist(), key=lambda value: value.lower())
 
 
def _apply_exact_text_filter(frame: pd.DataFrame, column: str, selected: str) -> pd.DataFrame:
    """Apply one exact dropdown filter while treating 'All' as no filter."""
    if selected == "All" or column not in frame.columns:
        return frame
    values = frame[column].astype(str).str.replace("\u00a0", " ", regex=False).str.strip()
    return frame.loc[values == selected].copy()
 
 
def render_retail_rm_filters(records: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Cascading Retail filters: ZONE -> REGION -> MKT TYPE."""
    retail = records.loc[records["Vertical"] == "Retail"].copy()
    selections: Dict[str, str] = {"ZONE": "All", "REGION": "All", "MKT TYPE": "All"}
 
    st.markdown("<div class='metric-label'>Retail filters</div>", unsafe_allow_html=True)
    columns = st.columns(3)
 
    zone_options = ["All"] + (
        _clean_filter_values(retail["ZONE"]) if "ZONE" in retail.columns else []
    )
    with columns[0]:
        reset_stale_selection("retail_rm_zone_filter", zone_options)
        selections["ZONE"] = st.selectbox("Zone", zone_options, index=0, key="retail_rm_zone_filter")
    filtered_retail = _apply_exact_text_filter(retail, "ZONE", selections["ZONE"])
 
    region_options = ["All"] + (
        _clean_filter_values(filtered_retail["REGION"]) if "REGION" in filtered_retail.columns else []
    )
    with columns[1]:
        selections["REGION"] = st.selectbox(
            "Region", region_options, index=0,
            key=f"retail_rm_region_filter_{selections['ZONE']}",
        )
    filtered_retail = _apply_exact_text_filter(filtered_retail, "REGION", selections["REGION"])
 
    if "MKT TYPE" in filtered_retail.columns:
        canonical_market = filtered_retail["MKT TYPE"].map(_scenario10_market_bucket)
        market_options = ["All"] + sorted(
            {
                value for value in canonical_market
                if value not in {"", "Unspecified", "nan", "None"}
            },
            key=lambda value: value.casefold(),
        )
    else:
        canonical_market = pd.Series("Unspecified", index=filtered_retail.index)
        market_options = ["All"]
    with columns[2]:
        selections["MKT TYPE"] = st.selectbox(
            "Market type", market_options, index=0,
            key=f"retail_rm_market_filter_{selections['ZONE']}_{selections['REGION']}",
            help="B30 includes B30 Select; T30 includes T30 Ext.",
        )
    if selections["MKT TYPE"] != "All":
        filtered_retail = filtered_retail.loc[
            canonical_market == selections["MKT TYPE"]
        ].copy()
 
    non_retail = records.loc[records["Vertical"] != "Retail"].copy()
    filtered_records = pd.concat([filtered_retail, non_retail], ignore_index=True, sort=False)
 
    active = [f"{column}: {value}" for column, value in selections.items() if value != "All"]
    active_text = " · ".join(active) if active else "All Retail RMs"
    st.markdown(
        f"<div class='glass-note'><span class='inline-pill gold'>{len(filtered_retail):,} Retail RMs"
        f"</span><span class='inline-pill'>{escape(active_text)}</span></div>",
        unsafe_allow_html=True,
    )
    return filtered_records, selections
 
 
def _star_card(rank: int, row: pd.Series) -> str:
    return (
        "<div class='glass-card stage-card jan'>"
        f"<div class='stage'>Star {rank:02d}</div>"
        f"<div class='metric-value'>{escape(str(row.get('Employee Name', 'RM')))}</div>"
        f"<div class='metric-hero gold' style='font-size:1.7rem'>"
        f"{escape(fmt_pct(row.get('YTD Achievement %')))}</div>"
        "<div class='metric-label'>YTD achievement</div>"
        "<div class='kpi-rows'>"
        + _kpi_row_html("Current run rate", fmt_cr(row.get("Current Run Rate")))
        + _kpi_row_html("Projected FY", fmt_pct(row.get("Projected FY Achievement %")))
        + _kpi_row_html(
            "Contribution", fmt_pct(row.get("Contribution to Overall Target %")), "gold"
        )
        + "</div></div>"
    )
 
 
def render_stars_of_month(detail: pd.DataFrame) -> None:
    section_header("05", "Stars of the month", "Ranked on YTD achievement against YTD target")
 
    if detail.empty:
        glass_note("No RM data is available for this selection.")
        return
 
    ranked = detail.sort_values(
        ["YTD Achievement %", "YTD Achievement"], ascending=[False, False]
    ).reset_index(drop=True)
 
    cards = "".join(_star_card(i + 1, row) for i, (_, row) in enumerate(ranked.head(3).iterrows()))
    st.markdown(f"<div class='trio-grid'>{cards}</div>", unsafe_allow_html=True)
 
    glass_note(
        "Stars are ranked from the fields the scorecard actually provides: overall YTD achievement "
        "against YTD target, with the achievement amount as the tie-breaker. The workbook has no "
        "standalone monthly actual, so no separate monthly score is invented."
    )
 
    columns = [
        c for c in [
            "Employee Name", "Emp Code", "ADID", "REGION", "EM City",
            "Achievement Category", "YTD Achievement %", "Current Run Rate",
            "Projected FY Achievement %", "Contribution to Overall Target %",
        ] if c in ranked.columns
    ]
    render_glass_table(
        ranked.head(10)[columns],
        {
            "Employee Name": "txt", "Emp Code": "txt", "ADID": "txt", "REGION": "txt",
            "EM City": "txt", "Achievement Category": "txt", "YTD Achievement %": "pct",
            "Current Run Rate": "cr", "Projected FY Achievement %": "pct",
            "Contribution to Overall Target %": "pct",
        },
    )
 
 
def render_rm_sales_segmentation(
    records: pd.DataFrame,
    final_metrics: Dict[str, Any],
    vertical: str,
    sales: str,
) -> None:
    final_target = _final_vertical_target(final_metrics, sales, vertical)
    detail = build_rm_performance_detail(records, vertical, sales, final_target)
    if detail.empty:
        glass_note(f"No RM records are available for {vertical} · {SALES_LABEL[sales]}.")
        return
 
    contribution = build_category_contribution(detail, final_target)
 
    projected_total = _num(detail["Estimated FY @ Current RR"].sum()) or 0.0
    ytd_total = _num(detail["YTD Achievement"].sum()) or 0.0
    ytd_target_total = _num(detail["YTD Target"].sum()) or 0.0
    rm_count = len(detail)
    ytd_pct = ytd_total / ytd_target_total if ytd_target_total > 0 else 0.0
    denominator = final_target if (final_target or 0) > 0 else (_num(detail["FY Target"].sum()) or 1.0)
    projected_pct = projected_total / denominator
    high_performers = int(
        detail["Achievement Category"].isin(["100% and above", "90% - 100%"]).sum()
    )
 
    kpi_strip([
        {"label": "RMs in view", "value": fmt_num(rm_count),
         "secondary": f"{vertical} · {SALES_LABEL[sales]}"},
        {"label": "YTD vs YTD target", "value": fmt_pct(ytd_pct),
         "secondary": fmt_cr(ytd_total)},
        {"label": "Projected FY achievement", "value": fmt_pct(projected_pct),
         "delta": fmt_pts(projected_pct - 1.0), "tone": _tone_for(projected_pct - 1.0),
         "secondary": fmt_cr(projected_total)},
        {"label": "RMs at 90% or better", "value": fmt_num(high_performers),
         "secondary": fmt_pct(high_performers / rm_count if rm_count else 0.0)},
        {"label": "FINAL FY27 target", "value": fmt_cr(final_target),
         "secondary": "published channel target"},
    ])
 
    section_header("02", "Achievement bands", "How the RM population is distributed")
 
    counts = dict(zip(contribution["Achievement Category"], contribution["RM Count"]))
    kpi_strip([
        {"label": band, "value": fmt_num(counts.get(band, 0)), "secondary": "RMs"}
        for band in ACHIEVEMENT_BAND_ORDER
    ])
 
    section_header("03", "Run-rate contribution to target", "If each band holds its current pace")
    glass_note(
        "Each band's RMs are annualised at their current run rate and divided by the FY27 target "
        "from FINAL for the same channel and sales basis. The result is how many percentage points "
        "of the overall target each band is on track to deliver."
    )
    render_glass_table(
        contribution,
        {
            "Achievement Category": "txt", "RM Count": "num", "FY Target": "cr",
            "YTD Target": "cr", "YTD Achievement": "cr", "Current YTD Achievement %": "pct",
            "Current Run Rate": "cr", "Estimated FY @ Current RR": "cr",
            "Category Projected FY %": "pct", "Contribution to Overall Target %": "pct",
            "Share of Projected Sales %": "pct",
        },
    )
 
    total_contribution = _num(contribution["Contribution to Overall Target %"].sum()) or 0.0
    glass_callout(
        f"<b>{vertical} · {SALES_LABEL[sales]}:</b> at today's RM run rates the six bands together "
        f"are on track to deliver <b>{fmt_pct(total_contribution)}</b> of the FINAL FY27 target"
        + (f" ({fmt_cr(final_target)})." if final_target is not None else ".")
    )
 
    section_header("04", "RM drill-down", "Every RM inside the selected band")
    band = st.radio(
        "Achievement band", ACHIEVEMENT_BAND_ORDER, index=0, horizontal=True,
        key=f"rm_band_{vertical}_{sales}", label_visibility="collapsed",
    )
    rows = detail.loc[detail["Achievement Category"] == band].copy()
    if rows.empty:
        glass_note(f"No RMs fall in {band} for this selection.")
    else:
        columns = [
            c for c in [
                "Employee Name", "Emp Code", "ADID", "ZONE", "REGION", "EM City",
                "Achievement Category", "FY Target", "YTD Target", "YTD Achievement",
                "YTD Achievement %", "Current Run Rate", "Estimated FY @ Current RR",
                "Projected FY Achievement %", "Contribution to Overall Target %",
            ] if c in rows.columns
        ]
        render_glass_table(
            rows[columns],
            {
                "Employee Name": "txt", "Emp Code": "txt", "ADID": "txt", "ZONE": "txt",
                "REGION": "txt", "EM City": "txt", "Achievement Category": "txt",
                "FY Target": "cr", "YTD Target": "cr", "YTD Achievement": "cr",
                "YTD Achievement %": "pct", "Current Run Rate": "cr",
                "Estimated FY @ Current RR": "cr", "Projected FY Achievement %": "pct",
                "Contribution to Overall Target %": "pct",
            },
        )
 
    render_stars_of_month(detail)
 
 
def render_rm_segmentation_page(records: pd.DataFrame, payload: bytes) -> None:
    """RM achievement segmentation and contribution analysis."""
    final_metrics = parse_final_dashboard_metrics(payload)
    render_apple_header(final_metrics, records, "FY27 · RM performance")
 
    section_header(
        "01",
        "RM performance segmentation",
        "Retail · DHNI · VRM, banded by achievement against the YTD target",
    )
    st.markdown(
        "".join(
            f"<span class='inline-pill{' gold' if band.startswith('100') else ''}'>{escape(band)}</span>"
            for band in ACHIEVEMENT_BAND_ORDER
        ),
        unsafe_allow_html=True,
    )
 
    st.markdown("<div class='metric-label'>Channel</div>", unsafe_allow_html=True)
    vertical = st.radio(
        "Channel", VERTICALS, index=0, horizontal=True,
        key="rm_seg_vertical", label_visibility="collapsed",
    )
 
    page_records = records
    selections: Dict[str, str] = {"ZONE": "All", "REGION": "All", "MKT TYPE": "All"}
    if vertical == "Retail":
        page_records, selections = render_retail_rm_filters(records)
 
    tabs = st.tabs([SALES_LABEL["NS"], SALES_LABEL["GS"]])
    for tab, sales in zip(tabs, ["NS", "GS"]):
        with tab:
            render_rm_sales_segmentation(page_records, final_metrics, vertical, sales)
 
    section_header("06", "Export", "The full RM segmentation pack")
    try:
        export_payload = make_rm_segmentation_export(page_records, final_metrics)
        st.download_button(
            "Download RM segmentation workbook",
            data=export_payload,
            file_name="rm_performance_segmentation.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        if vertical == "Retail":
            active = [f"{k}: {v}" for k, v in selections.items() if v != "All"]
            st.caption(
                "Retail sheets follow the filters above"
                + (f" ({' · '.join(active)})." if active else " (all Retail RMs).")
                + " DHNI and VRM remain unfiltered."
            )
    except Exception:  # pragma: no cover - defensive
        st.warning("The RM segmentation export could not be generated for this selection.")
 
    glass_note("Undefined numeric outputs are shown as 0.")
 
 

# =============================================================================
# 23A. BON VOYAGE + AI CHOICE
# =============================================================================

BON_VOYAGE_MARKETS: List[str] = ["T2", "T6", "T30", "B30", "EM"]
BON_VOYAGE_DIVISIONS: List[str] = ["Retail", "Insti", "Digital", "VRM", "DHNI"]
BON_VOYAGE_ASSETS: List[str] = ["Equity", "Debt", "Liquid"]

BON_VOYAGE_CUTS: Dict[str, Tuple[Optional[str], str]] = {
    "ALL": (None, "Overall"),
}
for _division in BON_VOYAGE_DIVISIONS:
    for _asset in BON_VOYAGE_ASSETS:
        BON_VOYAGE_CUTS[f"{_division.upper()} {_asset.upper()}"] = (
            _division,
            _asset,
        )


def inject_voyage_experience_css() -> None:
    """Extra page-level styling for Bon voyage and AI Choice."""
    st.markdown(
        """
        <style>
        .bv-hero {
            position: relative;
            overflow: hidden;
            border: 1px solid rgba(255,255,255,.55);
            border-radius: 26px;
            padding: 28px 30px;
            margin: 4px 0 20px 0;
            background:
                radial-gradient(circle at 8% 10%, rgba(255,232,74,.42), transparent 25%),
                radial-gradient(circle at 92% 18%, rgba(0,255,209,.24), transparent 25%),
                linear-gradient(130deg, #073b4c 0%, #005f73 48%, #0a9396 100%);
            box-shadow:
                14px 14px 30px rgba(15, 46, 57, .18),
                -10px -10px 24px rgba(255,255,255,.55);
            color: white;
        }
        .bv-hero .eyebrow {
            text-transform: uppercase;
            letter-spacing: .18em;
            font-size: .72rem;
            font-weight: 900;
            color: #ffef65;
        }
        .bv-hero .title {
            margin-top: 8px;
            font-size: 2rem;
            line-height: 1.08;
            font-weight: 900;
            color: white;
        }
        .bv-hero .sub {
            margin-top: 8px;
            max-width: 920px;
            font-size: .92rem;
            line-height: 1.55;
            color: rgba(255,255,255,.84);
        }
        .bv-cut-banner {
            border: 1px solid rgba(95, 73, 19, .20);
            border-radius: 18px;
            padding: 14px 18px;
            margin: 8px 0 12px 0;
            background:
                linear-gradient(145deg, rgba(255,248,210,.96), rgba(255,225,82,.72));
            box-shadow:
                8px 8px 18px rgba(98,75,20,.10),
                -8px -8px 18px rgba(255,255,255,.85);
        }
        .bv-cut-banner .cut {
            font-size: .72rem;
            letter-spacing: .12em;
            text-transform: uppercase;
            color: #7c5711;
            font-weight: 900;
        }
        .bv-cut-banner .name {
            margin-top: 3px;
            font-size: 1.2rem;
            font-weight: 900;
            color: #2d2515;
        }
        .bv-threshold-title {
            color: #073b4c;
            font-weight: 900;
            font-size: .82rem;
            text-transform: uppercase;
            letter-spacing: .08em;
            margin: 15px 0 5px 0;
        }
        .bv-summary {
            background: linear-gradient(135deg, rgba(255,255,255,.72), rgba(228,255,252,.72));
            border: 1px solid rgba(10,147,150,.20);
            border-radius: 20px;
            padding: 16px;
            box-shadow:
                inset 2px 2px 0 rgba(255,255,255,.75),
                8px 8px 22px rgba(3, 67, 78, .08);
        }
        .bv-table .glass-table thead th {
            background: linear-gradient(135deg, #075985, #0891b2) !important;
            color: white !important;
        }
        .bv-table .glass-table tbody tr.total td {
            background: #fff724 !important;
            color: #1f2937 !important;
            font-weight: 900 !important;
        }
        .bv-detail .glass-table thead th {
            background: linear-gradient(135deg, #4c1d95, #7c3aed) !important;
            color: white !important;
        }

        .ai-hero {
            position: relative;
            overflow: hidden;
            border-radius: 30px;
            padding: 30px;
            margin: 4px 0 18px 0;
            color: #fff;
            border: 1px solid rgba(255,255,255,.55);
            background:
                radial-gradient(circle at 15% 20%, rgba(255,244,71,.92), transparent 18%),
                radial-gradient(circle at 85% 22%, rgba(0,255,209,.65), transparent 22%),
                radial-gradient(circle at 78% 85%, rgba(255,61,172,.72), transparent 22%),
                linear-gradient(125deg, #6d28d9 0%, #d946ef 34%, #f97316 67%, #06b6d4 100%);
            box-shadow:
                15px 15px 32px rgba(103, 35, 122, .18),
                -12px -12px 28px rgba(255,255,255,.72);
        }
        .ai-hero .eyebrow {
            display: inline-block;
            background: rgba(0,0,0,.23);
            border: 1px solid rgba(255,255,255,.4);
            backdrop-filter: blur(10px);
            border-radius: 999px;
            padding: 5px 10px;
            font-weight: 900;
            font-size: .70rem;
            text-transform: uppercase;
            letter-spacing: .15em;
            color: #fff;
        }
        .ai-hero .title {
            font-size: 2.35rem;
            font-weight: 950;
            line-height: 1.05;
            margin-top: 12px;
            color: #fff;
        }
        .ai-hero .sub {
            max-width: 980px;
            margin-top: 10px;
            color: rgba(255,255,255,.92);
            font-size: .95rem;
            line-height: 1.55;
        }
        .ai-glow-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(190px,1fr));
            gap: 14px;
            margin: 10px 0 18px 0;
        }
        .ai-glow {
            min-height: 135px;
            padding: 16px;
            border-radius: 22px;
            border: 1px solid rgba(255,255,255,.58);
            backdrop-filter: blur(18px);
            box-shadow:
                10px 10px 24px rgba(57,39,98,.11),
                -8px -8px 20px rgba(255,255,255,.80);
        }
        .ai-glow:nth-child(5n+1) { background: linear-gradient(145deg,#fff56b,#ffd43b); }
        .ai-glow:nth-child(5n+2) { background: linear-gradient(145deg,#77f7df,#31d7c4); }
        .ai-glow:nth-child(5n+3) { background: linear-gradient(145deg,#ff93d2,#ff5fb2); }
        .ai-glow:nth-child(5n+4) { background: linear-gradient(145deg,#b9a3ff,#8b5cf6); color:white; }
        .ai-glow:nth-child(5n+5) { background: linear-gradient(145deg,#ffb66e,#ff7a18); color:white; }
        .ai-glow .label {
            font-size: .69rem;
            font-weight: 900;
            letter-spacing: .10em;
            text-transform: uppercase;
            opacity: .72;
        }
        .ai-glow .value {
            font-size: 1.65rem;
            font-weight: 950;
            margin-top: 8px;
        }
        .ai-glow .mini {
            margin-top: 7px;
            font-size: .77rem;
            line-height: 1.35;
            opacity: .82;
        }
        .ai-panel {
            border-radius: 24px;
            padding: 18px;
            margin: 10px 0 15px 0;
            background: rgba(255,255,255,.58);
            border: 1px solid rgba(255,255,255,.78);
            backdrop-filter: blur(20px);
            box-shadow:
                10px 10px 25px rgba(60,43,102,.09),
                -10px -10px 22px rgba(255,255,255,.80);
        }
        .ai-panel .ttl {
            font-weight: 950;
            color: #3b0764;
            font-size: 1rem;
        }
        .ai-panel .sub {
            margin-top: 3px;
            color: #6b7280;
            font-size: .78rem;
        }
        .ai-table .glass-table thead th {
            background: linear-gradient(135deg,#7c3aed,#ec4899,#f97316) !important;
            color: white !important;
        }
        .ai-trip .glass-table thead th {
            background: linear-gradient(135deg,#0369a1,#06b6d4,#22c55e) !important;
            color:white !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def _bon_voyage_first_nonempty(series: pd.Series) -> str:
    for value in series:
        value = "" if value is None else str(value).strip()
        if value and value.casefold() not in {"nan", "none"}:
            return value
    return ""


def _bon_voyage_rm_key(frame: pd.DataFrame) -> pd.Series:
    adid = text_column(frame, "ADID")
    emp_code = text_column(frame, "Emp Code")
    name = text_column(frame, "Employee Name")

    keys: List[str] = []
    for position in range(len(frame)):
        if adid.iloc[position]:
            keys.append(f"ADID::{adid.iloc[position]}")
        elif emp_code.iloc[position]:
            keys.append(f"EMP::{emp_code.iloc[position]}")
        elif name.iloc[position]:
            keys.append(f"NAME::{name.iloc[position]}")
        else:
            keys.append(f"ROW::{position}")
    return pd.Series(keys, index=frame.index, dtype="object")


def _bon_voyage_division(row: pd.Series) -> str:
    channel = str(row.get("Channel", "")).strip().casefold()
    vertical = str(row.get("Vertical", "")).strip().casefold()

    if channel == "digital":
        return "Digital"
    if channel in {"institutional", "insti"}:
        return "Insti"
    if vertical == "vrm":
        return "VRM"
    if vertical == "dhni":
        return "DHNI"
    return "Retail"


def _bon_voyage_market(value: Any) -> str:
    return _scenario10_market_bucket(value)


def build_bon_voyage_rm_detail(
    records: pd.DataFrame,
    projection_months: int,
) -> pd.DataFrame:
    """
    RM-level Net Sales projection used by Bon voyage and AI Choice.

    Run Rate = YTD Net Sales / 3 completed months
    Projection = YTD Net Sales + Run Rate × editable remaining months
    """
    if records is None or records.empty:
        return pd.DataFrame()

    work = records.copy()
    work["_BV RM Key"] = _bon_voyage_rm_key(work)
    work["Bon Voyage Division"] = work.apply(_bon_voyage_division, axis=1)

    if "MKT TYPE" in work.columns:
        work["Market Bucket"] = work["MKT TYPE"].map(_bon_voyage_market)
    else:
        work["Market Bucket"] = "Unspecified"

    for asset in ASSETS:
        for role in ("fy", "ach"):
            column = f"NS_{asset}_{role}"
            if column not in work.columns:
                work[column] = 0.0
            work[column] = pd.to_numeric(
                work[column], errors="coerce"
            ).fillna(0.0)

    metadata = [
        column
        for column in [
            "Employee Name", "Emp Code", "ADID", "Status", "Type",
            "ZONE", "REGION", "EM City", "MKT TYPE",
        ]
        if column in work.columns
    ]

    aggregations: Dict[str, Any] = {
        column: _bon_voyage_first_nonempty for column in metadata
    }
    for asset in ASSETS:
        aggregations[f"NS_{asset}_fy"] = "sum"
        aggregations[f"NS_{asset}_ach"] = "sum"

    grouped = (
        work.groupby(
            ["_BV RM Key", "Bon Voyage Division", "Market Bucket"],
            dropna=False,
            as_index=False,
        )
        .agg(aggregations)
    )

    months = max(int(projection_months), 0)

    for asset in ASSETS:
        target = pd.to_numeric(
            grouped[f"NS_{asset}_fy"], errors="coerce"
        ).fillna(0.0)
        ytd = pd.to_numeric(
            grouped[f"NS_{asset}_ach"], errors="coerce"
        ).fillna(0.0)
        rr = ytd / max(MONTHS_COMPLETED, 1)
        projection = ytd + rr * months

        grouped[f"{asset} FY Target"] = target
        grouped[f"{asset} YTD NS"] = ytd
        grouped[f"{asset} Current RR"] = rr
        grouped[f"{asset} Projected NS"] = projection
        grouped[f"{asset} Projected %"] = np.where(
            target > 0,
            projection / target,
            0.0,
        )

    grouped["Overall FY Target"] = sum(
        grouped[f"{asset} FY Target"] for asset in ASSETS
    )
    grouped["Overall YTD NS"] = sum(
        grouped[f"{asset} YTD NS"] for asset in ASSETS
    )
    grouped["Overall Current RR"] = (
        grouped["Overall YTD NS"] / max(MONTHS_COMPLETED, 1)
    )
    grouped["Overall Projected NS"] = (
        grouped["Overall YTD NS"]
        + grouped["Overall Current RR"] * months
    )
    grouped["Overall Projected %"] = np.where(
        grouped["Overall FY Target"] > 0,
        grouped["Overall Projected NS"] / grouped["Overall FY Target"],
        0.0,
    )
    grouped["Projection Months"] = months

    return grouped


def _bon_voyage_collapse_all_divisions(detail: pd.DataFrame) -> pd.DataFrame:
    """One RM × market row across all management divisions."""
    if detail.empty:
        return detail.copy()

    metadata = [
        column
        for column in [
            "Employee Name", "Emp Code", "ADID", "Status", "Type",
            "ZONE", "REGION", "EM City", "MKT TYPE",
        ]
        if column in detail.columns
    ]
    aggregations: Dict[str, Any] = {
        column: _bon_voyage_first_nonempty for column in metadata
    }

    for dimension in ["Overall", *ASSETS]:
        for suffix in ("FY Target", "YTD NS", "Current RR", "Projected NS"):
            aggregations[f"{dimension} {suffix}"] = "sum"

    result = (
        detail.groupby(
            ["_BV RM Key", "Market Bucket"],
            dropna=False,
            as_index=False,
        )
        .agg(aggregations)
    )
    result["Bon Voyage Division"] = "All"

    for dimension in ["Overall", *ASSETS]:
        target = pd.to_numeric(
            result[f"{dimension} FY Target"], errors="coerce"
        ).fillna(0.0)
        projected = pd.to_numeric(
            result[f"{dimension} Projected NS"], errors="coerce"
        ).fillna(0.0)
        result[f"{dimension} Projected %"] = np.where(
            target > 0,
            projected / target,
            0.0,
        )

    result["Projection Months"] = (
        int(detail["Projection Months"].iloc[0])
        if "Projection Months" in detail.columns and not detail.empty
        else MONTHS_REMAINING
    )
    return result


def _bon_voyage_cut_definition(
    label: str,
) -> Tuple[Optional[str], str]:
    return BON_VOYAGE_CUTS.get(label, (None, "Overall"))


def _bon_voyage_scope_for_cut(
    detail: pd.DataFrame,
    cut_label: str,
) -> Tuple[pd.DataFrame, str, str]:
    division, dimension = _bon_voyage_cut_definition(cut_label)

    if division is None:
        scope = _bon_voyage_collapse_all_divisions(detail)
        cut_name = "All employees · Overall Net Sales"
    else:
        scope = detail.loc[
            detail["Bon Voyage Division"] == division
        ].copy()
        cut_name = f"{division} · {dimension} Net Sales"

    return scope, dimension, cut_name


def _bon_voyage_metric_columns(
    dimension: str,
) -> Tuple[str, str, str, str]:
    return (
        f"{dimension} FY Target",
        f"{dimension} YTD NS",
        f"{dimension} Current RR",
        f"{dimension} Projected NS",
    )


def _bon_voyage_thresholds_for_cut(
    cut_label: str,
    key_prefix: str,
) -> Dict[str, float]:
    thresholds: Dict[str, float] = {}
    st.markdown(
        "<div class='bv-threshold-title'>Edit location thresholds</div>",
        unsafe_allow_html=True,
    )
    columns = st.columns(len(BON_VOYAGE_MARKETS))

    for index, market in enumerate(BON_VOYAGE_MARKETS):
        key = (
            f"{key_prefix}_{_scenario10_slug(cut_label)}_"
            f"{_scenario10_slug(market)}"
        )
        with columns[index]:
            value = st.number_input(
                f"{market} threshold %",
                min_value=0.0,
                value=float(st.session_state.get(key, 100.0)),
                step=5.0,
                format="%.1f",
                key=key,
                help=(
                    "RM qualifies when its projected achievement for the "
                    "selected cut reaches this threshold."
                ),
            )
        thresholds[market] = float(value) / 100.0
    return thresholds


def _bon_voyage_location_results(
    scope: pd.DataFrame,
    dimension: str,
    thresholds: Dict[str, float],
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Screenshot-style location result and the qualified RM population."""
    target_col, _, _, projection_col = _bon_voyage_metric_columns(dimension)
    pct_col = f"{dimension} Projected %"

    rows: List[Dict[str, Any]] = []
    qualified_parts: List[pd.DataFrame] = []

    total_population_keys: set = set()
    total_qualified_keys: set = set()
    total_amount = 0.0
    total_target = 0.0

    for market in BON_VOYAGE_MARKETS:
        subset = scope.loc[
            scope["Market Bucket"] == market
        ].copy()

        threshold = thresholds.get(market, 1.0)
        projected_pct = pd.to_numeric(
            subset[pct_col], errors="coerce"
        ).fillna(0.0)
        qualified = subset.loc[
            projected_pct >= threshold
        ].copy()

        if not qualified.empty:
            qualified["Threshold Applied"] = threshold
            qualified_parts.append(qualified)

        population = int(subset["_BV RM Key"].nunique()) if not subset.empty else 0
        qualified_count = (
            int(qualified["_BV RM Key"].nunique())
            if not qualified.empty else 0
        )
        amount = (
            float(
                pd.to_numeric(
                    qualified[projection_col], errors="coerce"
                ).fillna(0.0).sum()
            )
            if not qualified.empty else 0.0
        )
        location_target = (
            float(
                pd.to_numeric(
                    subset[target_col], errors="coerce"
                ).fillna(0.0).sum()
            )
            if not subset.empty else 0.0
        )
        contribution = (
            amount / location_target
            if location_target != 0 else 0.0
        )

        if not subset.empty:
            total_population_keys.update(
                subset["_BV RM Key"].astype(str).tolist()
            )
        if not qualified.empty:
            total_qualified_keys.update(
                qualified["_BV RM Key"].astype(str).tolist()
            )

        total_amount += amount
        total_target += location_target

        rows.append({
            "Location": market,
            "Employees": population,
            "Threshold": threshold,
            "Qualified RMs": qualified_count,
            "NS Contribution": amount,
            f"Location-wise {dimension} NS Target": location_target,
            "Contribution": contribution,
        })

    overall = {
        "Location": "OVERALL",
        "Employees": len(total_population_keys),
        "Threshold": None,
        "Qualified RMs": len(total_qualified_keys),
        "NS Contribution": total_amount,
        f"Location-wise {dimension} NS Target": total_target,
        "Contribution": (
            total_amount / total_target
            if total_target != 0 else 0.0
        ),
    }

    result = pd.DataFrame([overall, *rows])
    qualified_all = (
        pd.concat(qualified_parts, ignore_index=True)
        if qualified_parts else pd.DataFrame()
    )
    return result, qualified_all


def _bon_voyage_location_formats(
    dimension: str,
) -> Dict[str, str]:
    return {
        "Location": "txt",
        "Employees": "num",
        "Threshold": "pct",
        "Qualified RMs": "num",
        "NS Contribution": "cr",
        f"Location-wise {dimension} NS Target": "cr",
        "Contribution": "pct",
    }


def _bon_voyage_rm_detail_table(
    qualified: pd.DataFrame,
    selected_dimension: str,
) -> Tuple[pd.DataFrame, Dict[str, str]]:
    if qualified.empty:
        return pd.DataFrame(), {}

    columns = [
        column
        for column in [
            "Employee Name", "Emp Code", "ADID", "Status", "Type",
            "ZONE", "REGION", "EM City", "MKT TYPE", "Market Bucket",
            "Bon Voyage Division", "Threshold Applied",
            f"{selected_dimension} FY Target",
            f"{selected_dimension} YTD NS",
            f"{selected_dimension} Current RR",
            f"{selected_dimension} Projected NS",
            f"{selected_dimension} Projected %",
            "Overall FY Target", "Overall YTD NS",
            "Overall Current RR", "Overall Projected NS",
            "Overall Projected %",
            "Equity Projected NS", "Equity Projected %",
            "Debt Projected NS", "Debt Projected %",
            "Liquid Projected NS", "Liquid Projected %",
        ]
        if column in qualified.columns
    ]

    formats: Dict[str, str] = {}
    for column in columns:
        if column in {
            "Threshold Applied",
            f"{selected_dimension} Projected %",
            "Overall Projected %",
            "Equity Projected %",
            "Debt Projected %",
            "Liquid Projected %",
        }:
            formats[column] = "pct"
        elif any(
            phrase in column
            for phrase in (
                "FY Target", "YTD NS", "Current RR", "Projected NS"
            )
        ):
            formats[column] = "cr"
        else:
            formats[column] = "txt"

    return qualified[columns].copy(), formats


def _bon_voyage_projected_revenue(
    frame: pd.DataFrame,
) -> float:
    if frame is None or frame.empty:
        return 0.0

    total = 0.0
    for asset in ASSETS:
        column = f"{asset} Projected NS"
        if column not in frame.columns:
            continue
        amount = float(
            pd.to_numeric(
                frame[column], errors="coerce"
            ).fillna(0.0).sum()
        )
        total += amount * REVENUE_RATE.get(asset, 0.0)
    return total


def render_bon_voyage_page(
    records: pd.DataFrame,
    payload: bytes,
) -> None:
    """User-friendly screenshot-style threshold simulator."""
    inject_voyage_experience_css()
    final_metrics = parse_final_dashboard_metrics(payload)

    st.markdown(
        """
        <div class='bv-hero'>
          <div class='eyebrow'>Current NS Achievement · #Employees</div>
          <div class='title'>Bon voyage · Outperformer contribution simulator</div>
          <div class='sub'>
            Pick one business cut, edit the five location thresholds, and immediately see
            how many regional managers qualify, the projected Net Sales they contribute,
            the location-wise FY Net Sales target and the resulting contribution percentage.
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    controls = st.columns([1.4, 1.0, 2.3])
    with controls[0]:
        cut_options = list(BON_VOYAGE_CUTS.keys())
        cut_label = st.selectbox(
            "Select scenario cut",
            cut_options,
            index=cut_options.index(
                st.session_state.get("bv_selected_cut", "RETAIL EQUITY")
                if st.session_state.get("bv_selected_cut", "RETAIL EQUITY") in cut_options
                else "RETAIL EQUITY"
            ),
            key="bv_selected_cut",
            help=(
                "Choose ALL or a Division × Asset cut such as RETAIL EQUITY, "
                "DHNI DEBT, VRM LIQUID, INSTI EQUITY or DIGITAL DEBT."
            ),
        )
    with controls[1]:
        projection_months = int(
            st.number_input(
                "Months to project",
                min_value=0,
                max_value=12,
                value=int(
                    st.session_state.get(
                        "bv_projection_months",
                        MONTHS_REMAINING,
                    )
                ),
                step=1,
                key="bv_projection_months",
                help=(
                    "Projection = YTD Net Sales + "
                    "(YTD Net Sales ÷ 3) × selected months."
                ),
            )
        )
    with controls[2]:
        glass_callout(
            "<b>Qualification basis:</b> projected achievement % for the selected "
            "Division × Asset cut. Thresholds are location-specific and default to 100%."
        )

    detail = build_bon_voyage_rm_detail(
        records,
        projection_months,
    )
    if detail.empty:
        glass_note("No RM-level Net Sales records are available.")
        return

    scope, dimension, cut_name = _bon_voyage_scope_for_cut(
        detail,
        cut_label,
    )

    st.markdown(
        f"""
        <div class='bv-cut-banner'>
          <div class='cut'>Scenario cut</div>
          <div class='name'>{escape(cut_name)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    thresholds = _bon_voyage_thresholds_for_cut(
        cut_label,
        "bv_screen",
    )
    result, qualified = _bon_voyage_location_results(
        scope,
        dimension,
        thresholds,
    )

    section_header(
        "01",
        "Location-wise threshold contribution",
        (
            "Employees → editable threshold → qualified RMs → projected NS contribution "
            "→ location-wise NS target → contribution"
        ),
    )
    render_glass_table(
        result,
        _bon_voyage_location_formats(dimension),
        total_rows=("OVERALL",),
        css_class="bv-table",
    )

    overall = result.iloc[0]
    kpi_strip([
        {
            "label": "Employees in selected cut",
            "value": fmt_num(overall.get("Employees")),
            "secondary": cut_name,
        },
        {
            "label": "Qualified RMs",
            "value": fmt_num(overall.get("Qualified RMs")),
            "secondary": "after location thresholds",
        },
        {
            "label": "NS contribution",
            "value": fmt_cr(overall.get("NS Contribution")),
            "secondary": f"projected {dimension} NS",
        },
        {
            "label": "Selected-cut target",
            "value": fmt_cr(
                overall.get(f"Location-wise {dimension} NS Target")
            ),
            "secondary": f"{dimension} FY NS target",
        },
        {
            "label": "Contribution",
            "value": fmt_pct(overall.get("Contribution")),
            "secondary": "qualified projected NS ÷ target",
        },
    ])

    section_header(
        "02",
        "Qualified regional managers",
        f"RM detail for {cut_name} using the five thresholds above",
    )
    detail_table, detail_formats = _bon_voyage_rm_detail_table(
        qualified,
        dimension,
    )
    render_glass_table(
        detail_table,
        detail_formats,
        max_html_rows=350,
        css_class="bv-detail",
        empty_message=(
            "No regional managers qualify for the current thresholds."
        ),
    )

    if not detail_table.empty:
        st.download_button(
            "Download this qualified RM cut",
            data=detail_table.to_csv(index=False).encode("utf-8-sig"),
            file_name=(
                f"bon_voyage_{_scenario10_slug(cut_label)}_qualified.csv"
            ),
            mime="text/csv",
            key="bv_download_current_cut",
        )

    glass_note(
        "B30 Select is combined into B30 and T30 Ext is combined into T30. "
        "Contribution = projected Net Sales of qualified RMs ÷ the location-wise "
        "Net Sales FY target for the selected cut."
    )


def _ai_summary_by_market(
    scope: pd.DataFrame,
    dimension: str,
    threshold: float,
) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    target_col = f"{dimension} FY Target"
    projection_col = f"{dimension} Projected NS"
    pct_col = f"{dimension} Projected %"

    for market in BON_VOYAGE_MARKETS:
        subset = scope.loc[
            scope["Market Bucket"] == market
        ].copy()
        qualified = subset.loc[
            pd.to_numeric(
                subset[pct_col], errors="coerce"
            ).fillna(0.0) >= threshold
        ].copy()

        employees = int(subset["_BV RM Key"].nunique()) if not subset.empty else 0
        q = int(qualified["_BV RM Key"].nunique()) if not qualified.empty else 0
        amount = float(
            pd.to_numeric(
                qualified[projection_col], errors="coerce"
            ).fillna(0.0).sum()
        ) if not qualified.empty else 0.0
        target = float(
            pd.to_numeric(
                subset[target_col], errors="coerce"
            ).fillna(0.0).sum()
        ) if not subset.empty else 0.0

        rows.append({
            "Market": market,
            "Employees": employees,
            "Qualified": q,
            "Qualification Rate": q / employees if employees else 0.0,
            "Qualified Projected NS": amount,
            "FY NS Target": target,
            "Contribution": amount / target if target else 0.0,
            "Projected Revenue": _bon_voyage_projected_revenue(qualified),
        })
    return pd.DataFrame(rows)


def _ai_summary_by_division(
    detail: pd.DataFrame,
    threshold: float,
) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []

    for division in BON_VOYAGE_DIVISIONS:
        subset = detail.loc[
            detail["Bon Voyage Division"] == division
        ].copy()
        qualified = subset.loc[
            pd.to_numeric(
                subset["Overall Projected %"], errors="coerce"
            ).fillna(0.0) >= threshold
        ].copy()

        employees = int(subset["_BV RM Key"].nunique()) if not subset.empty else 0
        q = int(qualified["_BV RM Key"].nunique()) if not qualified.empty else 0
        amount = float(
            pd.to_numeric(
                qualified["Overall Projected NS"], errors="coerce"
            ).fillna(0.0).sum()
        ) if not qualified.empty else 0.0
        target = float(
            pd.to_numeric(
                subset["Overall FY Target"], errors="coerce"
            ).fillna(0.0).sum()
        ) if not subset.empty else 0.0

        rows.append({
            "Division": division,
            "Employees": employees,
            "Qualified": q,
            "Qualification Rate": q / employees if employees else 0.0,
            "Qualified Projected NS": amount,
            "FY NS Target": target,
            "Contribution": amount / target if target else 0.0,
            "Projected Revenue": _bon_voyage_projected_revenue(qualified),
        })
    return pd.DataFrame(rows)


def _ai_summary_by_asset(
    all_scope: pd.DataFrame,
    threshold: float,
) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []

    for asset in ASSETS:
        pct_col = f"{asset} Projected %"
        amount_col = f"{asset} Projected NS"
        target_col = f"{asset} FY Target"

        qualified = all_scope.loc[
            pd.to_numeric(
                all_scope[pct_col], errors="coerce"
            ).fillna(0.0) >= threshold
        ].copy()

        employees = int(all_scope["_BV RM Key"].nunique())
        q = int(qualified["_BV RM Key"].nunique()) if not qualified.empty else 0
        amount = float(
            pd.to_numeric(
                qualified[amount_col], errors="coerce"
            ).fillna(0.0).sum()
        ) if not qualified.empty else 0.0
        target = float(
            pd.to_numeric(
                all_scope[target_col], errors="coerce"
            ).fillna(0.0).sum()
        )
        revenue = amount * REVENUE_RATE.get(asset, 0.0)

        rows.append({
            "Asset": asset,
            "Employees": employees,
            "Qualified": q,
            "Qualification Rate": q / employees if employees else 0.0,
            "Qualified Projected NS": amount,
            "FY NS Target": target,
            "Contribution": amount / target if target else 0.0,
            "Projected Revenue": revenue,
        })
    return pd.DataFrame(rows)


def _ai_near_miss_pool(
    all_scope: pd.DataFrame,
    threshold: float,
) -> pd.DataFrame:
    pct = pd.to_numeric(
        all_scope["Overall Projected %"], errors="coerce"
    ).fillna(0.0)

    definitions = [
        ("Within 5 pp", max(threshold - 0.05, 0.0), threshold),
        ("5–10 pp away", max(threshold - 0.10, 0.0), max(threshold - 0.05, 0.0)),
        ("10–20 pp away", max(threshold - 0.20, 0.0), max(threshold - 0.10, 0.0)),
    ]

    rows = []
    for label, lower, upper in definitions:
        subset = all_scope.loc[
            (pct >= lower) & (pct < upper)
        ].copy()
        rows.append({
            "Conversion Pool": label,
            "Employees": int(subset["_BV RM Key"].nunique()) if not subset.empty else 0,
            "Projected NS": float(
                pd.to_numeric(
                    subset["Overall Projected NS"], errors="coerce"
                ).fillna(0.0).sum()
            ) if not subset.empty else 0.0,
            "Projected Revenue": _bon_voyage_projected_revenue(subset),
            "Avg Projected Achievement": float(
                pd.to_numeric(
                    subset["Overall Projected %"], errors="coerce"
                ).fillna(0.0).mean()
            ) if not subset.empty else 0.0,
        })
    return pd.DataFrame(rows)


def _ai_glow_cards(cards: Sequence[Dict[str, str]]) -> None:
    blocks = []
    for card in cards:
        blocks.append(
            "<div class='ai-glow'>"
            f"<div class='label'>{escape(card.get('label',''))}</div>"
            f"<div class='value'>{escape(card.get('value',''))}</div>"
            f"<div class='mini'>{escape(card.get('mini',''))}</div>"
            "</div>"
        )
    st.markdown(
        "<div class='ai-glow-grid'>" + "".join(blocks) + "</div>",
        unsafe_allow_html=True,
    )


def _ai_trip_rows(
    scope: pd.DataFrame,
    dimension: str,
    threshold: float,
    cost_per_employee_lakh: float,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    pct_col = f"{dimension} Projected %"
    selected_amount_col = f"{dimension} Projected NS"

    rows = []
    all_qualified_parts: List[pd.DataFrame] = []

    for market in BON_VOYAGE_MARKETS:
        subset = scope.loc[
            scope["Market Bucket"] == market
        ].copy()
        qualified = subset.loc[
            pd.to_numeric(
                subset[pct_col], errors="coerce"
            ).fillna(0.0) >= threshold
        ].copy()

        if not qualified.empty:
            all_qualified_parts.append(qualified)

        count = int(qualified["_BV RM Key"].nunique()) if not qualified.empty else 0
        selected_ns = float(
            pd.to_numeric(
                qualified[selected_amount_col], errors="coerce"
            ).fillna(0.0).sum()
        ) if not qualified.empty else 0.0
        overall_ns = float(
            pd.to_numeric(
                qualified["Overall Projected NS"], errors="coerce"
            ).fillna(0.0).sum()
        ) if not qualified.empty else 0.0
        revenue = _bon_voyage_projected_revenue(qualified)
        spend_lakh = count * float(cost_per_employee_lakh)
        spend_cr = spend_lakh / 100.0

        rows.append({
            "Market": market,
            "Qualified Employees": count,
            "Cost / Employee (₹ Lakh)": cost_per_employee_lakh,
            "Total Trip Spend (₹ Lakh)": spend_lakh,
            "Total Trip Spend (₹ Cr)": spend_cr,
            "Selected-cut Projected NS": selected_ns,
            "Overall Projected NS": overall_ns,
            "Projected Revenue (₹ Cr)": revenue,
            "Revenue / Trip Spend": revenue / spend_cr if spend_cr > 0 else 0.0,
        })

    qualified_all = (
        pd.concat(all_qualified_parts, ignore_index=True)
        if all_qualified_parts else pd.DataFrame()
    )

    if not qualified_all.empty:
        qualified_all = qualified_all.drop_duplicates(
            subset=["_BV RM Key", "Market Bucket"]
        )

    total_count = (
        int(qualified_all["_BV RM Key"].nunique())
        if not qualified_all.empty else 0
    )
    total_spend_lakh = total_count * float(cost_per_employee_lakh)
    total_spend_cr = total_spend_lakh / 100.0

    total_selected_ns = float(
        pd.to_numeric(
            qualified_all[selected_amount_col], errors="coerce"
        ).fillna(0.0).sum()
    ) if not qualified_all.empty else 0.0
    total_overall_ns = float(
        pd.to_numeric(
            qualified_all["Overall Projected NS"], errors="coerce"
        ).fillna(0.0).sum()
    ) if not qualified_all.empty else 0.0
    total_revenue = _bon_voyage_projected_revenue(qualified_all)

    overall = {
        "Market": "OVERALL",
        "Qualified Employees": total_count,
        "Cost / Employee (₹ Lakh)": cost_per_employee_lakh,
        "Total Trip Spend (₹ Lakh)": total_spend_lakh,
        "Total Trip Spend (₹ Cr)": total_spend_cr,
        "Selected-cut Projected NS": total_selected_ns,
        "Overall Projected NS": total_overall_ns,
        "Projected Revenue (₹ Cr)": total_revenue,
        "Revenue / Trip Spend": (
            total_revenue / total_spend_cr
            if total_spend_cr > 0 else 0.0
        ),
    }

    return pd.DataFrame([overall, *rows]), qualified_all


def render_ai_choice_page(
    records: pd.DataFrame,
    payload: bytes,
) -> None:
    """Bright decision-support page built from the same RM run-rate projections."""
    inject_voyage_experience_css()

    st.markdown(
        """
        <div class='ai-hero'>
          <span class='eyebrow'>AI Choice · Voyage Intelligence Lab</span>
          <div class='title'>Where are the winners, the near-misses and the best trip economics?</div>
          <div class='sub'>
            A management view of the same RM Net Sales data: winner concentration by market,
            division and asset class, the near-miss conversion pool, and a live comparison
            between projected revenue and the foreign-trip budget.
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    control_cols = st.columns([1.0, 1.0, 1.2])
    with control_cols[0]:
        months = int(
            st.number_input(
                "Projection months",
                min_value=0,
                max_value=12,
                value=int(
                    st.session_state.get(
                        "ai_projection_months",
                        st.session_state.get(
                            "bv_projection_months",
                            MONTHS_REMAINING,
                        ),
                    )
                ),
                step=1,
                key="ai_projection_months",
            )
        )
    with control_cols[1]:
        threshold_pct = float(
            st.number_input(
                "Winner threshold %",
                min_value=0.0,
                value=float(
                    st.session_state.get(
                        "ai_threshold_pct",
                        100.0,
                    )
                ),
                step=5.0,
                format="%.1f",
                key="ai_threshold_pct",
            )
        )
        threshold = threshold_pct / 100.0
    with control_cols[2]:
        trip_cost_lakh = float(
            st.number_input(
                "Foreign trip cost / employee (₹ Lakh)",
                min_value=0.0,
                value=float(
                    st.session_state.get(
                        "ai_trip_cost_lakh",
                        3.0,
                    )
                ),
                step=0.25,
                format="%.2f",
                key="ai_trip_cost_lakh",
                help=(
                    "Total trip budget = qualified employees × "
                    "this editable cost."
                ),
            )
        )

    detail = build_bon_voyage_rm_detail(records, months)
    if detail.empty:
        glass_note("No RM-level Net Sales records are available.")
        return

    all_scope = _bon_voyage_collapse_all_divisions(detail)
    overall_qualified = all_scope.loc[
        pd.to_numeric(
            all_scope["Overall Projected %"], errors="coerce"
        ).fillna(0.0) >= threshold
    ].copy()

    employee_count = int(all_scope["_BV RM Key"].nunique())
    qualified_count = (
        int(overall_qualified["_BV RM Key"].nunique())
        if not overall_qualified.empty else 0
    )
    projected_ns = float(
        pd.to_numeric(
            overall_qualified["Overall Projected NS"],
            errors="coerce",
        ).fillna(0.0).sum()
    ) if not overall_qualified.empty else 0.0
    revenue = _bon_voyage_projected_revenue(overall_qualified)

    _ai_glow_cards([
        {
            "label": "Total RM universe",
            "value": fmt_num(employee_count),
            "mini": "Across the available Net Sales RM data",
        },
        {
            "label": "Overall winners",
            "value": fmt_num(qualified_count),
            "mini": f"At or above {threshold_pct:.1f}% projected achievement",
        },
        {
            "label": "Winner rate",
            "value": fmt_pct(
                qualified_count / employee_count
                if employee_count else 0.0
            ),
            "mini": "Qualified employees ÷ RM universe",
        },
        {
            "label": "Projected NS",
            "value": fmt_cr(projected_ns),
            "mini": "Overall projected Net Sales from winners",
        },
        {
            "label": "Projected revenue",
            "value": fmt_cr(revenue),
            "mini": "Equity 60 bps · Debt 20 bps · Liquid 10 bps",
        },
    ])

    market_summary = _ai_summary_by_market(
        all_scope,
        "Overall",
        threshold,
    )
    division_summary = _ai_summary_by_division(
        detail,
        threshold,
    )
    asset_summary = _ai_summary_by_asset(
        all_scope,
        threshold,
    )
    near_miss = _ai_near_miss_pool(
        all_scope,
        threshold,
    )

    # Deterministic insight cards from the data.
    best_market = (
        market_summary.sort_values(
            "Contribution", ascending=False
        ).iloc[0]
        if not market_summary.empty else None
    )
    best_division = (
        division_summary.sort_values(
            "Contribution", ascending=False
        ).iloc[0]
        if not division_summary.empty else None
    )
    near_total = int(
        near_miss["Employees"].sum()
    ) if not near_miss.empty else 0

    st.markdown(
        "<div class='ai-panel'><div class='ttl'>AI Choice · management pulse</div>"
        "<div class='sub'>Data-driven highlights from the selected threshold and projection horizon.</div></div>",
        unsafe_allow_html=True,
    )

    pulse_cards = []
    if best_market is not None:
        pulse_cards.append({
            "label": "Strongest market",
            "value": str(best_market["Market"]),
            "mini": (
                f"{fmt_pct(best_market['Contribution'])} target contribution · "
                f"{fmt_num(best_market['Qualified'])} winners"
            ),
        })
    if best_division is not None:
        pulse_cards.append({
            "label": "Strongest division",
            "value": str(best_division["Division"]),
            "mini": (
                f"{fmt_pct(best_division['Contribution'])} target contribution · "
                f"{fmt_num(best_division['Qualified'])} winners"
            ),
        })
    pulse_cards.append({
        "label": "Near-miss conversion pool",
        "value": fmt_num(near_total),
        "mini": "RMs sitting within 20 percentage points below the threshold",
    })
    _ai_glow_cards(pulse_cards)

    section_header(
        "01",
        "Market winner concentration",
        "T2 · T6 · T30 · B30 · EM",
    )
    render_glass_table(
        market_summary,
        {
            "Market": "txt",
            "Employees": "num",
            "Qualified": "num",
            "Qualification Rate": "pct",
            "Qualified Projected NS": "cr",
            "FY NS Target": "cr",
            "Contribution": "pct",
            "Projected Revenue": "cr",
        },
        css_class="ai-table",
    )

    section_header(
        "02",
        "Division winner concentration",
        "Retail · Insti · Digital · VRM · DHNI",
    )
    render_glass_table(
        division_summary,
        {
            "Division": "txt",
            "Employees": "num",
            "Qualified": "num",
            "Qualification Rate": "pct",
            "Qualified Projected NS": "cr",
            "FY NS Target": "cr",
            "Contribution": "pct",
            "Projected Revenue": "cr",
        },
        css_class="ai-table",
    )

    section_header(
        "03",
        "Asset-class winner concentration",
        "Equity · Debt · Liquid",
    )
    render_glass_table(
        asset_summary,
        {
            "Asset": "txt",
            "Employees": "num",
            "Qualified": "num",
            "Qualification Rate": "pct",
            "Qualified Projected NS": "cr",
            "FY NS Target": "cr",
            "Contribution": "pct",
            "Projected Revenue": "cr",
        },
        css_class="ai-table",
    )

    section_header(
        "04",
        "Near-miss conversion pool",
        "Employees just below the winner threshold",
    )
    render_glass_table(
        near_miss,
        {
            "Conversion Pool": "txt",
            "Employees": "num",
            "Projected NS": "cr",
            "Projected Revenue": "cr",
            "Avg Projected Achievement": "pct",
        },
        css_class="ai-table",
    )

    section_header(
        "05",
        "Foreign-trip economics",
        "Qualified employees × editable trip cost versus projected revenue",
    )

    cut_options = list(BON_VOYAGE_CUTS.keys())
    trip_cut = st.selectbox(
        "Select qualification cut for trip economics",
        cut_options,
        index=cut_options.index(
            st.session_state.get(
                "ai_trip_cut",
                st.session_state.get(
                    "bv_selected_cut",
                    "RETAIL EQUITY",
                ),
            )
            if st.session_state.get(
                "ai_trip_cut",
                st.session_state.get(
                    "bv_selected_cut",
                    "RETAIL EQUITY",
                ),
            ) in cut_options
            else "RETAIL EQUITY"
        ),
        key="ai_trip_cut",
    )

    trip_scope, trip_dimension, trip_cut_name = _bon_voyage_scope_for_cut(
        detail,
        trip_cut,
    )
    trip_table, trip_qualified = _ai_trip_rows(
        trip_scope,
        trip_dimension,
        threshold,
        trip_cost_lakh,
    )

    overall_trip = trip_table.iloc[0]
    _ai_glow_cards([
        {
            "label": "Qualified employees",
            "value": fmt_num(
                overall_trip.get("Qualified Employees")
            ),
            "mini": f"{trip_cut_name} · threshold {threshold_pct:.1f}%",
        },
        {
            "label": "Total foreign-trip spend",
            "value": (
                f"₹{_z(overall_trip.get('Total Trip Spend (₹ Lakh)')):,.1f} L"
            ),
            "mini": (
                f"{fmt_num(overall_trip.get('Qualified Employees'))} × "
                f"₹{trip_cost_lakh:,.2f} lakh"
            ),
        },
        {
            "label": "Trip spend in Cr",
            "value": fmt_cr(
                overall_trip.get("Total Trip Spend (₹ Cr)")
            ),
            "mini": "Comparable with projected revenue",
        },
        {
            "label": "Projected revenue",
            "value": fmt_cr(
                overall_trip.get("Projected Revenue (₹ Cr)")
            ),
            "mini": "Revenue from the qualified employee pool",
        },
        {
            "label": "Revenue / trip spend",
            "value": (
                f"{_z(overall_trip.get('Revenue / Trip Spend')):,.1f}×"
            ),
            "mini": "Projected revenue divided by total trip budget",
        },
    ])

    render_glass_table(
        trip_table,
        {
            "Market": "txt",
            "Qualified Employees": "num",
            "Cost / Employee (₹ Lakh)": "num",
            "Total Trip Spend (₹ Lakh)": "num",
            "Total Trip Spend (₹ Cr)": "cr",
            "Selected-cut Projected NS": "cr",
            "Overall Projected NS": "cr",
            "Projected Revenue (₹ Cr)": "cr",
            "Revenue / Trip Spend": "num",
        },
        total_rows=("OVERALL",),
        css_class="ai-trip",
    )

    glass_note(
        "Projected revenue uses the same dashboard assumptions: Equity 60 bps, "
        "Debt 20 bps and Liquid 10 bps on projected Net Sales. Trip cost is an "
        "editable planning assumption, not a workbook field."
    )


# =============================================================================
# 24. APPLICATION ENTRY POINT
# =============================================================================
 
def reset_workbook() -> None:
    for key in (
        "workbook", "segment_mapping", "channel_mapping", "application_page_selector",
        "scenario_id", "scenario_navigator", "sidebar_scenario_selector",
        "sidebar_scenario_selector_v7",
    ):
        st.session_state.pop(key, None)

    # Scenario 10 widget state depends on the current FINAL workbook.
    for key in list(st.session_state.keys()):
        if str(key).startswith(("s10_", "s10v2_", "s10v3_", "bv_", "ai_")):
            st.session_state.pop(key, None)

    rerun()
 
 
def main() -> None:
    st.set_page_config(
        page_title=APP_TITLE,
        page_icon="\u25c6",
        layout="wide",
        initial_sidebar_state="collapsed",
    )
    inject_theme()
 
    if "workbook" not in st.session_state:
        render_upload_screen()
        return
 
    payload = st.session_state["workbook"]
 
    try:
        records = load_workbook(payload)
    except WorkbookError as error:
        st.error(str(error))
        if st.button("Use another workbook"):
            reset_workbook()
        return
    except Exception:  # pragma: no cover - never surface a traceback
        st.error(
            "This workbook could not be read. Upload the standard RM scorecard containing "
            "RM Retail Sales, RM DHNI, VRM and FINAL."
        )
        if st.button("Use another workbook"):
            reset_workbook()
        return
 
    page, segment_mapping, channel_mapping = render_sidebar(records)
    records = map_business_segments(records, segment_mapping)
    records = map_business_channels(records, channel_mapping)
 
    try:
        if page == "RM performance":
            render_rm_segmentation_page(records, payload)
        elif page == "Bon voyage":
            render_bon_voyage_page(records, payload)
        elif page == "AI Choice":
            render_ai_choice_page(records, payload)
        else:
            render_command_center(records, payload)
    except WorkbookError as error:
        st.error(str(error))
    except Exception:  # pragma: no cover - never surface a traceback
        st.error(
            "This view could not be prepared from the uploaded workbook. Select another scenario "
            "or scope, or upload a workbook that matches the standard RM scorecard format."
        )
 
 
if __name__ == "__main__":
    main()
